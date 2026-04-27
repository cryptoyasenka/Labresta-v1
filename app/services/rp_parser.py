"""РП (rp.ua) Google Sheets feed parser.

The РП price list is a Google Sheets workbook in section-grouped layout
that no other supplier in the project uses, so the generic
excel_parser.parse_excel_products header-detection path cannot consume
it. Instead this module walks the sheet row-by-row, tracking the current
brand from section-header rows, and emits the same `list[dict]`
save_supplier_products() expects.

Sheet shape (verified against rp.ua sheet 1075932276, 2026-04-26):

  row 0..2   : contact / metadata rows — D is a datetime (price-list date)
  row 3      : "Курс EUR" cell (sometimes 52.00 in col D); skipped
  row 4      : section header e.g. ``B="AIRHOT (Китай)"``,
               cells A and C are None, D is 0.0 (filler).
  rows 5..N  : product rows — A is URL on rp.ua, B is product name,
               C is float EUR price, D is integer stock OR string
               "БАГАТО" (Ukrainian for "many"), or "" / 0 for out-of-stock.
  blank row  : separates sections.

Invariants honored:
  - external_id = canonicalised URL (stripped, query/fragment removed).
    URLs are the only stable identifier — there are no SKUs in the sheet.
  - currency = "EUR" (the rest of the pricing pipeline assumes EUR).
  - brand inheritance flows top-to-bottom from the most recent section
    header. A product row before any header is skipped (would have no
    brand and matcher requires brand).
  - Country-suffix in brand header (``"SIRMAN (Італія)"``) is stripped
    so brand normalisation matches MARESTO/Кодаки/НП conventions.
  - "БАГАТО" / non-zero numeric stock → available=True.
    "" / 0 / 0.0 → available=False (matches existing
    no-stock semantics).
  - Empty / unparseable price → available=False (cannot quote a 0 EUR
    offer to Horoshop), same rule as excel_parser._parse_price.
"""

from __future__ import annotations

import logging
import re
from datetime import datetime
from urllib.parse import urlparse, urlunparse

import openpyxl

logger = logging.getLogger(__name__)

# Brand-header heuristic constants (verified empirically on the rp.ua
# sheet). Keep loose so a 60-char brand label like "HELIA SMOKER (Німеччина)"
# still matches but a multi-line contact block (140+ chars) does not.
_BRAND_HEADER_MAX_LEN = 80

# "БАГАТО" is the only non-numeric stock marker we've seen.
_STOCK_MANY = "багато"

_COUNTRY_SUFFIX_RE = re.compile(r"\s*\([^)]+\)\s*$")


def _strip_country_suffix(text: str) -> str:
    """Drop a trailing ``"(Country)"`` clause from a brand-header label.

    "SIRMAN (Італія)" → "SIRMAN".  Idempotent for already-clean strings.
    """
    return _COUNTRY_SUFFIX_RE.sub("", text).strip()


def _canonical_url(url: str) -> str:
    """Strip whitespace, query and fragment from a product URL.

    The same SKU appears in the sheet with mixed protocols (``http`` vs
    ``https``) and trailing whitespace; we want a stable external_id
    across feed refreshes.
    """
    cleaned = url.strip()
    parsed = urlparse(cleaned)
    # Normalise scheme to https so http→https flips don't churn external_id.
    scheme = "https" if parsed.scheme in ("http", "https") else parsed.scheme
    return urlunparse((scheme, parsed.netloc, parsed.path, "", "", ""))


def _parse_price_cents(value) -> int | None:
    """Parse the price column into integer EUR cents.

    Accepts:
      - openpyxl-decoded number (int/float) → most rows
      - string with optional ``€`` / ``EUR`` suffix and comma decimals →
        defensive against future format drift / CSV exports

    Returns None if the value can't be coerced.
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if value <= 0:
            return None
        return int(round(float(value) * 100))
    if isinstance(value, str):
        cleaned = (
            value.replace(" ", "")
            .replace(" ", "")
            .replace("€", "")
            .replace("EUR", "")
            .replace("eur", "")
            .replace(",", ".")
            .strip()
        )
        if not cleaned:
            return None
        try:
            num = float(cleaned)
        except ValueError:
            return None
        if num <= 0:
            return None
        return int(round(num * 100))
    return None


def _parse_stock_available(value) -> bool:
    """Decide availability from the stock column.

    - "БАГАТО" (any case) → available.
    - Positive integer/float → available.
    - "" / None / 0 / 0.0 → unavailable.
    - Anything else stringy and non-empty → available (defensive: "є",
      "так", "in stock" — RP sometimes types free text).
    """
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return float(value) > 0
    if isinstance(value, str):
        text = value.strip().lower()
        if not text:
            return False
        if text in ("0", "0.0", "немає", "нет", "no", "out", "false"):
            return False
        if text == _STOCK_MANY:
            return True
        # Numeric string (e.g. "5"): treat 0 as out-of-stock.
        try:
            num = float(text.replace(",", "."))
        except ValueError:
            # Free-form non-empty text → assume in stock.
            return True
        return num > 0
    return False


def _is_brand_header(a, b, c, d) -> bool:
    """Row classification: brand-header (vs product, vs blank, vs metadata)."""
    if a is not None:
        return False
    if c is not None and not (isinstance(c, str) and not c.strip()):
        # A real brand header has C = None. A non-empty C means this is
        # a product row whose A cell happens to be missing (corrupt feed).
        return False
    if isinstance(d, datetime):
        # row 0 — contact block carries a datetime in D.
        return False
    if not isinstance(b, str):
        return False
    text = b.strip()
    if not text:
        return False
    if len(text) > _BRAND_HEADER_MAX_LEN:
        return False
    # The contact-info row has multiple lines and commas; brand headers
    # are single-line short labels, possibly with "(Country)".
    if "\n" in text:
        return False
    return True


def parse_rp_sheet(file_path: str, supplier_id: int) -> tuple[list[dict], list[str]]:
    """Parse a downloaded РП xlsx into product dicts + warnings.

    Args:
        file_path: Path to the .xlsx downloaded from Google Sheets export.
        supplier_id: Supplier row id to attach.

    Returns:
        ``(products, errors)`` — products compatible with
        ``save_supplier_products()``; errors is a list of human-readable
        warnings (duplicate URLs, brandless products, unparseable price).
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    # The sheet currently lives in a tab named "ЗАЛИШКИ" but we don't
    # rely on the name — just use the first (or only) sheet. If the
    # operator splits the workbook in the future, sync_pipeline can pass
    # an explicit sheet name as a follow-up.
    ws = wb[wb.sheetnames[0]]

    products: list[dict] = []
    errors: list[str] = []
    seen_ext_ids: set[str] = set()

    current_brand: str | None = None

    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        # Pad row to 4 cells so destructuring is safe even when openpyxl
        # trims trailing Nones.
        a = row[0] if len(row) > 0 else None
        b = row[1] if len(row) > 1 else None
        c = row[2] if len(row) > 2 else None
        d = row[3] if len(row) > 3 else None

        # Skip fully blank rows.
        if (a is None or (isinstance(a, str) and not a.strip())) \
                and (b is None or (isinstance(b, str) and not b.strip())) \
                and (c is None or (isinstance(c, str) and not c.strip())) \
                and (d is None or (isinstance(d, str) and not d.strip())):
            continue

        # Section header → update current_brand and move on.
        if _is_brand_header(a, b, c, d):
            current_brand = _strip_country_suffix(str(b).strip())
            continue

        # Skip header / metadata rows: contact line (datetime in D), or
        # rows without a URL in A.
        if not isinstance(a, str) or not a.strip().startswith("http"):
            continue

        url = a.strip()
        ext_id = _canonical_url(url)
        if not ext_id:
            errors.append(f"Row {row_idx + 1}: empty canonical URL ({url!r})")
            continue

        name = ""
        if isinstance(b, str):
            name = b.strip()
        if not name:
            errors.append(f"Row {row_idx + 1}: missing name for {url}")
            continue

        if not current_brand:
            errors.append(
                f"Row {row_idx + 1}: product {url!r} has no preceding "
                f"brand-header section, skipping"
            )
            continue

        if ext_id in seen_ext_ids:
            errors.append(f"Row {row_idx + 1}: duplicate URL {ext_id!r}")
            continue
        seen_ext_ids.add(ext_id)

        price_cents = _parse_price_cents(c)
        if price_cents is None:
            errors.append(
                f"Row {row_idx + 1}: unparseable / zero price {c!r} for {url}"
            )

        available = _parse_stock_available(d)
        # Cannot quote a row with no price even if stock says "БАГАТО".
        if price_cents is None:
            available = False

        products.append({
            "external_id": ext_id,
            "name": name,
            "brand": current_brand,
            "model": None,
            "article": None,
            "price_cents": price_cents,
            "currency": "EUR",
            "available": available,
            "supplier_id": supplier_id,
        })

    wb.close()

    logger.info(
        "RP parser: %d products parsed (%d available), %d warnings",
        len(products),
        sum(1 for p in products if p["available"]),
        len(errors),
    )
    return products, errors
