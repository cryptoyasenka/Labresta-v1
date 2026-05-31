"""Recover the Horoshop category corpus from a downloaded store export.

`catalog_import.py` reads the Horoshop export to upsert PromProduct rows, but
its COLUMN_ALIASES has NO alias for «Раздел», so the category column is read and
then discarded (RESEARCH Q6). The smart category resolver (plan 09-02) needs that
column as a corpus: the set of real categories the store already uses, plus a
per-card {brand, name → category} table the analogy tier ranks against.

Rather than migrate the PromProduct schema to persist «Раздел», this reads it
straight off the on-disk export at generate-time — read-only, no DB.

The CANONICAL export is `horoshop-export 26.05.26.xlsx` (sheet "Sheet1", 59
cols). `horoshop-export-extended.xlsx` and the "Кодаки …категории.xlsx" files
are CORRUPT (no sharedStrings — RESEARCH Q1/Q4/Pitfall 9) and are NEVER read by
this code; tests build a tiny valid export in-memory.

Columns are located by HEADER LABEL (mirroring np_parser's header-drift guard):
fixed positions can't be trusted across a store-side export change, so a missing
«Раздел» or «Артикул» header ABORTS with one error rather than mis-reading.
"""

from __future__ import annotations

import logging

import openpyxl

logger = logging.getLogger(__name__)

# Header LABELS in the canonical Horoshop dealer export (verified 2026-05-31).
# Located by exact case-insensitive match — NOT by fixed index.
_LBL_ARTICLE = "Артикул"
_LBL_DISPLAY_ARTICLE = "Артикул для отображения на сайте"
_LBL_NAME_UA = "Название (UA)"
_LBL_BRAND = "Бренд"
_LBL_CATEGORY = "Раздел"

_EXPORT_SHEET = "Sheet1"


def _clean(value) -> str | None:
    """Stripped string, or None for empty/None."""
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _find_col(header, label: str) -> int | None:
    """0-based index of the first header cell whose label == ``label`` (ci)."""
    target = label.strip().lower()
    for idx, cell in enumerate(header):
        if isinstance(cell, str) and cell.strip().lower() == target:
            return idx
    return None


def read_category_corpus(export_path: str) -> tuple[list[dict], list[str]]:
    """Read a Horoshop export into a category corpus.

    Args:
        export_path: local path to a downloaded Horoshop dealer-export .xlsx.

    Returns:
        ``(rows, errors)`` where each row is
        ``{"external_id", "display_article", "name", "brand", "category"}``
        (category == «Раздел»). Rows with an empty category are skipped (no
        category signal). De-dup is not done — multiple cards per category is
        the norm and the analogy tier wants every example. On a missing
        «Раздел»/«Артикул» header the rows are empty and errors holds one abort
        message (fixed positions can't be trusted — mirror np_parser).
    """
    wb = openpyxl.load_workbook(export_path, read_only=True, data_only=True)
    ws = wb[_EXPORT_SHEET] if _EXPORT_SHEET in wb.sheetnames else wb[wb.sheetnames[0]]

    rows: list[dict] = []
    errors: list[str] = []

    row_iter = ws.iter_rows(values_only=True)
    try:
        header = next(row_iter)
    except StopIteration:
        wb.close()
        return [], ["Horoshop export is empty (no header row)"]

    col_article = _find_col(header, _LBL_ARTICLE)
    col_category = _find_col(header, _LBL_CATEGORY)
    if col_article is None or col_category is None:
        wb.close()
        missing = []
        if col_article is None:
            missing.append(_LBL_ARTICLE)
        if col_category is None:
            missing.append(_LBL_CATEGORY)
        return [], [
            f"Horoshop export missing required header(s): {', '.join(missing)}. "
            f"Aborting — fixed column positions can't be trusted."
        ]

    col_display = _find_col(header, _LBL_DISPLAY_ARTICLE)
    col_name = _find_col(header, _LBL_NAME_UA)
    col_brand = _find_col(header, _LBL_BRAND)

    def _cell(row, idx):
        if idx is None:
            return None
        return row[idx] if len(row) > idx else None

    skipped_no_category = 0
    for row in row_iter:
        category = _clean(_cell(row, col_category))
        if not category:
            skipped_no_category += 1
            continue
        rows.append({
            "external_id": _clean(_cell(row, col_article)),
            "display_article": _clean(_cell(row, col_display)),
            "name": _clean(_cell(row, col_name)),
            "brand": _clean(_cell(row, col_brand)),
            "category": category,
        })

    wb.close()

    logger.info(
        "Category export: %d rows with a category (%d skipped no-category), "
        "%d distinct categories",
        len(rows), skipped_no_category,
        len({r["category"] for r in rows}),
    )
    return rows, errors
