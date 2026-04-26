"""Excel/Google Sheets parser service for supplier product feeds.

Detects Google Sheets URLs, converts to download URLs, auto-detects
column mappings by keyword matching (ukr/rus/eng), and parses products
into the same list[dict] format used by save_supplier_products().
"""

import logging
import re
from urllib.parse import parse_qsl, urlparse

import openpyxl

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

GOOGLE_SHEETS_PATTERN = re.compile(
    r"docs\.google\.com/spreadsheets/d/([a-zA-Z0-9_-]+)"
)

COLUMN_KEYWORDS = {
    # "title" covers Horoshop-style dealer exports (title_uk / title_ru).
    # "наименование/найменування" covers common 1C and distributor templates.
    "name": ["назва", "название", "наименование", "найменування", "title", "name"],
    "price": ["ціна", "цена", "price"],
    "available": ["наявність", "наличие", "available", "в наявності", "в наличии"],
    # "brend" (Latin-only transliteration) is a frequent real-world header —
    # e.g. NP dealer export uses attr_brend_uk / attr_brend_ru.
    "brand": ["бренд", "brand", "brend", "виробник", "производитель"],
    "model": ["модель", "model"],
    # SKU/vendor-code column. When present, it becomes the stable external_id
    # (takes priority over brand+model) — see parse_excel_products.
    "article": ["артикул", "sku", "vendor", "vendorcode", "article"],
}

REQUIRED_FIELDS = {"name", "brand", "model", "price"}


# ---------------------------------------------------------------------------
# URL helpers
# ---------------------------------------------------------------------------


def is_google_sheets_url(url: str) -> bool:
    """Check whether *url* is a Google Sheets sharing/editing URL."""
    if not url:
        return False
    return bool(GOOGLE_SHEETS_PATTERN.search(url))


def is_xlsx_url(url: str) -> bool:
    """Heuristic: URL likely serves an .xlsx file (non-Google).

    Covers two signals commonly seen with dealer/b2b export endpoints:
      - path ends with ``.xlsx`` (with or without query string)
      - query param ``filetype=xlsx`` is present (e.g. np.com.ua dealer-export)
    """
    if not url:
        return False
    parsed = urlparse(url)
    if parsed.path.lower().endswith(".xlsx"):
        return True
    for key, value in parse_qsl(parsed.query, keep_blank_values=True):
        if key.lower() == "filetype" and value.lower() == "xlsx":
            return True
    return False


def convert_google_sheets_url(url: str) -> str:
    """Convert a Google Sheets sharing URL to an .xlsx download URL.

    Extracts the spreadsheet ID and optional gid parameter, then builds
    ``/export?format=xlsx[&gid=GID]``.

    Raises:
        ValueError: If *url* is not a recognised Google Sheets URL.
    """
    match = GOOGLE_SHEETS_PATTERN.search(url)
    if not match:
        raise ValueError(f"Not a Google Sheets URL: {url}")

    spreadsheet_id = match.group(1)
    gid_match = re.search(r"gid=(\d+)", url)
    gid_param = f"&gid={gid_match.group(1)}" if gid_match else ""

    return (
        f"https://docs.google.com/spreadsheets/d/"
        f"{spreadsheet_id}/export?format=xlsx{gid_param}"
    )


# ---------------------------------------------------------------------------
# XLSX validation
# ---------------------------------------------------------------------------


def validate_xlsx_response(raw_bytes: bytes) -> None:
    """Validate that *raw_bytes* look like a genuine .xlsx (ZIP) file.

    Raises:
        ValueError: If bytes are HTML (not-shared sheet) or otherwise invalid.
    """
    if raw_bytes[:2] == b"PK":
        return  # Valid ZIP / .xlsx header

    if raw_bytes[:1] == b"<" or b"<!DOCTYPE" in raw_bytes[:200]:
        raise ValueError(
            "Google Sheet is not shared publicly. "
            "Got HTML instead of .xlsx. "
            "Please set sharing to 'Anyone with the link can view'."
        )

    raise ValueError(
        "Downloaded file is not a valid .xlsx (expected ZIP/PK header)"
    )


# ---------------------------------------------------------------------------
# Column detection
# ---------------------------------------------------------------------------


def detect_columns(ws, max_scan_rows: int = 10) -> tuple[int | None, dict]:
    """Scan the first *max_scan_rows* rows for keyword-matching headers.

    A row is considered a header when it contains the ``name`` field **and**
    at least one other recognised field.

    Returns:
        ``(header_row_index, {col_index: field_name})`` or ``(None, {})``.
    """
    for row_idx, row in enumerate(
        ws.iter_rows(max_row=max_scan_rows, values_only=True)
    ):
        mapping: dict[int, str] = {}
        for col_idx, cell_value in enumerate(row):
            if cell_value is None:
                continue
            cell_lower = str(cell_value).strip().lower()
            for field, keywords in COLUMN_KEYWORDS.items():
                if any(kw in cell_lower for kw in keywords):
                    # Only assign first column that matches a field
                    if field not in mapping.values():
                        mapping[col_idx] = field
                        break

        # Header = "name" + at least one more recognised field
        if "name" in mapping.values() and len(mapping) >= 2:
            return row_idx, mapping

    return None, {}


# ---------------------------------------------------------------------------
# Price parsing helper
# ---------------------------------------------------------------------------


def _parse_price(price_str: str) -> int | None:
    """Parse a price string into integer cents.

    Handles comma-as-decimal and space-as-thousands separator.
    Returns ``None`` if unparseable.
    """
    cleaned = price_str.replace("\u00a0", "").replace(" ", "")
    # Comma as decimal separator (European style)
    cleaned = cleaned.replace(",", ".")
    try:
        return int(float(cleaned) * 100)
    except (ValueError, TypeError):
        return None


# ---------------------------------------------------------------------------
# Product parsing
# ---------------------------------------------------------------------------


def parse_excel_products(
    file_path: str,
    column_mapping: dict,
    header_row: int,
    supplier_id: int,
) -> tuple[list[dict], list[str]]:
    """Parse an Excel file using a confirmed column mapping.

    Args:
        file_path: Path to the .xlsx file.
        column_mapping: ``{col_index: field_name}`` (keys may be int or str).
        header_row: Zero-based index of the header row.
        supplier_id: Supplier ID attached to every product dict.

    Returns:
        ``(products, errors)`` — *products* is compatible with
        ``save_supplier_products()``, *errors* is a list of human-readable
        error strings.
    """
    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb[wb.sheetnames[0]]

    # Normalise column_mapping keys to int
    col_map: dict[int, str] = {int(k): v for k, v in column_mapping.items()}

    products: list[dict] = []
    errors: list[str] = []
    seen_ids: set[str] = set()

    has_available_col = "available" in col_map.values()

    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if row_idx <= header_row:
            continue  # Skip header and rows above it

        # Extract values by mapping
        values: dict[str, str] = {}
        for col_idx, field in col_map.items():
            if col_idx < len(row):
                val = row[col_idx]
                values[field] = str(val).strip() if val is not None else ""
            else:
                values[field] = ""

        brand = values.get("brand", "").strip()
        model_val = values.get("model", "").strip()
        article = values.get("article", "").strip()

        name = values.get("name", "").strip()
        if not name:
            continue

        # Stable external_id priority: article (SKU) → brand+model → name.
        # SKUs are the most stable identifier across feed refreshes; fall back
        # only when the feed doesn't carry one.
        if article:
            ext_id = article.lower()
            dup_label = f"article '{article}'"
        elif brand and model_val:
            ext_id = f"{brand.lower()}|{model_val.lower()}"
            dup_label = f"brand+model '{brand} {model_val}'"
        else:
            ext_id = name.lower()
            dup_label = f"name '{name}'"

        # Skip duplicates
        if ext_id in seen_ids:
            errors.append(f"Row {row_idx + 1}: duplicate {dup_label}")
            continue
        seen_ids.add(ext_id)

        # Parse price
        price_str = values.get("price", "").strip()
        price_cents: int | None = None
        price_error = False

        if price_str:
            price_cents = _parse_price(price_str)
            if price_cents is None:
                errors.append(
                    f"Row {row_idx + 1}: unparseable price '{price_str}'"
                )
                price_error = True
        else:
            # Empty price: keep the row but surface it as a warning so the
            # operator doesn't wonder why 9 of 686 products came in at 0 EUR.
            errors.append(f"Row {row_idx + 1}: empty price")

        # Parse availability
        if has_available_col:
            avail_raw = values.get("available", "").strip().lower()
            available = avail_raw in (
                "",
                "да",
                "так",
                "yes",
                "true",
                "+",
                "1",
                "в наявності",
                "в наличии",
            )
        else:
            available = True

        # Unparseable price => mark unavailable
        if price_error:
            available = False
        # Missing price => also unavailable, even if the feed says "в наличии"
        # (we can't quote a 0-EUR offer to Horoshop).
        if price_cents is None:
            available = False

        products.append(
            {
                "external_id": ext_id,
                "name": name,
                "brand": brand,
                "model": model_val,
                "article": article or None,
                "price_cents": price_cents,
                # LabResta pipeline treats supplier prices as EUR — MARESTO YML
                # uses <currencyId>EUR</currencyId>, calculate_price_eur and
                # calculate_auto_discount assume EUR cents. Horoshop-style xlsx
                # dealer exports likewise quote in EUR. Hard-coded to EUR is the
                # pragmatic choice until a real UAH supplier appears.
                "currency": "EUR",
                "available": available,
                "supplier_id": supplier_id,
            }
        )

    wb.close()
    return products, errors


# ---------------------------------------------------------------------------
# Preview data generation
# ---------------------------------------------------------------------------


def get_preview_data(file_path: str, max_rows: int = 10) -> dict:
    """Generate preview data for the column-mapping confirmation page.

    Returns a dict with keys:
        - ``header_row``: detected header row index (or ``None``).
        - ``detected_mapping``: ``{col_idx: field_name}``.
        - ``all_headers``: list of header cell strings.
        - ``preview_rows``: list of row-lists (up to *max_rows* data rows).
    """
    # First pass: detect columns
    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row, detected = detect_columns(ws)
    wb.close()

    # Second pass: collect header + preview rows (iterator consumed above)
    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb[wb.sheetnames[0]]

    all_rows: list[list[str]] = []
    limit = (header_row if header_row is not None else 0) + max_rows + 1
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        all_rows.append([str(c) if c is not None else "" for c in row])
        if i >= limit:
            break
    wb.close()

    if header_row is not None and header_row < len(all_rows):
        headers = all_rows[header_row]
        data_start = header_row + 1
        data_rows = all_rows[data_start : data_start + max_rows]
    elif all_rows:
        # No header detected — show first row as column labels
        # and remaining rows as preview so user can map manually
        max_cols = max(len(r) for r in all_rows) if all_rows else 0
        headers = [f"Колонка {i + 1}" for i in range(max_cols)]
        data_rows = all_rows[:max_rows]
    else:
        headers = []
        data_rows = []

    return {
        "header_row": header_row,
        "detected_mapping": detected,
        "all_headers": headers,
        "preview_rows": data_rows,
    }
