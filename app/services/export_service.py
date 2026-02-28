"""Export service: CSV and XLSX generation for match data."""

import csv
import io

import openpyxl


def _match_row(match):
    """Extract a row dict from a ProductMatch object."""
    sp = match.supplier_product
    pp = match.prom_product

    supplier_price = ""
    if sp and sp.price_cents is not None:
        supplier_price = f"{sp.price_cents / 100:.2f}"

    prom_price = ""
    if pp and pp.price is not None:
        prom_price = f"{pp.price / 100:.2f}"

    confirmed_at = ""
    if match.confirmed_at:
        confirmed_at = match.confirmed_at.strftime("%d.%m.%Y %H:%M")

    return {
        "supplier_name": sp.name if sp else "",
        "prom_name": pp.name if pp else "",
        "brand": (sp.brand if sp and sp.brand else ""),
        "score": f"{match.score:.0f}" if match.score is not None else "",
        "status": match.status or "",
        "supplier_price": supplier_price,
        "prom_price": prom_price,
        "confirmed_by": match.confirmed_by or "",
        "confirmed_at": confirmed_at,
    }


COLUMNS = [
    ("supplier_name", "Товар поставщика"),
    ("prom_name", "Товар prom.ua"),
    ("brand", "Бренд"),
    ("score", "Скор матчинга"),
    ("status", "Статус"),
    ("supplier_price", "Цена поставщика (EUR)"),
    ("prom_price", "Расч. цена prom.ua (EUR)"),
    ("confirmed_by", "Подтвердил"),
    ("confirmed_at", "Дата подтверждения"),
]


def export_matches_csv(matches):
    """Export matches to CSV with UTF-8 BOM encoding for Excel compatibility.

    Args:
        matches: iterable of ProductMatch objects with eager-loaded relationships.

    Returns:
        StringIO with CSV content (utf-8-sig encoded).
    """
    output = io.StringIO()
    # Write BOM for Excel
    output.write("\ufeff")

    writer = csv.writer(output)
    # Header row
    writer.writerow([col[1] for col in COLUMNS])

    for match in matches:
        row = _match_row(match)
        writer.writerow([row[col[0]] for col in COLUMNS])

    output.seek(0)
    return output


def export_matches_xlsx(matches):
    """Export matches to XLSX with openpyxl.

    Args:
        matches: iterable of ProductMatch objects with eager-loaded relationships.

    Returns:
        BytesIO with workbook content.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Matches"

    # Header row
    headers = [col[1] for col in COLUMNS]
    ws.append(headers)

    # Bold header style
    from openpyxl.styles import Font

    bold_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold_font

    # Data rows
    for match in matches:
        row = _match_row(match)
        ws.append([row[col[0]] for col in COLUMNS])

    # Auto-fit column widths (approximate based on header + content)
    for col_idx, (key, header) in enumerate(COLUMNS, 1):
        max_length = len(header)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
