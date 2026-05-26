"""Build the НП Channel-2 catalog import file (native Horoshop xlsx).

Generalizes the proven 1-product canary (build_canary_xlsx.py) to EVERY matched
NP-exclusive SKU, producing one native [КАТАЛОГ]-schema workbook that Horoshop
auto-maps on import (FINAL-MODEL.md §1, §3). Yana imports it by hand
(invariant #13) — this module only builds the rows/workbook, never touches the
store.

For each in-scope match (supplier_id=2, confirmed/manual, published=1, brand in
the 9 NP-exclusive brands):
  - content (name UA/RU, body UA/RU, brand, photo gallery) comes from the NP
    feed, looked up by Артикул via parse_np_feed();
  - Артикул column := pp.external_id (Horoshop matches the card by Артикул);
  - price / availability come from the SAME helpers the live YML feed uses
    (yml_generator._compute_price_eur / _is_available_for_offer) so the Excel
    price equals the price the operator approved on screen — no pricing mirror.

Output headers carry Horoshop's exact catalog field names so columns auto-map;
the photo gallery lands in «[КАТАЛОГ] Галерея» (Yana confirms that one column in
the import preview — see FINAL-MODEL.md §3).

Read-only against the DB. Both the script (scripts/build_np_catalog_xlsx.py) and
the operator download endpoint (views/feed.py) call into here, so the view never
imports from scripts/ — the dependency direction stays app ← entry points.
"""

import json
from pathlib import Path

import openpyxl

from app.models.product_match import ProductMatch
from app.models.supplier_product import SupplierProduct
from app.services.np_parser import parse_np_feed
from app.services.yml_generator import (
    _compute_price_eur,
    _is_available_for_offer,
    is_valid_price,
)

NP_SUPPLIER_ID = 2
DEFAULT_FEED = str(
    Path(__file__).resolve().parent.parent.parent
    / ".planning" / "plans" / "np-feed" / "np-feed.xlsx"
)
SCOPE_BRANDS = {
    "HURAKAN", "APACH", "FAGOR", "TATRA", "COLD",
    "PROJECT SYSTEMS", "ASTORIA", "ARRIS", "MAXIMA",
}

# Horoshop catalog field names (exact, from the live-proven canary). The "[КАТАЛОГ] "
# prefix is taken from the feed header at runtime so we never hand-type cyrillic.
_LEAF = {
    "id": "id",
    "article": "Артикул",
    "price": "Цена",
    "gallery": "Галерея",
    "available": "Наличие",
    "show": "Отображать",
    "name_uk": "Название модификации (UA)",
    "desc_uk": "Описание товара (UA)",
    "brand": "Бренд",
    "name_ru": "Название модификации (RU)",
    "desc_ru": "Описание товара (RU)",
    "oldprice": "Старая цена",
    "currency": "Валюта",
}


def _detect_prefix(feed_path: str) -> str:
    """Extract the "[КАТАЛОГ] " prefix from the feed's price header."""
    wb = openpyxl.load_workbook(feed_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = list(next(ws.iter_rows(values_only=True)))
    wb.close()
    for h in header:
        s = str(h or "")
        if s.startswith("[") and s.rstrip().endswith("Цена"):
            # "[КАТАЛОГ] Цена" → "[КАТАЛОГ] "
            return s[: s.rfind("Цена")]
    return "[КАТАЛОГ] "


def _headers(prefix: str) -> list[str]:
    p = prefix
    # Top-level fields (id, Артикул) carry no prefix; catalog leaves do.
    return [
        _LEAF["id"],
        _LEAF["article"],
        p + _LEAF["price"],
        p + _LEAF["gallery"],
        p + _LEAF["available"],
        p + _LEAF["show"],
        p + _LEAF["name_uk"],
        p + _LEAF["desc_uk"],
        p + _LEAF["brand"],
        p + _LEAF["name_ru"],
        p + _LEAF["desc_ru"],
        p + _LEAF["oldprice"],
        p + _LEAF["currency"],
    ]


def _gallery_str(images_json: str | None, image_url: str | None) -> str:
    """Join the photo gallery as `url;url`, main first, deduplicated."""
    urls: list[str] = []
    seen: set[str] = set()
    if image_url:
        urls.append(image_url)
        seen.add(image_url)
    if images_json:
        try:
            for u in json.loads(images_json):
                if u and u not in seen:
                    seen.add(u)
                    urls.append(u)
        except (ValueError, TypeError):
            pass
    return ";".join(urls)


def _in_scope(brand: str | None) -> bool:
    return bool(brand) and brand.strip().upper() in SCOPE_BRANDS


def build_catalog_rows(feed_path: str = DEFAULT_FEED,
                       supplier_id: int = NP_SUPPLIER_ID):
    """Build (headers, rows, errors) for the NP catalog import file.

    Must run inside a Flask app context. Reads the feed + DB; no writes.
    Each row is a list aligned to ``headers``.
    """
    feed_rows, feed_errors = parse_np_feed(feed_path, supplier_id)
    by_article = {r["article"]: r for r in feed_rows}

    errors = list(feed_errors)
    prefix = _detect_prefix(feed_path)
    headers = _headers(prefix)

    matches = (
        ProductMatch.query
        .join(SupplierProduct,
              ProductMatch.supplier_product_id == SupplierProduct.id)
        .filter(
            SupplierProduct.supplier_id == supplier_id,
            ProductMatch.status.in_(["confirmed", "manual"]),
            ProductMatch.published.is_(True),
        )
        .all()
    )

    rows: list[list] = []
    for m in matches:
        sp = m.supplier_product
        pp = m.prom_product
        if not sp or not pp or not _in_scope(sp.brand):
            continue
        feed = by_article.get((sp.article or "").strip())
        if feed is None:
            errors.append(
                f"Артикул {sp.article!r} (pp {pp.external_id}) not in feed — skipped"
            )
            continue

        available = _is_available_for_offer(m)
        price_eur = _compute_price_eur(m) if is_valid_price(sp.price_cents) else 0.0
        retail_eur = (sp.price_cents or 0) / 100.0
        currency = sp.currency if sp.currency in ("EUR", "UAH") else "EUR"

        rows.append([
            "",                                              # id (match by Артикул)
            str(pp.external_id),                             # Артикул = our card
            f"{price_eur:.1f}",                              # Цена (our discount)
            _gallery_str(feed.get("images"), feed.get("image_url")),  # Галерея
            "В наличии" if available else "Нет в наличии",   # Наличие
            "1",                                             # Отображать
            feed.get("name") or "",                          # Название UA
            feed.get("description") or "",                   # Описание UA
            feed.get("brand") or "",                         # Бренд
            feed.get("name_ru") or "",                       # Название RU
            feed.get("description_ru") or "",                # Описание RU
            f"{retail_eur:.1f}",                             # Старая цена (retail)
            currency,                                        # Валюта
        ])

    return headers, rows, errors


def build_workbook(headers: list[str], rows: list[list]) -> "openpyxl.Workbook":
    """Build (but do not save) the catalog workbook. Caller saves to path or stream."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Worksheet"
    ws.append(headers)
    for r in rows:
        ws.append(r)
    return wb


def write_workbook(headers: list[str], rows: list[list], out_path: str) -> str:
    wb = build_workbook(headers, rows)
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    wb.close()
    return out_path
