"""
Catalog CSV/XLS/XLSX import service.

Parses export files from Horoshop or prom.ua admin panels,
normalizes headers, and upserts products into the PromProduct table.
"""

import csv
import io
from datetime import datetime, timezone

import chardet
import openpyxl

from app.extensions import db
from app.models.catalog import PromProduct

# Column alias mapping: export headers -> internal field names
# Supports both Horoshop and prom.ua formats
COLUMN_ALIASES = {
    # --- Horoshop export headers ---
    "артикул": "external_id",
    "артикул для отображения на сайте": "display_article",
    "артикул для відображення на сайті": "display_article",
    "назва (ua)": "name",
    "название (ua)": "name",
    "назва модифікації (ua)": "name",
    "название модификации (ua)": "name",
    "назва (ru)": "name_ru",
    "название (ru)": "name_ru",
    "бренд": "brand",
    "ціна": "price",
    "цена": "price",
    "валюта": "currency",
    "посилання": "page_url",
    "ссылка": "page_url",
    "фото": "image_url",
    "галерея": "images",
    "описание товара (ua)": "description_ua",
    "опис товару (ua)": "description_ua",
    "описание товара (ru)": "description_ru",
    "опис товару (ru)": "description_ru",
    # --- Prom.ua export headers ---
    "унікальний_ідентифікатор": "external_id",
    "уникальный_идентификатор": "external_id",
    "ідентифікатор_товару": "external_id",
    "идентификатор_товара": "external_id",
    "назва_позиції": "name",
    "название_позиции": "name",
    "код_товару": "article",
    "код_товара": "article",
    "виробник": "brand",
    "производитель": "brand",
    "посилання_на_сторінку_товару": "page_url",
    "ссылка_на_страницу_товара": "page_url",
}


def normalize_header(header: str) -> str:
    """Strip whitespace and lowercase a header string."""
    return header.strip().lower()


def map_headers(raw_headers: list[str]) -> dict[int, str]:
    """
    Map column indices to internal field names via COLUMN_ALIASES.

    Args:
        raw_headers: List of raw header strings from the first row of the file.

    Returns:
        Dict mapping column index -> internal field name for recognized columns.

    Raises:
        ValueError: If 'external_id' or 'name' column is not found.
    """
    mapping = {}
    mapped_fields = set()
    for idx, header in enumerate(raw_headers):
        normalized = normalize_header(header)
        if normalized in COLUMN_ALIASES:
            field = COLUMN_ALIASES[normalized]
            # First column wins — skip duplicates for the same field
            if field not in mapped_fields:
                mapping[idx] = field
                mapped_fields.add(field)

    # Validate required columns
    missing = []
    if "external_id" not in mapped_fields:
        missing.append("external_id")
    if "name" not in mapped_fields:
        missing.append("name")

    if missing:
        recognized = [f"{idx}: {field}" for idx, field in mapping.items()]
        unrecognized = [
            raw_headers[i] for i in range(len(raw_headers)) if i not in mapping
        ]
        raise ValueError(
            f"Missing required columns: {missing}. "
            f"Recognized: {recognized}. "
            f"Unrecognized headers: {unrecognized}"
        )

    return mapping


def parse_csv(file_stream: io.TextIOBase, encoding: str = "utf-8") -> list[dict]:
    """
    Parse a CSV file stream into a list of product dicts.

    Args:
        file_stream: Text-mode file stream.
        encoding: Not used directly (stream should already be decoded), kept for API consistency.

    Returns:
        List of dicts with mapped field names.
    """
    reader = csv.reader(file_stream)

    # First row = headers
    try:
        raw_headers = next(reader)
    except StopIteration:
        raise ValueError("CSV file is empty")

    mapping = map_headers(raw_headers)

    products = []
    for row in reader:
        if not row or all(cell.strip() == "" for cell in row):
            continue  # skip empty rows
        product = {}
        for idx, field_name in mapping.items():
            if idx < len(row):
                product[field_name] = row[idx].strip()
            else:
                product[field_name] = ""
        products.append(product)

    return products


def parse_xlsx(file_path: str) -> list[dict]:
    """
    Parse an XLSX file into a list of product dicts.

    Args:
        file_path: Path to the .xlsx file on disk.

    Returns:
        List of dicts with mapped field names.
    """
    wb = openpyxl.load_workbook(file_path, read_only=True)

    # Try known sheet name first, fall back to first sheet
    if "Export Products Sheet" in wb.sheetnames:
        ws = wb["Export Products Sheet"]
    else:
        ws = wb[wb.sheetnames[0]]

    rows = ws.iter_rows(values_only=True)

    # First row = headers
    try:
        raw_headers = [str(h) if h is not None else "" for h in next(rows)]
    except StopIteration:
        wb.close()
        raise ValueError("XLSX file is empty")

    mapping = map_headers(raw_headers)

    products = []
    for row in rows:
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue  # skip empty rows
        product = {}
        for idx, field_name in mapping.items():
            if idx < len(row):
                val = row[idx]
                product[field_name] = str(val).strip() if val is not None else ""
            else:
                product[field_name] = ""
        products.append(product)

    wb.close()
    return products


def parse_catalog_file(file_path: str, filename: str) -> list[dict]:
    """
    Dispatch to the correct parser based on file extension.

    Args:
        file_path: Path to the uploaded file on disk.
        filename: Original filename (used to detect extension).

    Returns:
        List of product dicts.

    Raises:
        ValueError: For unsupported file extensions or parse errors.
    """
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext == "csv":
        # Detect encoding
        with open(file_path, "rb") as f:
            raw_data = f.read()
        detected = chardet.detect(raw_data)
        encoding = detected.get("encoding") or "utf-8"

        # Try detected encoding, fall back to cp1251
        try:
            text = raw_data.decode(encoding)
        except (UnicodeDecodeError, LookupError):
            text = raw_data.decode("cp1251")

        return parse_csv(io.StringIO(text), encoding=encoding)

    elif ext in ("xls", "xlsx"):
        return parse_xlsx(file_path)

    else:
        raise ValueError(
            f"Unsupported file format: .{ext}. Supported formats: .csv, .xls, .xlsx"
        )


# Catalog-owned fields: a Horoshop export is authoritative — overwritten on update.
CATALOG_FIELDS = [
    "name",
    "brand",
    "article",
    "display_article",
    "price",
    "currency",
    "page_url",
    "image_url",
    "images",
    "description_ua",
]
# Worker-owned fields: filled by the translation terminals, protected on update.
TRANSLATION_FIELDS = ["name_ru", "description_ru"]


def _normalize_product(product: dict) -> dict | None:
    """Normalize one parsed row into typed field values.

    Returns a dict keyed by internal field name (external_id, name, price as
    int cents, the rest as str-or-None), or None if the row is missing a
    required field (external_id / name) and must be skipped.
    """
    ext_id = product.get("external_id", "").strip()
    name = product.get("name", "").strip()
    if not ext_id or not name:
        return None

    # Parse price: string -> float -> cents (int)
    price = None
    raw_price = product.get("price", "").strip()
    if raw_price:
        try:
            price = int(round(float(raw_price) * 100))
        except (ValueError, TypeError):
            price = None

    def _opt(field: str) -> str | None:
        return product.get(field, "").strip() or None

    return {
        "external_id": ext_id,
        "name": name,
        "price": price,
        "currency": _opt("currency"),
        "article": _opt("article"),
        "display_article": _opt("display_article"),
        "brand": _opt("brand"),
        "page_url": _opt("page_url"),
        "name_ru": _opt("name_ru"),
        "image_url": _opt("image_url"),
        "images": _opt("images"),
        "description_ua": _opt("description_ua"),
        "description_ru": _opt("description_ru"),
    }


def preview_catalog_import(
    products: list[dict], preserve_translations: bool = True
) -> dict:
    """Compute what a catalog import WOULD do, without writing anything.

    Read-only counterpart of save_catalog_products: it walks the same fields
    via the same _normalize_product helper and compares each row to the current
    DB state, so the preview can never drift from what the real import does.

    Returns a dict:
      * created / updated / skipped / total — same shape as save result.
      * changed: {field: n} — among existing products, how many would have this
        CATALOG-owned field's value actually change.
      * cleared: {field: n} — among existing products, how many would have a
        non-empty value REPLACED BY EMPTY (the file lacks/blanks that column).
        This is the key "wrong file" signal — e.g. price cleared on 5000 rows.
      * translations_protected — existing products whose name_ru/description_ru
        currently hold a value that the import will leave untouched.
      * samples — up to 12 example field changes [{external_id, field, old, new}].
    """
    created = 0
    updated = 0
    skipped = 0
    changed: dict[str, int] = {f: 0 for f in CATALOG_FIELDS}
    cleared: dict[str, int] = {f: 0 for f in CATALOG_FIELDS}
    translations_protected = 0
    samples: list[dict] = []

    for product in products:
        norm = _normalize_product(product)
        if norm is None:
            skipped += 1
            continue

        existing = db.session.execute(
            db.select(PromProduct).where(
                PromProduct.external_id == norm["external_id"]
            )
        ).scalar_one_or_none()

        if not existing:
            created += 1
            continue

        updated += 1

        if preserve_translations and (
            existing.name_ru or existing.description_ru
        ):
            translations_protected += 1

        for field in CATALOG_FIELDS:
            old = getattr(existing, field)
            new = norm[field]
            if old == new:
                continue
            changed[field] += 1
            if old not in (None, "") and new in (None, ""):
                cleared[field] += 1
            if len(samples) < 12:
                samples.append(
                    {
                        "external_id": norm["external_id"],
                        "field": field,
                        "old": old,
                        "new": new,
                    }
                )

    return {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "total": created + updated,
        "changed": changed,
        "cleared": cleared,
        "translations_protected": translations_protected,
        "samples": samples,
    }


def save_catalog_products(products: list[dict], preserve_translations: bool = True) -> dict:
    """
    Upsert parsed products into the PromProduct table.

    Matches on external_id (unique). Updates existing records, inserts new ones.

    Two-channel separation (mirrors the feed Path B design):
      * Matcher / catalog channel — name (UA), brand, article, display_article,
        price, currency, page_url, image_url, images, description_ua.
        These are CATALOG-owned: a Horoshop export is authoritative for them,
        so they are always overwritten on update.
      * Translation channel — name_ru, description_ru.
        These are WORKER-owned (filled by the parallel translation terminals
        and stored only here in pp). Horoshop's RU fields are corrupted/stale
        (the YML <name>->"Назва модифікації (RU)" bug), so re-importing a
        Horoshop export would clobber clean translations with corrupted text.

    With preserve_translations=True (default), an UPDATE of an existing product
    leaves name_ru / description_ru UNTOUCHED — the catalog import can never
    overwrite worker output. INSERTs (brand-new products) still take all fields
    from the row, since no translation exists yet to protect.

    Set preserve_translations=False only when the source IS authoritative for
    RU text (e.g. a native [КАТАЛОГ] export produced after translation).

    Note: with translations preserved, this function NO LONGER does a blind
    full-overwrite of every column, so a partial Horoshop export is safer than
    before — but it will still null catalog-owned fields absent from the row.

    Args:
        products: List of product dicts from parse_catalog_file.
        preserve_translations: When True, never overwrite name_ru/description_ru
            on update of an existing product.

    Returns:
        Dict with keys: created, updated, skipped, total.
    """
    created = 0
    updated = 0
    skipped = 0

    for product in products:
        norm = _normalize_product(product)

        # Skip rows missing required fields
        if norm is None:
            skipped += 1
            continue

        # Check if product already exists
        existing = db.session.execute(
            db.select(PromProduct).where(
                PromProduct.external_id == norm["external_id"]
            )
        ).scalar_one_or_none()

        if existing:
            # Catalog-owned fields: Horoshop export is authoritative — always overwrite.
            for field in CATALOG_FIELDS:
                setattr(existing, field, norm[field])
            # Worker-owned translation fields: protected unless caller opts out.
            if not preserve_translations:
                for field in TRANSLATION_FIELDS:
                    setattr(existing, field, norm[field])
            existing.imported_at = datetime.now(timezone.utc)
            updated += 1
        else:
            new_product = PromProduct(
                external_id=norm["external_id"],
                name=norm["name"],
                name_ru=norm["name_ru"],
                brand=norm["brand"],
                article=norm["article"],
                display_article=norm["display_article"],
                price=norm["price"],
                currency=norm["currency"],
                page_url=norm["page_url"],
                image_url=norm["image_url"],
                images=norm["images"],
                description_ua=norm["description_ua"],
                description_ru=norm["description_ru"],
            )
            db.session.add(new_product)
            created += 1

    db.session.commit()

    return {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "total": created + updated,
    }
