"""НП (Новый проект, np.com.ua) native-Horoshop feed parser — CONTENT only.

Unlike rp_parser / excel_parser (which emit save_supplier_products() dicts),
this module reads the NP dealer-export XLSX purely as a *content lookup*:
``article -> {brand, description, description_ru, photos}``. It is consumed by
np_horoshop_file.build_np_file(), which joins this map onto the matcher DB
(SupplierProduct.article → ProductMatch → PromProduct.external_id) to emit the
native [КАТАЛОГ] Excel file the operator imports into Horoshop (Channel 2).

Why a separate parser:
  - NP cards have an EMPTY display_article, so ProductMatch is the only bridge,
    and the join key is the feed's «Артикул» column == SupplierProduct.article.
  - Price / availability for NP already flow through the matcher YML feed
    (Channel 1, Path B). This file carries the *content* the YML deliberately
    omits: description UA/RU + photo gallery. So the parser only extracts those.
  - brand is extracted NOT to write it (the output file omits «Бренд») but so
    the UI can filter rows by the 9 exclusive brands via checkboxes.

Feed shape (np.com.ua dealer-export xlsx, sheet "Worksheet", verified 2026-05-19,
690 rows / 24 cols — column positions are STABLE and are the whole contract):

  row 0      : header row (col B label == "Артикул"); skipped + sanity-checked.
  rows 1..N  : product rows. Fixed columns we read (0-based):
                 1  (B) Артикул              → join key (== sp.article)
                 3  (D) [КАТАЛОГ] Фото       → ';'-separated gallery URLs
                 7  (H) description_uk       → raw HTML, passed through as-is
                 9  (J) attr_brend_uk        → brand (for UI filtering)
                 16 (Q) description_ru       → raw HTML, passed through as-is

  Additionally located by HEADER LABEL (not a fixed index — they were never read
  by the original content-only parser and their position is not part of the
  legacy fixed-column contract). Verified live positions 2026-05-31 are
  title_uk=6, title_ru=15, categories_uk=8, but a label scan is used so a feed
  column reorder cannot silently mis-read these (consistent with the col-B
  «Артикул» header-drift guard philosophy):
                 (label) title_uk           → name UA  (create-card name)
                 (label) title_ru           → name RU  (create-card name, RU)
                 (label) categories_uk      → feed category path UA
                                              (e.g. "Холодильне обладнання/Льодогенератори")

Invariants:
  - Keyed by stripped Артикул string. Empty Артикул → row skipped (cannot join);
    flagged only if the row actually carries content (description / photo).
  - Duplicate Артикул → first wins, warning emitted (feed should be unique).
  - Descriptions kept verbatim (Horoshop renders the supplier's HTML).
  - Photos split on ';', stripped, de-duplicated, order preserved.
  - Header drift (col B != «Артикул») ABORTS the parse with one error — fixed
    indices are meaningless if the layout changed, and silently mis-reading a
    live store's content is worse than refusing.
"""

from __future__ import annotations

import logging

import openpyxl

logger = logging.getLogger(__name__)

# 0-based column indices in the NP native-Horoshop export. See module docstring.
_COL_ARTICLE = 1     # B — «Артикул» (join key)
_COL_PHOTOS = 3      # D — «[КАТАЛОГ] Фото» (';'-separated gallery)
_COL_DESC_UA = 7     # H — description_uk (raw HTML)
_COL_BRAND_UA = 9    # J — attr_brend_uk (brand, for UI filtering)
_COL_DESC_RU = 16    # Q — description_ru (raw HTML)

# Header LABELS for the name/category columns located by scan (not fixed index).
# Lower-cased exact-match against the header row labels.
_LBL_NAME_UA = "title_uk"
_LBL_NAME_RU = "title_ru"
_LBL_CATEGORY_UA = "categories_uk"

_NP_SHEET = "Worksheet"


def _cell(row, idx):
    """Safe positional read — openpyxl trims trailing Nones on short rows."""
    return row[idx] if len(row) > idx else None


def _clean_text(value) -> str | None:
    """Stripped string, or None for empty/None. Preserves inner HTML verbatim."""
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _split_photos(value) -> list[str]:
    """Split a ';'-separated gallery cell into a de-duplicated URL list.

    Order is preserved (Horoshop treats the first URL as the main image).
    """
    if value is None:
        return []
    text = str(value).strip()
    if not text:
        return []
    seen: set[str] = set()
    out: list[str] = []
    for part in text.split(";"):
        url = part.strip()
        if not url or url in seen:
            continue
        seen.add(url)
        out.append(url)
    return out


def _find_col(header, label: str) -> int | None:
    """0-based index of the first header cell whose label == ``label`` (ci).

    Returns None if no cell matches — the caller then leaves that field None
    and emits a non-fatal warning (these columns are additive, not part of the
    legacy fixed-column contract).
    """
    target = label.strip().lower()
    for idx, cell in enumerate(header):
        if isinstance(cell, str) and cell.strip().lower() == target:
            return idx
    return None


def parse_np_feed(file_path: str) -> tuple[dict[str, dict], list[str]]:
    """Parse an NP dealer-export xlsx into an article-keyed content map.

    Args:
        file_path: Path to the downloaded NP native-Horoshop .xlsx.

    Returns:
        ``(content, errors)`` where content maps ``article -> {"brand",
        "description", "description_ru", "photos": [url], "name", "name_ru",
        "category"}`` and errors is a list of human-readable warnings. The
        name/name_ru/category fields come from the title_uk/title_ru/
        categories_uk columns (located by header label); a missing label leaves
        that field None with a non-fatal warning. On col-B header drift the
        content is empty and errors holds a single abort message.
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[_NP_SHEET] if _NP_SHEET in wb.sheetnames else wb[wb.sheetnames[0]]

    content: dict[str, dict] = {}
    errors: list[str] = []

    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        wb.close()
        return {}, ["NP feed is empty (no header row)"]

    # Header sanity: column B must be the «Артикул» anchor. Fixed indices are
    # the whole contract — refuse rather than mis-read a live store's content.
    hdr_article = _cell(header, _COL_ARTICLE)
    if not (isinstance(hdr_article, str) and "артикул" in hdr_article.lower()):
        wb.close()
        return {}, [
            f"NP feed header mismatch: column B = {hdr_article!r}, "
            f"expected 'Артикул'. Aborting — fixed column indices can't be trusted."
        ]

    # Locate the name/category columns by header label (additive — not part of
    # the legacy fixed-index contract). A missing label is non-fatal: that field
    # stays None and a warning is appended, while brand/desc/photo keep working.
    col_name_ua = _find_col(header, _LBL_NAME_UA)
    col_name_ru = _find_col(header, _LBL_NAME_RU)
    col_category = _find_col(header, _LBL_CATEGORY_UA)
    for label, col in (
        (_LBL_NAME_UA, col_name_ua),
        (_LBL_NAME_RU, col_name_ru),
        (_LBL_CATEGORY_UA, col_category),
    ):
        if col is None:
            errors.append(
                f"NP feed missing optional column {label!r} — that field "
                f"will be None (brand/desc/photos unaffected)."
            )

    for row_idx, row in enumerate(rows, start=2):  # header consumed = row 1
        article_raw = _cell(row, _COL_ARTICLE)
        article = str(article_raw).strip() if article_raw is not None else ""

        if not article:
            # A keyless row with real content can never be joined — surface it.
            if _clean_text(_cell(row, _COL_DESC_UA)) or _split_photos(_cell(row, _COL_PHOTOS)):
                errors.append(f"Row {row_idx}: content row with empty Артикул, skipped")
            continue

        if article in content:
            errors.append(f"Row {row_idx}: duplicate Артикул {article!r}, keeping first")
            continue

        content[article] = {
            "brand": _clean_text(_cell(row, _COL_BRAND_UA)),
            "description": _clean_text(_cell(row, _COL_DESC_UA)),
            "description_ru": _clean_text(_cell(row, _COL_DESC_RU)),
            "photos": _split_photos(_cell(row, _COL_PHOTOS)),
            "name": _clean_text(_cell(row, col_name_ua)) if col_name_ua is not None else None,
            "name_ru": _clean_text(_cell(row, col_name_ru)) if col_name_ru is not None else None,
            "category": _clean_text(_cell(row, col_category)) if col_category is not None else None,
        }

    wb.close()

    logger.info(
        "NP parser: %d articles parsed (%d with photos), %d warnings",
        len(content),
        sum(1 for v in content.values() if v["photos"]),
        len(errors),
    )
    return content, errors
