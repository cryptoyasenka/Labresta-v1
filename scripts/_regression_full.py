"""Comprehensive regression detection across ALL 9 IMP columns.

Regression = broken import made things STRICTLY WORSE.
For RU fields:
  - imp has UA-leak AND exp clean (no UA-leak)
  - imp empty AND exp non-empty
  - imp structurally broken (desc only) AND exp not
  - imp shorter by >50% AND exp non-empty (suspect truncation)
For UA fields:
  - imp empty AND exp non-empty
  - imp structurally broken AND exp not
  - imp shorter by >50%
For names (mod_name, name):
  - imp == name_in_other_lang (UA leak pattern from Track #1)
  - imp empty AND exp non-empty
"""
import os, sys, re, json
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
from collections import Counter, defaultdict

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
IMP  = os.path.join(ROOT, 'horoshop-import-2026-05-20.xlsx')
EXP  = os.path.join(ROOT, 'horoshop-export 20.05.26.xlsx')

UA_ONLY = re.compile(r'[іїєґІЇЄҐ]')
def ua_leak(s): return bool(s) and bool(UA_ONLY.search(s))
def desc_struct(s):
    if not s: return None
    if s.startswith('<li>'): return '<li>'
    if s.startswith('`') or s.startswith(' `'): return 'backtick'
    if len(s) < 80 and not re.search(r'<(p|h[1-6]|ul|div)\b', s, re.I): return 'short'
    if not re.match(r'^\s*<(p|h[1-6]|ul|div|table|ol|hr|br)\b', s, re.I): return 'no-block'
    return None

def load(path):
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    hdr = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    idx = {h:i for i,h in enumerate(hdr)}
    rows = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        a = r[idx['Артикул']]
        if a is None: continue
        a = str(a).strip()
        if not a: continue
        rows[a] = {h: str(r[i] or '').strip() for h,i in idx.items()}
    return rows

imp = load(IMP)
exp = load(EXP)
print(f'IMP: {len(imp)}, EXP: {len(exp)}')

# Regression rules per column
RU_DESC_COLS = ['Описание товара (RU)']
UA_DESC_COLS = ['Описание товара (UA)']
RU_NAME_COLS = ['Название модификации (RU)', 'Название (RU)']
UA_NAME_COLS = ['Название модификации (UA)', 'Название (UA)']
RU_META_COLS = ['META keywords (RU)']
UA_META_COLS = ['META keywords (UA)']

def normalize(s):
    """Strip HTML, lowercase, collapse whitespace."""
    s = re.sub(r'<[^>]+>', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip().lower()
    return s

def is_truncation(iv, ev):
    """True if imp is a head-prefix of exp (content-wise), and significantly shorter."""
    if not iv or not ev: return False
    if len(iv) >= len(ev) * 0.5: return False
    if len(ev) < 60: return False
    ivn = normalize(iv)
    evn = normalize(ev)
    if not ivn or not evn or len(ivn) < 20: return False
    # Take first 60% of imp normalized — must appear at start of exp (within first 30% of exp)
    head = ivn[:max(20, int(len(ivn)*0.8))]
    return head in evn[:int(len(evn)*0.5)+len(head)]

def regressed(col, iv, ev, full_imp_row, full_exp_row):
    """Return list of regression labels for this cell, or [] if none."""
    out = []
    if iv == ev:
        return []

    is_ru_field = col in RU_DESC_COLS + RU_NAME_COLS + RU_META_COLS
    is_ua_field = col in UA_DESC_COLS + UA_NAME_COLS + UA_META_COLS

    # FILTER OUT translation success: RU field where imp clean RU and exp had UA-leak
    # In this case broken import was a SUCCESSFUL translation, NOT a regression
    if is_ru_field and ua_leak(ev) and not ua_leak(iv) and iv:
        return []  # broken import fixed UA-leak in RU field

    # Empty in imp but filled in exp (data wipe)
    if not iv and ev:
        out.append('imp_empty_exp_filled')
        return out

    # Real truncation (imp is prefix of exp)
    if is_truncation(iv, ev):
        out.append(f'imp_truncated_to_{len(iv)}_from_{len(ev)}')

    # RU fields: UA-leak regression
    if is_ru_field:
        if ua_leak(iv) and not ua_leak(ev):
            out.append('imp_has_ua_leak_exp_clean')

    # Desc structural regression (imp newly broken, exp clean)
    if col in RU_DESC_COLS + UA_DESC_COLS:
        ib = desc_struct(iv); eb = desc_struct(ev)
        if ib and not eb:
            out.append(f'imp_struct_{ib}_exp_ok')

    # Track #1 pattern: RU mod_name == UA name (wholesale UA copy into RU)
    if col == 'Название модификации (RU)':
        ua_name = full_imp_row.get('Название (UA)', '') or full_exp_row.get('Название (UA)', '')
        if iv and ua_name and iv == ua_name and iv != full_imp_row.get('Название (RU)', ''):
            out.append('mod_name_RU_equals_name_UA')

    return out

regressions = defaultdict(list)  # col -> [(art, labels, iv, ev)]
for art, imp_row in imp.items():
    exp_row = exp.get(art, {})
    if not exp_row: continue
    for col in [c for c in imp_row if c != 'Артикул']:
        iv = imp_row.get(col, '')
        ev = exp_row.get(col, '')
        labels = regressed(col, iv, ev, imp_row, exp_row)
        if labels:
            regressions[col].append((art, labels, iv[:80], ev[:80]))

print()
print('=== Regression summary ===')
total = 0
for col, lst in regressions.items():
    print(f'\n  {col}: {len(lst)} regressions')
    total += len(lst)
    cnt = Counter()
    for _, labels, _, _ in lst:
        for l in labels:
            cnt[l] += 1
    for lbl, n in cnt.most_common():
        print(f'    {n:5d}  {lbl}')
    # samples
    print(f'    samples (first 3):')
    for art, labels, iv, ev in lst[:3]:
        print(f'      {art} [{",".join(labels)}]')
        print(f'        imp={iv!r}')
        print(f'        exp={ev!r}')

print(f'\n=== TOTAL regression cells: {total} ===')

# Save per-column regression lists
out = {}
for col, lst in regressions.items():
    out[col] = [{'art':a, 'labels':l, 'imp_snip':iv, 'exp_snip':ev} for a,l,iv,ev in lst]
with open(os.path.join(ROOT, '.planning', '_regressions.json'), 'w', encoding='utf-8') as f:
    json.dump(out, f, ensure_ascii=False, indent=2)
print(f'\nSaved → .planning/_regressions.json')
