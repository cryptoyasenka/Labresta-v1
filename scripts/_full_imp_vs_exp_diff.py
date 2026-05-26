"""Full diff between broken import (943 SKUs) and pre-incident export.
For each of 9 IMP columns, count:
  - same as exp
  - imp != exp (real change applied by broken import)
For each "imp != exp" group, break down by whether the change made things
better, worse, or just different.

Reveals which columns the broken import actually mutated on LIVE.
"""
import os, sys, re
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
from collections import Counter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
IMP  = os.path.join(ROOT, 'horoshop-import-2026-05-20.xlsx')
EXP  = os.path.join(ROOT, 'horoshop-export 20.05.26.xlsx')

UA_ONLY = re.compile(r'[іїєґІЇЄҐ]')
def has_ua_leak(s): return bool(s) and bool(UA_ONLY.search(s))
def desc_struct_broken(s):
    if not s: return None
    if s.startswith('<li>'): return 'starts <li>'
    if s.startswith('`') or s.startswith(' `'): return 'leading backtick'
    if len(s) < 80 and not re.search(r'<(p|h[1-6]|ul|div)\b', s, re.I): return 'short, no block tags'
    if not re.match(r'^\s*<(p|h[1-6]|ul|div|table|ol|hr|br)\b', s, re.I): return 'no opening block tag'
    return None

def load_xlsx(path):
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    hdr = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    idx = {h:i for i,h in enumerate(hdr)}
    rows = {}
    art_i = idx['Артикул']
    for r in ws.iter_rows(min_row=2, values_only=True):
        a = r[art_i]
        if a is None: continue
        a = str(a).strip()
        if not a: continue
        rows[a] = {h: (r[i] or '') for h,i in idx.items()}
    return hdr, rows

imp_hdr, imp = load_xlsx(IMP)
exp_hdr, exp = load_xlsx(EXP)
print(f'IMP: {len(imp)} rows, {len(imp_hdr)} cols')
print(f'EXP: {len(exp)} rows, {len(exp_hdr)} cols')

# 9 IMP columns to check
IMP_COLS = [c for c in imp_hdr if c != 'Артикул']
print(f'\nIMP columns to diff: {IMP_COLS}')

print()
print('=== Per-column diff (imp 943 vs exp 5632, joined by Артикул) ===')
for col in IMP_COLS:
    same = diff = missing_in_exp = both_empty = imp_only_filled = exp_only_filled = 0
    diff_samples = []
    for art, imp_row in imp.items():
        if art not in exp:
            missing_in_exp += 1
            continue
        iv = str(imp_row.get(col, '')).strip()
        ev = str(exp.get(art, {}).get(col, '')).strip()
        if not iv and not ev:
            both_empty += 1
            same += 1
            continue
        if iv == ev:
            same += 1
        else:
            diff += 1
            if not iv: exp_only_filled += 1
            elif not ev: imp_only_filled += 1
            if len(diff_samples) < 3:
                diff_samples.append((art, iv[:80], ev[:80]))
    print(f'\n  --- {col} ---')
    print(f'    same: {same}/{len(imp)} (both_empty={both_empty})')
    print(f'    diff: {diff} (imp_only_filled={imp_only_filled} exp_only_filled={exp_only_filled})')
    print(f'    missing_in_exp: {missing_in_exp}')
    if diff_samples:
        print(f'    samples:')
        for art, iv, ev in diff_samples:
            print(f'      {art}: imp={iv!r}')
            print(f'              exp={ev!r}')
