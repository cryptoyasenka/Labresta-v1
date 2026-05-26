"""DRY-RUN MINI-5: pick 5 artikuls from different chunk eras and slice them
out of the already-built horoshop-import-2026-05-20.xlsx for visual verification
in Horoshop UI before the full 911-SKU upload.

Eras chosen (5 chunks across formats):
  chunk-002 — legacy format, earliest closed chunk (Hurakan era, pre-SKIP-НП lock)
  chunk-006 — modern v2 format, first post-legacy
  chunk-015 — middle of the closed range
  chunk-022 — modern v2 format, dense batching epoch
  chunk-028 — modern v3 format, newest closed chunk (POL5 forward-only cleanup baseline)

For each chunk, picks the FIRST artikul with edits (parse_chunk()[0][0]).
Writes horoshop-import-mini-5.xlsx with the same 9-col header as the full file.
"""
import sys, os, importlib.util
sys.stdout.reconfigure(encoding='utf-8')

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# Import the build script as a module so we can reuse parse_chunk()
spec = importlib.util.spec_from_file_location('_bhi', os.path.join(ROOT, 'scripts', '_build_horoshop_import.py'))
bhi = importlib.util.module_from_spec(spec)
spec.loader.exec_module(bhi)

ERA_CHUNKS = [2, 6, 15, 22, 28]

print('=' * 60)
print('MINI-5 DRY-RUN — picking 1 artikul per era chunk')
print('=' * 60)

picks = []  # list of (chunk_n, artikul, edits_count)
for n in ERA_CHUNKS:
    edits, skipped = bhi.parse_chunk(n)
    if not edits:
        print(f'  chunk-{n:03d}: NO EDITS (skipped: {len(skipped)}) — skip era')
        continue
    # edits is list of (artikul, field_canonical, value); we want first artikul that has ANY edit
    first_art = edits[0][0]
    art_edits = [e for e in edits if e[0] == first_art]
    picks.append((n, first_art, len(art_edits), art_edits))
    print(f'  chunk-{n:03d}: artikul={first_art} edits={len(art_edits)} fields={sorted({e[1] for e in art_edits})}')

print()
print(f'Picked {len(picks)} artikuls: {[p[1] for p in picks]}')
print()

# Now load the already-built full import xlsx, filter to the 5 rows
from openpyxl import load_workbook, Workbook

src_path = os.path.join(ROOT, 'horoshop-import-2026-05-20.xlsx')
out_path = os.path.join(ROOT, 'horoshop-import-mini-5.xlsx')

wb_src = load_workbook(src_path, read_only=False)
ws_src = wb_src.active

# Build artikul -> row_idx map in source xlsx
header = [c.value for c in ws_src[1]]
print('Source xlsx header:')
for i, h in enumerate(header):
    print(f'  col{i+1}: {h!r}')
print()

art_col_idx = header.index('Артикул') + 1  # 1-based
art_to_row = {}
for row_idx in range(2, ws_src.max_row + 1):
    a = ws_src.cell(row=row_idx, column=art_col_idx).value
    if a is not None:
        art_to_row[str(a).strip()] = row_idx

print(f'Source has {len(art_to_row)} unique artikuls')
print()

# Create destination workbook
wb_dst = Workbook()
ws_dst = wb_dst.active
ws_dst.title = ws_src.title or 'Sheet1'

# Copy header row 1:1
for col_idx, val in enumerate(header, start=1):
    ws_dst.cell(row=1, column=col_idx, value=val)

# Copy the 5 picked rows
dst_row = 2
copied = []
for (chunk_n, artikul, n_edits, _art_edits) in picks:
    src_row = art_to_row.get(str(artikul))
    if not src_row:
        print(f'  WARN: artikul {artikul} (chunk-{chunk_n:03d}) NOT in built xlsx — skip')
        continue
    for col_idx in range(1, len(header) + 1):
        v = ws_src.cell(row=src_row, column=col_idx).value
        ws_dst.cell(row=dst_row, column=col_idx, value=v)
    copied.append((chunk_n, artikul, src_row, dst_row))
    dst_row += 1

wb_dst.save(out_path)

print()
print('=' * 60)
print(f'WROTE: {out_path}')
print(f'Rows: {dst_row - 1} (excl header)')
print('=' * 60)
for chunk_n, artikul, src_row, dst_row_n in copied:
    print(f'  row {dst_row_n}: artikul={artikul} (chunk-{chunk_n:03d}, src_row={src_row})')

print()
print('Verification sample — Назв.мод (RU) for each row:')
nazv_mod_ru_col = header.index('Назв.мод (RU)') + 1 if 'Назв.мод (RU)' in header else None
nazv_mod_ua_col = header.index('Назв.мод (UA)') + 1 if 'Назв.мод (UA)' in header else None
opisanie_ru_col = header.index('Описание товара (RU)') + 1 if 'Описание товара (RU)' in header else None

for row_idx in range(2, dst_row):
    a = ws_dst.cell(row=row_idx, column=art_col_idx).value
    if nazv_mod_ru_col:
        v = ws_dst.cell(row=row_idx, column=nazv_mod_ru_col).value
        print(f'  art={a}  nm_ru={(v[:80] + "…") if v and len(v) > 80 else v!r}')
    if opisanie_ru_col:
        v = ws_dst.cell(row=row_idx, column=opisanie_ru_col).value
        snippet = (v[:120] + '…') if v and len(v) > 120 else v
        print(f'           opis_ru={snippet!r}')

print()
print('DONE.')
