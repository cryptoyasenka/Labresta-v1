"""Audit: which rows in horoshop-import-2026-05-20.xlsx have suspiciously
short / fragmentary values in description fields (likely partial-fix entries
that the parser incorrectly treated as full replacements).

A "suspicious" cell:
- starts with <li>  (single bullet without surrounding <ul>/<p>)
- is a single short line without any tags  (likely a name fragment)
- starts with something other than <p>, <h1>-<h6>, <div>, <ul> for Описание
"""
import sys, os, re
sys.stdout.reconfigure(encoding='utf-8')

# write to file too so we don't lose output to pipe buffering
LOG = open(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), '.planning', '_audit_partial_fix.txt'), 'w', encoding='utf-8')
def p(*args):
    s = ' '.join(str(a) for a in args)
    print(s, flush=True)
    LOG.write(s + '\n')
    LOG.flush()

from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
src = os.path.join(ROOT, 'horoshop-import-2026-05-20.xlsx')

wb = load_workbook(src, read_only=True)
ws = wb.active

header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
header = list(header)
art_col = header.index('Артикул') + 1

DESC_COLS = [
    ('Описание товара (RU)', header.index('Описание товара (RU)') + 1 if 'Описание товара (RU)' in header else None),
    ('Описание товара (UA)', header.index('Описание товара (UA)') + 1 if 'Описание товара (UA)' in header else None),
]

# Suspicion rules for Описание fields
def is_suspicious_desc(v):
    if v is None: return False, ''
    s = str(v).strip()
    if not s: return False, ''
    # Strong sign of partial-fix: starts with <li> (orphan bullet)
    if s.startswith('<li>'): return True, 'starts <li>'
    # Strong sign: <ul>/<li> wrap of only one bullet (single <li> wrapped) — keep
    # Strong sign: very short and no block tags
    if len(s) < 80 and not re.search(r'<(p|h[1-6]|ul|div)\b', s, re.I):
        return True, 'short, no block tags'
    # Strong sign: starts with closing tag or text fragment without opening block
    if not re.match(r'^\s*<(p|h[1-6]|ul|div|table|ol|hr|br)\b', s, re.I):
        return True, f'no opening block tag (start: {s[:40]!r})'
    return False, ''

# Suspicion for name fields: should be a single line of plain text, not HTML fragment
NAME_COLS = [
    ('Название модификации (RU)', header.index('Название модификации (RU)') + 1 if 'Название модификации (RU)' in header else None),
    ('Название модификации (UA)', header.index('Название модификации (UA)') + 1 if 'Название модификации (UA)' in header else None),
    ('Название (RU)', header.index('Название (RU)') + 1 if 'Название (RU)' in header else None),
    ('Название (UA)', header.index('Название (UA)') + 1 if 'Название (UA)' in header else None),
]
def is_suspicious_name(v):
    if v is None: return False, ''
    s = str(v).strip()
    if not s: return False, ''
    if '<' in s and '>' in s: return True, 'contains HTML tags'
    if len(s) > 250: return True, 'very long (>250 chars)'
    return False, ''

# Suspicion for META keywords: comma-separated short phrases
META_COLS = [
    ('META keywords (RU)', header.index('META keywords (RU)') + 1 if 'META keywords (RU)' in header else None),
    ('META keywords (UA)', header.index('META keywords (UA)') + 1 if 'META keywords (UA)' in header else None),
]
def is_suspicious_meta(v):
    if v is None: return False, ''
    s = str(v).strip()
    if not s: return False, ''
    if '<' in s and '>' in s: return True, 'contains HTML tags'
    return False, ''

p('Scanning', ws.max_row - 1, 'rows via iter_rows...')

flagged = []
row_idx = 1  # we skip header
for row in ws.iter_rows(min_row=2, values_only=True):
    row_idx += 1
    art = row[art_col - 1]
    if art is None: continue
    art_str = str(art).strip()
    row_flags = []
    for col_name, col_idx in DESC_COLS:
        if col_idx is None: continue
        v = row[col_idx - 1]
        ok, reason = is_suspicious_desc(v)
        if ok:
            row_flags.append((col_name, reason, (str(v) or '')[:120]))
    for col_name, col_idx in NAME_COLS:
        if col_idx is None: continue
        v = row[col_idx - 1]
        ok, reason = is_suspicious_name(v)
        if ok:
            row_flags.append((col_name, reason, (str(v) or '')[:120]))
    for col_name, col_idx in META_COLS:
        if col_idx is None: continue
        v = row[col_idx - 1]
        ok, reason = is_suspicious_meta(v)
        if ok:
            row_flags.append((col_name, reason, (str(v) or '')[:120]))
    if row_flags:
        flagged.append((row_idx, art_str, row_flags))

p(f'\nFlagged: {len(flagged)} / {ws.max_row - 1} rows ({100*len(flagged)/(ws.max_row-1):.1f}%)')
p()

# Group by reason
from collections import Counter
reason_counts = Counter()
for _, _, row_flags in flagged:
    for col, reason, _ in row_flags:
        reason_counts[(col, reason)] += 1

p('Top issue patterns:')
for (col, reason), cnt in reason_counts.most_common(20):
    print(f'  {cnt:5d}  {col}  —  {reason}')

p()
p('First 20 flagged rows (sample):')
for row_idx, art_str, row_flags in flagged[:20]:
    print(f'\n  row={row_idx} artikul={art_str}')
    for col, reason, snippet in row_flags:
        print(f'    {col}: [{reason}]')
        print(f'      → {snippet!r}')
