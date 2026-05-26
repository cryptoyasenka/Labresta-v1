"""Track #1 fix: Название модификации (RU) UA-leak.

Reads horoshop-import-2026-05-20.xlsx (the broken import that's already live),
finds rows where:
    Название модификации (RU) != Название (RU)  AND  Название модификации (RU) == Название (UA)
For each such row, writes a TWO-column xlsx [Артикул, Название модификации (RU)]
where Название модификации (RU) = the CORRECT Название (RU) from the same source.

NEVER writes empty cells. NEVER includes other columns.

Output: horoshop-fix-mod-name-2026-05-21.xlsx
"""
import os, sys
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import load_workbook, Workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, 'horoshop-import-2026-05-20.xlsx')
OUT = os.path.join(ROOT, 'horoshop-fix-mod-name-2026-05-21.xlsx')

wb_in = load_workbook(SRC, read_only=True)
ws_in = wb_in.active
header = list(next(ws_in.iter_rows(min_row=1, max_row=1, values_only=True)))

art_i = header.index('Артикул')
mod_ru_i = header.index('Название модификации (RU)')
name_ru_i = header.index('Название (RU)')
name_ua_i = header.index('Название (UA)')

fixes = []
for row in ws_in.iter_rows(min_row=2, values_only=True):
    art = row[art_i]
    if art is None: continue
    art = str(art).strip()
    if not art: continue
    mod_ru = (row[mod_ru_i] or '').strip()
    name_ru = (row[name_ru_i] or '').strip()
    name_ua = (row[name_ua_i] or '').strip()
    # Strict pattern: mod_ru differs from correct RU AND matches UA name
    if mod_ru and name_ru and mod_ru != name_ru and mod_ru == name_ua:
        fixes.append((art, name_ru))

print(f'Building Track #1 fix xlsx: {len(fixes)} rows')

wb_out = Workbook()
ws_out = wb_out.active
ws_out.title = 'fix-mod-name'
ws_out.append(['Артикул', 'Название модификации (RU)'])
for art, val in fixes:
    ws_out.append([art, val])
wb_out.save(OUT)
print(f'Wrote: {OUT}')
print()
print('Preview:')
for art, val in fixes:
    print(f'  {art:>12}  →  {val!r}')
