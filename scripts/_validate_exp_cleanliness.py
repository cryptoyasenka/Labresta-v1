"""Validate pre-incident export cleanliness for broken SKUs:
1. exp desc_ru must NOT have UA-leak (і/ї/є/ґ)
2. exp desc_ru/ua must NOT be structurally broken (orphan <li>, leading `, etc.)
3. exp meta_kw_ru must NOT have UA-leak

If exp is dirty for our broken SKU subset → restore from exp would propagate
broken data. We need to know.
"""
import os, sys, json, re
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
from collections import Counter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXP  = os.path.join(ROOT, 'horoshop-export 20.05.26.xlsx')
ROWS = os.path.join(ROOT, '.planning', '_recovery_audit_rows.json')

with open(ROWS, 'r', encoding='utf-8') as f:
    broken = json.load(f)
print(f'broken SKUs to validate in exp: {len(broken)}')

UA_ONLY = re.compile(r'[іїєґІЇЄҐ]')
def desc_struct_broken(s):
    if not s: return None
    if s.startswith('<li>'): return 'starts <li>'
    if s.startswith('`') or s.startswith(' `'): return 'leading backtick'
    if len(s) < 80 and not re.search(r'<(p|h[1-6]|ul|div)\b', s, re.I):
        return 'short, no block tags'
    if not re.match(r'^\s*<(p|h[1-6]|ul|div|table|ol|hr|br)\b', s, re.I):
        return 'no opening block tag'
    return None

wb = load_workbook(EXP, read_only=True)
ws = wb.active
hdr = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
idx = {h:i for i,h in enumerate(hdr)}
art_i = idx['Артикул']
dru_i = idx['Описание товара (RU)']
dua_i = idx['Описание товара (UA)']
mru_i = idx['META keywords (RU)']
mua_i = idx['META keywords (UA)']

broken_set = set(broken.keys())
issues = Counter()
samples = {}

for row in ws.iter_rows(min_row=2, values_only=True):
    art = row[art_i]
    if art is None: continue
    art = str(art).strip()
    if art not in broken_set: continue
    dru = (row[dru_i] or '').strip()
    dua = (row[dua_i] or '').strip()
    mru = (row[mru_i] or '').strip()
    mua = (row[mua_i] or '').strip()

    if UA_ONLY.search(dru):
        issues['exp_desc_ru_has_ua_leak'] += 1
        samples.setdefault('exp_desc_ru_ua_leak', []).append((art, dru[:120]))
    sb = desc_struct_broken(dru)
    if sb:
        issues[f'exp_desc_ru_struct_{sb}'] += 1
        samples.setdefault(f'exp_desc_ru_{sb}', []).append((art, dru[:120]))
    sb = desc_struct_broken(dua)
    if sb:
        issues[f'exp_desc_ua_struct_{sb}'] += 1
        samples.setdefault(f'exp_desc_ua_{sb}', []).append((art, dua[:120]))
    if UA_ONLY.search(mru):
        issues['exp_meta_kw_ru_has_ua_leak'] += 1
        samples.setdefault('exp_meta_kw_ru_ua_leak', []).append((art, mru[:120]))
    if not dru:
        issues['exp_desc_ru_EMPTY'] += 1
    if not dua:
        issues['exp_desc_ua_EMPTY'] += 1
    if not mru:
        issues['exp_meta_kw_ru_EMPTY'] += 1

print()
print(f'Issues in exp for {len(broken_set)} broken SKUs:')
for k,v in issues.most_common():
    print(f'  {v:5d}  {k}')

print()
print('Samples (up to 5 per issue):')
for k, lst in samples.items():
    print(f'\n  --- {k} ---')
    for art, snip in lst[:5]:
        print(f'    {art}: {snip!r}')

# Save dirty list
dirty_path = os.path.join(ROOT, '.planning', '_exp_dirty_broken_sku.json')
dirty = {k:lst for k,lst in samples.items()}
with open(dirty_path, 'w', encoding='utf-8') as f:
    json.dump(dirty, f, ensure_ascii=False, indent=2)
print(f'\nDirty SKU samples → {dirty_path}')
