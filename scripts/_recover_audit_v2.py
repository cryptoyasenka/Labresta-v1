"""Recovery audit V2: precise delta between broken import and pre-incident export.

Classification:
  A: imp != exp, exp clean       → FIX (restore from exp)
  B: imp != exp, exp dirty       → uncertain (broken import potentially made worse, but exp also bad)
  C: imp == exp, both dirty      → DO NOT touch (pre-existing, will be cleaned by translation pipeline)
  D: imp clean and exp clean     → not in broken list

We only fix Category A. B requires human review.
"""
import os, sys, json, re
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
from collections import Counter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PLAN = os.path.join(ROOT, '.planning')
IMP  = os.path.join(ROOT, 'horoshop-import-2026-05-20.xlsx')
EXP  = os.path.join(ROOT, 'horoshop-export 20.05.26.xlsx')

UA_ONLY = re.compile(r'[іїєґІЇЄҐ]')
def has_ua_leak(s):
    return bool(s) and bool(UA_ONLY.search(s))

def desc_struct_broken(s):
    if not s: return None
    if s.startswith('<li>'): return 'starts <li>'
    if s.startswith('`') or s.startswith(' `'): return 'leading backtick'
    if len(s) < 80 and not re.search(r'<(p|h[1-6]|ul|div)\b', s, re.I):
        return 'short, no block tags'
    if not re.match(r'^\s*<(p|h[1-6]|ul|div|table|ol|hr|br)\b', s, re.I):
        return 'no opening block tag'
    return None

def ru_is_broken(s):
    """Combined: UA-leak OR struct broken."""
    if has_ua_leak(s): return 'ua_leak'
    sb = desc_struct_broken(s)
    if sb: return sb
    return None

def ua_is_struct_broken(s):
    """For UA we only flag structural issues, NOT presence of UA letters (legitimate)."""
    return desc_struct_broken(s)

def kw_ru_is_broken(s):
    return 'ua_leak' if has_ua_leak(s) else None

# --- Load IMP ---
wb = load_workbook(IMP, read_only=True)
ws = wb.active
hdr = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
i_imp = {h:i for i,h in enumerate(hdr)}
imp = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    art = row[i_imp['Артикул']]
    if art is None: continue
    art = str(art).strip()
    if not art: continue
    imp[art] = {
        'desc_ru':    (row[i_imp['Описание товара (RU)']] or '').strip(),
        'desc_ua':    (row[i_imp['Описание товара (UA)']] or '').strip(),
        'meta_kw_ru': (row[i_imp['META keywords (RU)']] or '').strip(),
    }
print(f'IMP: {len(imp)} rows')

# --- Load EXP ---
wb = load_workbook(EXP, read_only=True)
ws = wb.active
hdr = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
i_exp = {h:i for i,h in enumerate(hdr)}
exp = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    art = row[i_exp['Артикул']]
    if art is None: continue
    art = str(art).strip()
    if not art: continue
    exp[art] = {
        'desc_ru':    (row[i_exp['Описание товара (RU)']] or '').strip(),
        'desc_ua':    (row[i_exp['Описание товара (UA)']] or '').strip(),
        'meta_kw_ru': (row[i_exp['META keywords (RU)']] or '').strip(),
    }
print(f'EXP: {len(exp)} rows')

# --- Per-field classification ---
def classify(field):
    """field in {'desc_ru', 'desc_ua', 'meta_kw_ru'}"""
    check = ua_is_struct_broken if field == 'desc_ua' else (ru_is_broken if field.startswith('desc') else kw_ru_is_broken)
    cats = Counter()
    A = []  # fixable
    B = []  # uncertain
    C = []  # pre-existing
    for art, imp_row in imp.items():
        imp_val = imp_row[field]
        exp_row = exp.get(art, {})
        exp_val = exp_row.get(field, '')
        imp_b = check(imp_val)
        exp_b = check(exp_val)
        if not imp_b:
            continue  # imp is fine, no fix needed
        # imp is broken
        if imp_val == exp_val:
            cats['C_pre_existing'] += 1
            C.append((art, imp_b))
        else:
            if not exp_b:
                cats['A_fixable'] += 1
                A.append((art, imp_b, exp_val))
            else:
                cats['B_uncertain'] += 1
                B.append((art, imp_b, exp_b, imp_val[:60], exp_val[:60]))
    return cats, A, B, C

print()
print('=== Classification ===')
result = {}
for f in ('desc_ru', 'desc_ua', 'meta_kw_ru'):
    cats, A, B, C = classify(f)
    print(f'\n--- {f} ---')
    print(f'  A fixable (imp broken, exp clean, differ):       {len(A)}')
    print(f'  B uncertain (imp broken, exp also broken, differ):{len(B)}')
    print(f'  C pre-existing (imp==exp, both broken):           {len(C)}')
    result[f] = {'A': A, 'B': B, 'C': C}
    # Sample B
    if B:
        print(f'  B samples (first 3):')
        for art, ib, eb, iv, ev in B[:3]:
            print(f'    {art}: imp[{ib}]={iv!r}')
            print(f'             exp[{eb}]={ev!r}')

# Save Category A directly for fix-builder
fix_path = os.path.join(PLAN, '_recovery_fixable_A.json')
fixable = {
    'desc_ru': [{'art':a, 'reason':r, 'exp_value':v} for a,r,v in result['desc_ru']['A']],
    'desc_ua': [{'art':a, 'reason':r, 'exp_value':v} for a,r,v in result['desc_ua']['A']],
    'meta_kw_ru': [{'art':a, 'reason':r, 'exp_value':v} for a,r,v in result['meta_kw_ru']['A']],
}
with open(fix_path, 'w', encoding='utf-8') as f:
    json.dump(fixable, f, ensure_ascii=False, indent=2)
print(f'\nFixable (Category A) → {fix_path}')

# Save full classification
full_path = os.path.join(PLAN, '_recovery_classification.json')
serialize = {
    f: {
        'A': [{'art':a, 'reason':r} for a,r,v in result[f]['A']],
        'B': [{'art':a, 'imp_reason':ib, 'exp_reason':eb, 'imp_snip':iv, 'exp_snip':ev} for a,ib,eb,iv,ev in result[f]['B']],
        'C': [{'art':a, 'reason':r} for a,r in result[f]['C']],
    } for f in ('desc_ru','desc_ua','meta_kw_ru')
}
with open(full_path, 'w', encoding='utf-8') as f:
    json.dump(serialize, f, ensure_ascii=False, indent=2)
print(f'Full classification → {full_path}')
