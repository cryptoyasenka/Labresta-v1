"""Inspect the Labresta X Excel — list columns and sample rows."""
import openpyxl


def main():
    wb = openpyxl.load_workbook("Лабреста Х.xlsx", read_only=True, data_only=True)
    for ws in wb.worksheets:
        print(f"=== Sheet: {ws.title} ({ws.max_row} rows, {ws.max_column} cols) ===")
        rows = list(ws.iter_rows(values_only=True, max_row=3))
        for i, r in enumerate(rows):
            print(f"  Row {i}: {r[:20]}")
        print()


if __name__ == "__main__":
    main()
