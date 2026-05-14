"""
Split `horoshop-export 13.05.26.xlsx` into adaptive chunks for UA/RU revision.

Strategy:
- Iterate rows in original (Артикул) order — Yana decision: no priority bucketing
- Sum text length across 10 operator-owned UA/RU pairs per SKU
- Flush a new chunk when accumulated chars cross TARGET_CHARS (or hard cap MAX_SKU)
- Each chunk = full copy of all 59 columns for its SKU slice → safe partial upload later

Output:
  .planning/translation-audit/chunks/chunk-NN.xlsx    (each, full 59 cols, ~30-100 SKU)
  .planning/translation-audit/chunks/INDEX.md         (status board for all chunks)
"""
from __future__ import annotations
import sys
from pathlib import Path

import openpyxl
from openpyxl import Workbook

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "horoshop-export 13.05.26.xlsx"
OUT_DIR = ROOT / ".planning" / "translation-audit" / "chunks"

# Adaptive chunking — target ~50K input tokens per chat session.
# Cyrillic UTF-8: ~3-4 chars per token, so 50K tokens ≈ 150K chars.
TARGET_CHARS = 150_000   # soft threshold — flush when crossed
MIN_SKU = 30             # don't flush a chunk smaller than this unless EOF
MAX_SKU = 100            # hard cap regardless of length

# Operator-owned text pairs we'll be revising in each chunk
PAIR_HEADERS = [
    "Название (UA)", "Название (RU)",
    "Название модификации (UA)", "Название модификации (RU)",
    "HTML title (UA)", "HTML title (RU)",
    "META keywords (UA)", "META keywords (RU)",
    "META description (UA)", "META description (RU)",
    "h1 заголовок (UA)", "h1 заголовок (RU)",
    "Описание товара (UA)", "Описание товара (RU)",
    "Короткое описание (UA)", "Короткое описание (RU)",
    "Текст акции (UA)", "Текст акции (RU)",
    "Описание для маркетплейсов (UA)", "Описание для маркетплейсов (RU)",
]


def row_text_chars(row: tuple, indices: list[int]) -> int:
    total = 0
    for i in indices:
        v = row[i]
        if isinstance(v, str):
            total += len(v)
    return total


def write_chunk(headers: list, rows: list, out_path: Path):
    """Write one chunk XLSX preserving the full 59-column structure."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(headers)
    for r in rows:
        ws.append(list(r))
    wb.save(str(out_path))


def main():
    print(f"Reading {XLSX} ...")
    wb_in = openpyxl.load_workbook(str(XLSX), read_only=True, data_only=True)
    ws_in = wb_in.active
    rows_iter = ws_in.iter_rows(values_only=True)
    headers = list(next(rows_iter))
    col = {h: i for i, h in enumerate(headers) if h}
    text_indices = [col[h] for h in PAIR_HEADERS if h in col]
    sku_i = col["Артикул"]
    brand_i = col["Бренд"]

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    # Clean prior chunks if any
    for old in OUT_DIR.glob("chunk-*.xlsx"):
        old.unlink()
    for old in OUT_DIR.glob("chunk-*-fixed.xlsx"):
        old.unlink()

    chunks_info = []
    current_rows: list = []
    current_chars = 0
    chunk_idx = 1

    def flush(force=False):
        nonlocal current_rows, current_chars, chunk_idx
        if not current_rows:
            return
        if not force and len(current_rows) < MIN_SKU:
            return  # too small, keep accumulating
        name = f"chunk-{chunk_idx:03d}.xlsx"
        out_path = OUT_DIR / name
        write_chunk(headers, current_rows, out_path)
        skus = [str(r[sku_i]) for r in current_rows]
        brands = sorted({(r[brand_i] or "—").strip() for r in current_rows})
        chunks_info.append({
            "n": chunk_idx,
            "file": name,
            "sku_count": len(current_rows),
            "chars": current_chars,
            "sku_first": skus[0],
            "sku_last": skus[-1],
            "brands": brands,
        })
        print(f"  flushed chunk-{chunk_idx:03d}: {len(current_rows):3d} SKU, {current_chars:>7,d} chars, brands={','.join(brands[:3])}{'…' if len(brands)>3 else ''}")
        chunk_idx += 1
        current_rows = []
        current_chars = 0

    total_rows = 0
    for row in rows_iter:
        if row is None:
            continue
        total_rows += 1
        chars = row_text_chars(row, text_indices)
        current_rows.append(row)
        current_chars += chars
        # Flush condition: hit target chars (and have at least MIN_SKU) OR hit hard cap
        if len(current_rows) >= MAX_SKU:
            flush(force=True)
        elif current_chars >= TARGET_CHARS and len(current_rows) >= MIN_SKU:
            flush(force=True)

    # Trailing chunk
    flush(force=True)

    print(f"\nTotal: {total_rows} rows → {len(chunks_info)} chunks")

    # INDEX.md
    md = [
        "# Translation chunks — UA/RU revision workflow",
        "",
        f"- Source: `{XLSX.name}` ({total_rows} SKU)",
        f"- Total chunks: **{len(chunks_info)}**",
        f"- Target ~{TARGET_CHARS:,} chars input per chunk (≈50K tokens)",
        f"- Bounds: {MIN_SKU}-{MAX_SKU} SKU per chunk",
        "",
        "## Workflow per chunk",
        "",
        "1. Open chunk-NN.xlsx — that's the source (do not modify)",
        "2. Revise UA + RU operator text fields per memory rules `feedback_labresta_ua_ru_translation_rules.md`",
        "3. Write chunk-NN-fixed.xlsx (same structure, edited cells)",
        "4. Write chunk-NN-diff.md (human-readable list of changes)",
        "5. If any doubts → chunk-NN-questions.md (Yana batch-answers)",
        "6. Update status below: `pending → in_progress → done`",
        "",
        "## Chunks",
        "",
        "| # | File | SKU | Chars | First Артикул | Last Артикул | Brands | Status |",
        "|--:|---|--:|--:|---|---|---|---|",
    ]
    for c in chunks_info:
        brands_str = ", ".join(c["brands"][:4]) + ("…" if len(c["brands"]) > 4 else "")
        md.append(
            f"| {c['n']:03d} | `{c['file']}` | {c['sku_count']} | {c['chars']:,} | "
            f"`{c['sku_first']}` | `{c['sku_last']}` | {brands_str} | pending |"
        )
    md.append("")
    md.append("## Per-chunk artifacts (filled as work progresses)")
    md.append("")
    md.append("- `chunk-NN.xlsx` — source slice (read-only)")
    md.append("- `chunk-NN-fixed.xlsx` — revised version")
    md.append("- `chunk-NN-diff.md` — change list")
    md.append("- `chunk-NN-questions.md` — Yana's batch-question file")
    md.append("")

    index_path = OUT_DIR / "INDEX.md"
    index_path.write_text("\n".join(md), encoding="utf-8")
    print(f"\nWrote {index_path}")


if __name__ == "__main__":
    main()
