"""Aggregate translation edits from closed chunks into a Horoshop-import xlsx.

DRY-RUN by default. Use --write to produce xlsx.

Two edit types (auto-detected per Стало content):
  FULL    — Стало = entire new field value (e.g., RU=UA full translation rewrites).
            Behavior: overwrite cell with Стало verbatim.
  PARTIAL — Стало = small fragment(s) to be PATCHED INTO current cell value
            (e.g., `.`→`,` decimal fix, single-line typo fix, single <li> patch).
            Behavior: load current value from --base, do sequential
                          new = current.replace(was, stale)
            for each fragment pair. Multi-fragment partials (` + `, ` / `, ` … `
            separators inside backticks) handle each pair sequentially.

Detection rules:
  - Names/META/Title fields → ALWAYS FULL.
  - Описание: multi-fragment OR starts <li> w/o block tags OR plain short text → PARTIAL.
              <p>+<ul>/<h2> shape OR length ≥ 250 OR starts with block tag → FULL.

Parses 3 diff.md formats:
  A) Modern v2  (006–028): ### Field — note  + **Было[ (UA|RU)]:**  + **Стало[ (UA|RU)]:**
  B) Modern v3  (027–028+): **Поле:** Field name  + **Было:** value  + **Стало:** value
  C) Legacy v1  (002, 005): like A with **Стало (RU):** etc, often multi-line single-backtick
"""

import argparse
import re
import sys
from collections import defaultdict, namedtuple

sys.stdout.reconfigure(encoding="utf-8")

CLOSED_CHUNKS = [2, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20,
                 21, 22, 23, 24, 25, 26, 27, 28]

CHUNK_DIR = ".planning/translation-audit/chunks"

# ---- SKU section split ----------------------------------------------------------
SKU_SPLIT = re.compile(r"^(## SKU\s+\d+/\d+\s+—\s+.*?)$", re.M)
ART_PAREN = re.compile(r"\(Артикул\s+(\d{6,})\)")
ART_BACKTICK = re.compile(r"`\s*(\d{6,})\s*`")

# ---- Field declaration patterns -------------------------------------------------
RE_FIELD_H3 = re.compile(r"^###\s+(.+?)\s*(?:—.*)?$", re.M)
RE_FIELD_POLE = re.compile(r"^\*\*Поле:\*\*\s+(.+?)\s*$", re.M)

# ---- Было / Стало head markers --------------------------------------------------
# Locale qualifier variants: bare, ` UA`, ` (UA)`, ` (UA \`<li>\`)`, ` (RU)`.
LOC_Q = r"(?:\s*[A-ZА-Я]{2,}|\s*\([^)]*\))?"
RE_WAS_HEAD = re.compile(rf"\*\*Было{LOC_Q}:\*\*[^\S\n]*", re.M)
RE_STALO_HEAD = re.compile(rf"\*\*Стало{LOC_Q}:\*\*[^\S\n]*", re.M)

# No-change markers we explicitly skip
RE_NOCHANGE = re.compile(
    r"^(?:без изменений|не трогаем|без правок|N/A|—\s*$|–\s*$)\b", re.I
)

# Boundary patterns that end a Было/Стало payload chunk
# (next field decl, next Было/Стало head, SKU header, --- separator).
# We intentionally DO NOT use `\n\n` so fenced blocks with blank lines survive.
RE_BOUNDARY = re.compile(
    rf"(?:^---\s*$|^## SKU\s|^###\s|^\*\*Поле:\*\*\s|"
    rf"^\*\*Было{LOC_Q}:\*\*|^\*\*Стало{LOC_Q}:\*\*)",
    re.M,
)


def normalize_field(raw):
    """Canonicalize field declaration to a stable key."""
    s = raw.strip()
    s = re.split(r"\s+—\s+", s, maxsplit=1)[0].strip()
    s = s.replace("**", "").strip()
    s = re.sub(r"\s+", " ", s)
    if re.match(r"^Описание\s*\([UR][AU]\)$", s):
        s = s.replace("Описание", "Описание товара")
    if re.match(r"^Назв\.?\s?мод\.?\s?\([UR][AU]\)$", s):
        s = re.sub(r"^Назв\.?\s?мод", "Название модификации", s)
    s = s.replace("Назв. модификации", "Название модификации")
    return s


def extract_payload_chunk(body, head_match_end):
    """Return text from head_match_end up to next boundary (next head/field/SKU/---)."""
    rest = body[head_match_end:]
    bm = RE_BOUNDARY.search(rest)
    return rest[:bm.start()] if bm else rest


def strip_ellipsis_brackets(s):
    """Strip leading/trailing ellipsis markers (U+2026 '…') used as 'continues elsewhere'."""
    return s.strip(' \t…').strip()


def parse_payload(raw):
    """Parse raw payload chunk into (fragments_list, kind_string).

    Kinds: 'fence', 'quote', 'inline' (1+ backtick frags), 'text', 'nochange', 'empty'.
    """
    s = raw.lstrip()
    if not s:
        return [], 'empty'
    if RE_NOCHANGE.match(s.rstrip()):
        return [], 'nochange'
    # Fenced ```...```
    if s.startswith('```'):
        m = re.match(r"```(?:html|xml)?\s*\n(?P<p>.*?)\n```", s, re.S)
        if m:
            return [m.group('p').strip()], 'fence'
    # Quoted `> ...` markdown blockquote
    if s.startswith('>'):
        out_lines = []
        for line in s.splitlines():
            if line.startswith('>'):
                out_lines.append(re.sub(r"^>\s?", "", line))
            elif line.strip() == '':
                out_lines.append('')
            else:
                break
        joined = '\n'.join(out_lines).rstrip()
        if joined:
            return [joined], 'quote'
    # Backtick inline (1+ fragments, possibly multi-line within first pair)
    if '`' in s:
        frags = re.findall(r"`([^`]+)`", s, re.S)
        if frags:
            cleaned = [strip_ellipsis_brackets(f.strip()) for f in frags]
            cleaned = [f for f in cleaned if f]
            if cleaned:
                return cleaned, 'inline'
    # Plain text (defensive: trim stray leading backtick/whitespace artifacts)
    cleaned = s.strip().lstrip('`').rstrip('`').strip()
    return ([cleaned], 'text') if cleaned else ([], 'empty')


# Heuristic: fields that are always FULL replacement (single-value entities).
ALWAYS_FULL_FIELDS = {
    'Название (RU)', 'Название (UA)',
    'Название модификации (RU)', 'Название модификации (UA)',
    'META keywords (RU)', 'META keywords (UA)',
    'META description (RU)', 'META description (UA)',
    'HTML title (RU)', 'HTML title (UA)',
    'h1 заголовок (RU)', 'h1 заголовок (UA)',
}


def detect_replace_type(field, stale_frags, was_frags):
    """Return 'FULL' or 'PARTIAL' per content shape."""
    if field in ALWAYS_FULL_FIELDS:
        return 'FULL'
    if not stale_frags:
        return 'FULL'
    if len(stale_frags) >= 2:
        return 'PARTIAL'
    if 'Описание' not in field:
        return 'FULL'
    s = stale_frags[0].strip()
    if not s:
        return 'FULL'
    # Single <li> bullet without surrounding block structure → PARTIAL fragment patch
    if s.startswith('<li>') and not re.search(r'<(p|h[1-6]|ul|div|table)\b', s, re.I):
        return 'PARTIAL'
    # Full template shape: has <p> AND structural element (<ul>|<h2>+) OR long
    has_p = bool(re.search(r'<p[\s>]', s, re.I))
    has_struct = bool(re.search(r'<(ul|h[2-6])\b', s, re.I))
    if (has_p and has_struct) or len(s) >= 250:
        return 'FULL'
    if re.match(r'^<(p|h[1-6]|ul|div|table|ol)\b', s, re.I):
        return 'FULL'
    # Short plain text without HTML → PARTIAL typo fix
    if len(s) < 100 and not re.search(r'<\w', s):
        return 'PARTIAL'
    # Ambiguous tail — safer to treat as PARTIAL (avoid blowing away a full description)
    return 'PARTIAL'


def strip_noise(payload):
    """Strip leading/trailing whitespace + stray single backticks."""
    s = payload.strip()
    while s.startswith('`'):
        s = s[1:].lstrip()
    while s.endswith('`'):
        s = s[:-1].rstrip()
    return s


Edit = namedtuple('Edit', ['art', 'field', 'stale', 'was', 'stale_frags', 'was_frags', 'rtype', 'kind', 'n'])


def parse_chunk(n):
    """Return (edits_list, skipped_list) for chunk n. Edits are Edit namedtuples."""
    path = f"{CHUNK_DIR}/chunk-{n:03d}-diff.md"
    src = open(path, encoding="utf-8").read()

    parts = SKU_SPLIT.split(src)
    edits = []
    skipped = []
    for i in range(1, len(parts), 2):
        header = parts[i]
        body = parts[i + 1] if i + 1 < len(parts) else ""

        m = ART_PAREN.search(header) or ART_BACKTICK.search(header)
        if not m:
            skipped.append((n, header[:80], "no_artikul"))
            continue
        art = m.group(1)

        events = []
        for fm in RE_FIELD_H3.finditer(body):
            label = normalize_field(fm.group(1))
            if label:
                events.append((fm.start(), 'field', label))
        for fm in RE_FIELD_POLE.finditer(body):
            label = normalize_field(fm.group(1))
            if label:
                events.append((fm.start(), 'field', label))
        for hm in RE_WAS_HEAD.finditer(body):
            chunk = extract_payload_chunk(body, hm.end())
            frags, kind = parse_payload(chunk)
            events.append((hm.start(), 'was', (frags, kind)))
        for hm in RE_STALO_HEAD.finditer(body):
            chunk = extract_payload_chunk(body, hm.end())
            frags, kind = parse_payload(chunk)
            events.append((hm.start(), 'stale', (frags, kind)))

        events.sort(key=lambda x: x[0])

        current_field = None
        pending_was = None
        for pos, etype, data in events:
            if etype == 'field':
                current_field = data
                pending_was = None
            elif etype == 'was':
                pending_was = data
            elif etype == 'stale':
                if current_field is None:
                    skipped.append((n, art, "orphan_stalo"))
                    continue
                stale_frags, stale_kind = data
                if stale_kind in ('nochange', 'empty'):
                    pending_was = None
                    continue
                was_frags, was_kind = pending_was if pending_was else ([], 'missing')
                rtype = detect_replace_type(current_field, stale_frags, was_frags)
                stale_joined = stale_frags[0] if len(stale_frags) == 1 else ' / '.join(stale_frags)
                was_joined = was_frags[0] if len(was_frags) == 1 else (' / '.join(was_frags) if was_frags else '')
                edits.append(Edit(
                    art=art, field=current_field,
                    stale=stale_joined, was=was_joined,
                    stale_frags=stale_frags, was_frags=was_frags,
                    rtype=rtype, kind=stale_kind, n=n,
                ))
                pending_was = None

    return edits, skipped


def apply_partial(current, was_frags, stale_frags, art, field):
    """Apply sequential str.replace per fragment pair. Returns (new_value, warnings)."""
    warnings = []
    if current is None or current == '':
        warnings.append(f'art={art} field={field}: base value EMPTY, cannot patch')
        return current or '', warnings
    if not was_frags:
        warnings.append(f'art={art} field={field}: PARTIAL has no Было fragments — cannot patch safely')
        return current, warnings
    if len(was_frags) != len(stale_frags):
        warnings.append(f'art={art} field={field}: was/stale frag count mismatch ({len(was_frags)} vs {len(stale_frags)})')
    new_val = current
    for w, s in zip(was_frags, stale_frags):
        w = strip_noise(w)
        s = strip_noise(s)
        if not w:
            continue
        if w not in new_val:
            warnings.append(f'art={art} field={field}: was NOT FOUND: {w[:80]!r}')
            continue
        new_val = new_val.replace(w, s, 1)
    return new_val, warnings


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--chunks", default=None,
                    help="comma-sep chunk numbers; default = all closed (002,005,006..028)")
    ap.add_argument("--write", action="store_true",
                    help="write horoshop-import xlsx (default dry-run summary)")
    ap.add_argument("--out", default="horoshop-import-2026-05-20.xlsx")
    ap.add_argument("--base", default="horoshop-export 20.05.26.xlsx",
                    help="base xlsx to read current cells from")
    args = ap.parse_args()

    chunks = ([int(x) for x in args.chunks.split(",")] if args.chunks else CLOSED_CHUNKS)
    print(f"Parsing {len(chunks)} chunks: {chunks}\n")

    all_edits = []
    all_skipped = []
    type_counter = defaultdict(int)
    for n in chunks:
        edits, skipped = parse_chunk(n)
        all_edits.extend(edits)
        all_skipped.extend(skipped)
        types = defaultdict(int)
        for e in edits:
            types[e.rtype] += 1
            type_counter[e.rtype] += 1
        print(f"chunk-{n:03d}: {len(edits):4d} edits  ({len(skipped)} skipped)  FULL={types['FULL']:3d} PARTIAL={types['PARTIAL']:3d}")

    print(f"\nTOTAL: {len(all_edits)} edits across {len(chunks)} chunks")
    print(f"  FULL={type_counter['FULL']}  PARTIAL={type_counter['PARTIAL']}")
    print(f"SKIPPED orphans: {len(all_skipped)}")

    # Aggregate per art -> per field -> list of edits (ordered).
    agg = defaultdict(lambda: defaultdict(list))
    field_counter = defaultdict(int)
    for e in all_edits:
        agg[e.art][e.field].append(e)
        field_counter[e.field] += 1

    print(f"\nUnique артикулов с правками: {len(agg)}")
    print(f"Уникальных полей затронуто:   {len(field_counter)}")
    print("\nDistribution by field (top 20):")
    for fld, c in sorted(field_counter.items(), key=lambda x: -x[1])[:20]:
        print(f"  {c:5d}  {fld}")

    if all_skipped[:5]:
        print("\nFirst 5 skipped:")
        for s in all_skipped[:5]:
            print(f"  chunk-{s[0]:03d}  {s[1]}  -> {s[2]}")

    if not args.write:
        print("\nDry-run only. To write XLSX use --write.")
        return

    # ---- WRITE -----------------------------------------------------------------
    import openpyxl

    wb_in = openpyxl.load_workbook(args.base)
    ws_in = wb_in.active
    hdr = [c.value for c in next(ws_in.iter_rows(min_row=1, max_row=1))]
    print(f"\nBase xlsx loaded: {args.base} ({ws_in.max_row - 1} rows, {len(hdr)} cols)")

    FIELD_MAP_ALIASES = {
        "Описание товара (RU)": ["Описание товара (RU)", "Описание (RU)"],
        "Описание товара (UA)": ["Описание товара (UA)", "Описание (UA)"],
        "Название модификации (RU)": ["Название модификации (RU)", "Назв.мод (RU)"],
        "Название модификации (UA)": ["Название модификации (UA)", "Назв.мод (UA)"],
    }

    col_idx = {}
    for fld in field_counter:
        aliases = FIELD_MAP_ALIASES.get(fld, [fld])
        for cand in aliases:
            if cand in hdr:
                col_idx[fld] = hdr.index(cand)
                break
        if fld not in col_idx:
            print(f"  ⚠️  field NOT FOUND in base hdr: {fld!r}")

    art_col = None
    for cand in ["Артикул", "артикул"]:
        if cand in hdr:
            art_col = hdr.index(cand)
            break
    if art_col is None:
        raise SystemExit("No Артикул column in base xlsx — fatal")

    used_cols = ["Артикул"] + [c for c in field_counter if c in col_idx]

    base_rows = {}
    for r in ws_in.iter_rows(min_row=2, values_only=True):
        a = str(r[art_col]) if r[art_col] is not None else ""
        if a:
            base_rows[a] = r

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.append(used_cols)

    written = 0
    missing = []
    all_warnings = []
    rows_with_partial_fail = 0
    partial_apply_count = 0
    full_apply_count = 0
    for art, fld_map in agg.items():
        if art not in base_rows:
            missing.append(art)
            continue
        row_vals = base_rows[art]
        out_row = []
        row_had_partial_fail = False
        for col in used_cols:
            if col == "Артикул":
                out_row.append(art)
                continue
            if col not in fld_map:
                out_row.append(row_vals[col_idx[col]])
                continue
            field_edits = fld_map[col]
            full_edits = [e for e in field_edits if e.rtype == 'FULL']
            if full_edits:
                out_row.append(full_edits[-1].stale)
                full_apply_count += 1
                continue
            current = row_vals[col_idx[col]]
            current_str = '' if current is None else str(current)
            new_val = current_str
            row_partial_failed_here = False
            for e in field_edits:
                new_val, warns = apply_partial(new_val, e.was_frags, e.stale_frags, art, col)
                all_warnings.extend(warns)
                if warns:
                    row_partial_failed_here = True
            if row_partial_failed_here:
                row_had_partial_fail = True
                # Fallback: keep base value to avoid corrupting Horoshop
                # (only if EVERY edit failed; otherwise keep partially-patched value)
                if new_val == current_str:
                    new_val = current_str  # explicit
            partial_apply_count += 1
            out_row.append(new_val)
        if row_had_partial_fail:
            rows_with_partial_fail += 1
        ws_out.append(out_row)
        written += 1

    wb_out.save(args.out)
    print(f"\n  wrote {written} rows → {args.out}")
    print(f"  field-edits applied: FULL={full_apply_count}  PARTIAL={partial_apply_count}")
    if missing:
        print(f"  ⚠️  {len(missing)} артикулов из правок не найдены в базе:")
        for a in missing[:10]:
            print(f"     {a}")
        if len(missing) > 10:
            print(f"     ... +{len(missing) - 10} more")
    if all_warnings:
        print(f"\n  ⚠️  PARTIAL warnings: {len(all_warnings)} (affecting {rows_with_partial_fail} rows)")
        for w in all_warnings[:30]:
            print(f"     {w}")
        if len(all_warnings) > 30:
            print(f"     ... +{len(all_warnings) - 30} more")
    wb_in.close()
    wb_out.close()


if __name__ == "__main__":
    main()
