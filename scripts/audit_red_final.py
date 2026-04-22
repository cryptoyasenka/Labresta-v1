"""Final audit of RED rows in Лабреста Х.xlsx before feed upload.

Reads cell colors directly (not read_only) to identify RED rows, then for each:
  1. Broad HKN-prefix search — ALL matching supplier products (no .first()).
  2. Shows existing matches (status) and suggests action.

Classifies each red row into:
  - READY: SP exists, confirmed/manual match to PP — false red, fine.
  - NEEDS_CLICK: SP has only candidate match → confirm in UI.
  - NEEDS_MANUAL: SP exists but no match row → manual match in UI.
  - NO_SP: no supplier product found even with broad search.
  - AMBIGUOUS: multiple SPs match — list all, human picks.
  - SUSPICIOUS: SP found but name hints it's wrong variant.

Output: audit_red_final.json + audit_red_final.md (human-readable).
"""
import json
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
from sqlalchemy import or_

from app import create_app
from app.extensions import db
from app.models.catalog import PromProduct
from app.models.product_match import ProductMatch
from app.models.supplier_product import SupplierProduct


HKN_TOKEN_RE = re.compile(r"HKN[- ]([A-Z0-9][A-Z0-9()]*(?:[- ][A-Z0-9()]+)*)", re.IGNORECASE)


def extract_hkn_tail(name: str, url: str) -> str | None:
    """Return the full tail after 'HKN-' — preserves color/size tokens."""
    if url:
        m = re.search(r"hkn-([a-z0-9\-]+?)/?\s*$", url, re.IGNORECASE)
        if m:
            return m.group(1).rstrip("/").upper()
    if name:
        m = HKN_TOKEN_RE.search(name.upper())
        if m:
            return m.group(1).replace(" ", "-").upper()
    return None


def first_token(tail: str) -> str:
    return tail.split("-")[0] if tail else tail


def find_all_sps(session, supplier_id: int, tail: str, file_name: str) -> list[SupplierProduct]:
    """Broad search — return EVERY active SP whose article/name contains the
    first token of the tail. No .first(), no ordering tricks."""
    token = first_token(tail)
    if not token:
        return []
    pats = [f"%HKN-{token}%", f"%HKN {token}%", f"%HKN{token}%", f"%{token}%"]
    seen: set[int] = set()
    hits: list[SupplierProduct] = []
    for pat in pats:
        q = session.query(SupplierProduct).filter(
            SupplierProduct.supplier_id == supplier_id,
            or_(
                SupplierProduct.article.ilike(pat),
                SupplierProduct.external_id.ilike(pat),
                SupplierProduct.name.ilike(pat),
            ),
        )
        for sp in q.all():
            if sp.id not in seen:
                seen.add(sp.id)
                hits.append(sp)
    return hits


def extract_colour(name: str) -> str | None:
    """Colour/variant tokens that must match between file and SP."""
    up = name.upper()
    for tok, key in [
        ("BLACK", "BLACK"), ("ЧОРН", "BLACK"), ("ЧЕРН", "BLACK"),
        ("BRONZ", "BRONZE"), ("БРОНЗ", "BRONZE"),
        ("SILVER", "SILVER"), ("СРІБ", "SILVER"), ("СЕРЕБ", "SILVER"),
        ("INOX", "INOX"),
    ]:
        if tok in up:
            return key
    return None


def get_red_rows(workbook_path: str) -> list[dict]:
    """Read workbook with formatting → extract rows with red fill."""
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    red = []
    for ws in wb.worksheets:
        if ws.max_row < 100:
            continue
        for row_num, row in enumerate(ws.iter_rows(), start=1):
            if row_num < 6:
                continue
            if not row or not row[0].value or not isinstance(row[0].value, int):
                continue
            # check if any cell in the row has red-ish fill
            is_red = False
            # This workbook uses FA8072 (salmon) for red; FFFAD9 (beige) for normal.
            RED_FILLS = {"FFFA8072", "FA8072", "FFFF0000", "FF0000", "FFC7CE"}
            for cell in row[:8]:
                f = cell.fill
                if not f or not f.fgColor:
                    continue
                rgb = f.fgColor.rgb
                if rgb and isinstance(rgb, str) and rgb.upper() in RED_FILLS:
                    is_red = True
                    break
            if not is_red:
                continue
            vals = [c.value for c in row[:9]]
            red.append({
                "row": row_num,
                "num": vals[0],
                "np_url": vals[1],
                "name": vals[2] or "",
                "np_price": vals[3],
                "rts_price": vals[4],
                "lab_price": vals[5],
                "lab_url": vals[6],
            })
    return red


def classify_row(row: dict, sps: list[SupplierProduct], session) -> dict:
    tail = extract_hkn_tail(row.get("name") or "", row.get("np_url") or "")
    file_colour = extract_colour(row.get("name") or "")
    result = {
        "row": row["row"],
        "name": row["name"],
        "np_url": row["np_url"],
        "lab_url": row["lab_url"],
        "hkn_tail": tail,
        "file_colour": file_colour,
        "sp_hits": [],
        "verdict": "UNKNOWN",
        "notes": [],
    }
    if not tail:
        result["verdict"] = "NO_ANCHOR"
        return result
    if not sps:
        result["verdict"] = "NO_SP"
        return result

    # filter SPs by colour if file specifies one
    filtered = []
    for sp in sps:
        sp_colour = extract_colour((sp.article or "") + " " + (sp.name or ""))
        sp_info = {
            "sp_id": sp.id,
            "article": sp.article,
            "name": sp.name,
            "colour": sp_colour,
        }
        if file_colour and sp_colour and file_colour != sp_colour:
            sp_info["filtered_out"] = f"colour mismatch ({file_colour} vs {sp_colour})"
            result["sp_hits"].append(sp_info)
            continue
        if file_colour and not sp_colour:
            sp_info["filtered_out"] = f"file is {file_colour}, SP has no colour token"
            result["sp_hits"].append(sp_info)
            continue
        if sp_colour and not file_colour:
            sp_info["filtered_out"] = f"file has no colour, SP is {sp_colour}"
            result["sp_hits"].append(sp_info)
            continue
        filtered.append(sp)
        result["sp_hits"].append(sp_info)

    if not filtered:
        result["verdict"] = "NO_SP"
        result["notes"].append("all sp hits filtered out by colour")
        return result

    if len(filtered) > 1:
        # Tier A: exact first-token match. "HBH2000S" must equal SP's first
        # article/name token after HKN- (strips trailing letters like "STH").
        tok = first_token(tail)
        def sp_first_token(sp):
            art = (sp.article or "").upper().replace("HKN-", "").replace("HKN ", "").strip()
            return re.split(r"[\s\-]", art, 1)[0] if art else ""
        exact = [sp for sp in filtered if sp_first_token(sp) == tok]
        if len(exact) == 1:
            filtered = exact
        else:
            # Tier B: full tail contained. Handles "LPD150 BLACK" filtering
            # out "LPD150S BLACK" if the file tail includes "LPD150-BLACK".
            tail_re = re.sub(r"[-\s]+", r"[-\\s]?", re.escape(tail))
            import re as _re
            tail_pat = _re.compile(tail_re)
            deeper = [sp for sp in filtered if tail_pat.search(((sp.article or "") + " " + (sp.name or "")).upper())]
            if len(deeper) == 1:
                filtered = deeper
            else:
                result["verdict"] = "AMBIGUOUS"
                result["notes"].append(f"{len(filtered)} SPs match after colour filter (tok exact={len(exact)}, tail contained={len(deeper)})")
                return result

    sp = filtered[0]
    matches = session.query(ProductMatch).filter_by(
        supplier_product_id=sp.id
    ).all()
    confirmed = [m for m in matches if m.status in ("confirmed", "manual")]
    candidates = [m for m in matches if m.status == "candidate"]

    chosen = {
        "sp_id": sp.id,
        "article": sp.article,
        "name": sp.name,
        "matches": [{"id": m.id, "status": m.status, "pp_id": m.prom_product_id} for m in matches],
    }
    result["chosen_sp"] = chosen

    if confirmed:
        result["verdict"] = "READY"
    elif candidates:
        result["verdict"] = "NEEDS_CLICK"
        result["notes"].append(f"confirm candidate match#{candidates[0].id}")
    else:
        result["verdict"] = "NEEDS_MANUAL"

    return result


def main():
    print("Reading Лабреста Х.xlsx with cell colours…")
    red = get_red_rows("Лабреста Х.xlsx")
    print(f"Red rows: {len(red)}")

    app = create_app()
    with app.app_context():
        results = []
        for row in red:
            tail = extract_hkn_tail(row.get("name") or "", row.get("np_url") or "")
            sps = find_all_sps(db.session, 2, tail, row.get("name") or "") if tail else []
            results.append(classify_row(row, sps, db.session))

    # summary
    from collections import Counter
    counts = Counter(r["verdict"] for r in results)
    print("\n=== SUMMARY ===")
    for v, c in counts.most_common():
        print(f"  {v}: {c}")

    with open("audit_red_final.json", "w", encoding="utf-8") as f:
        json.dump({"counts": dict(counts), "rows": results}, f, ensure_ascii=False, indent=2, default=str)
    print(f"\nDetails in audit_red_final.json")

    # Human-readable report
    lines = ["# Red-rows audit — готовность к загрузке фида\n"]
    lines.append(f"**Всего красных строк:** {len(red)}\n")
    for v, c in counts.most_common():
        lines.append(f"- `{v}`: {c}")
    lines.append("")
    for verdict in ["NO_SP", "NEEDS_MANUAL", "NEEDS_CLICK", "AMBIGUOUS", "NO_ANCHOR", "READY"]:
        rows_v = [r for r in results if r["verdict"] == verdict]
        if not rows_v:
            continue
        lines.append(f"\n## {verdict} ({len(rows_v)})\n")
        for r in rows_v:
            lines.append(f"- **{r['name']}** (row {r['row']}, tail={r['hkn_tail']}, colour={r['file_colour']})")
            if r.get("chosen_sp"):
                cs = r["chosen_sp"]
                lines.append(f"    - sp#{cs['sp_id']} {cs['article']} — {cs['name']}")
                for m in cs["matches"]:
                    lines.append(f"      - match#{m['id']} status={m['status']} pp_id={m['pp_id']}")
            if r.get("notes"):
                for n in r["notes"]:
                    lines.append(f"    - _note: {n}_")
            if verdict in ("AMBIGUOUS", "NO_SP") and r.get("sp_hits"):
                lines.append(f"    - sp_hits ({len(r['sp_hits'])}):")
                for h in r["sp_hits"][:8]:
                    if "filtered_out" in h:
                        lines.append(f"      - sp#{h['sp_id']} {h['article']} — FILTERED: {h['filtered_out']}")
                    else:
                        lines.append(f"      - sp#{h['sp_id']} {h['article']} — {h['name']}")
    Path("audit_red_final.md").write_text("\n".join(lines), encoding="utf-8")
    print(f"Report in audit_red_final.md")


if __name__ == "__main__":
    main()
