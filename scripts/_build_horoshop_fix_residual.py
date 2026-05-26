"""Build residual-bug fix xlsx files after Track #1.

Three outputs:
1. horoshop-fix-mod-name-2026-05-21-b.xlsx — [Артикул, Название модификации (RU)] x 1
   (1766968161: take from imp's own Название (RU) — already clean)
2. horoshop-fix-desc-ru-2026-05-21.xlsx — [Артикул, Описание товара (RU)] x 2
   (restore from pre-incident exp 20.05.26 16:08)
3. horoshop-fix-desc-ua-2026-05-21.xlsx — [Артикул, Описание товара (UA)] x 1
   (restore from pre-incident exp 20.05.26 16:08)

NEVER touches source xlsx. Empty cells NEVER allowed. Author = manual import
by Yana, no Co-Authored-By traces.
"""
import os, sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook, Workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
IMP  = os.path.join(ROOT, 'horoshop-import-2026-05-20.xlsx')
EXP  = os.path.join(ROOT, 'horoshop-export 20.05.26.xlsx')

# Targets (extracted from _regressions.json after manual review)
MOD_NAME_FIX = ['1766968161']                # take from imp's Название (RU)
DESC_RU_FIX  = ['2110282234', '1582804831']  # restore from exp
DESC_UA_FIX  = ['2062006550']                # restore from exp

def load(path):
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    hdr = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    idx = {h:i for i,h in enumerate(hdr)}
    rows = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        a = r[idx['Артикул']]
        if a is None: continue
        a = str(a).strip()
        if not a: continue
        rows[a] = {h: (r[i] or '') for h,i in idx.items()}
    return rows

imp = load(IMP)
exp = load(EXP)

def write_xlsx(fname, header_cols, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = fname.replace('.xlsx','')[:30]
    ws.append(header_cols)
    for r in rows:
        # Strict: NO empty values allowed
        if any(v is None or str(v).strip() == '' for v in r):
            raise ValueError(f'Empty cell forbidden in {fname}: {r}')
        ws.append(list(r))
    p = os.path.join(ROOT, fname)
    wb.save(p)
    print(f'Wrote: {p}  ({len(rows)} rows)')
    # Preview
    for r in rows:
        print(f'  {r[0]}: {str(r[1])[:80]!r}{"..." if len(str(r[1]))>80 else ""}')

# --- 1. mod-name-b ---
mod_rows = []
for art in MOD_NAME_FIX:
    fix_val = imp[art].get('Название (RU)', '').strip()
    if not fix_val:
        raise ValueError(f'imp.Название (RU) empty for {art}')
    # Sanity: not UA-leak
    import re
    if re.search(r'[іїєґІЇЄҐ]', fix_val):
        raise ValueError(f'imp.Название (RU) still has UA leak for {art}: {fix_val!r}')
    mod_rows.append((art, fix_val))
write_xlsx('horoshop-fix-mod-name-2026-05-21-b.xlsx',
           ['Артикул', 'Название модификации (RU)'], mod_rows)

print()

# --- 2. desc-ru ---
ru_rows = []
for art in DESC_RU_FIX:
    v = exp[art].get('Описание товара (RU)', '').strip()
    if not v:
        raise ValueError(f'exp.Описание товара (RU) empty for {art}')
    ru_rows.append((art, v))
write_xlsx('horoshop-fix-desc-ru-2026-05-21.xlsx',
           ['Артикул', 'Описание товара (RU)'], ru_rows)

print()

# --- 3. desc-ua ---
ua_rows = []
for art in DESC_UA_FIX:
    v = exp[art].get('Описание товара (UA)', '').strip()
    if not v:
        raise ValueError(f'exp.Описание товара (UA) empty for {art}')
    ua_rows.append((art, v))
write_xlsx('horoshop-fix-desc-ua-2026-05-21.xlsx',
           ['Артикул', 'Описание товара (UA)'], ua_rows)
