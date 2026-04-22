"""Dump Лабреста Х.xlsx to JSON for analysis (utf-8 safe)."""
import json
import openpyxl


def main():
    wb = openpyxl.load_workbook("Лабреста Х.xlsx", read_only=True, data_only=True)
    out = {}
    for ws in wb.worksheets:
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        out[ws.title] = {
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "rows": rows,
        }
    with open("labresta_x_dump.json", "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2, default=str)
    print(f"Wrote {len(out)} sheets to labresta_x_dump.json")
    for name, data in out.items():
        print(f"  {name}: {data['max_row']} rows x {data['max_col']} cols")


if __name__ == "__main__":
    main()
