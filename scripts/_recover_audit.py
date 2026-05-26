"""Recovery audit: classify broken rows in horoshop-import-2026-05-20.xlsx
and probe horoshop-export 13.05.26.xlsx (pre-incident clean snapshot) as
canonical source for RU+UA descriptions and META keywords.

Output:
  .planning/_recovery_audit.txt  — classification report + coverage
  .planning/_recovery_audit_rows.json — per-row decisions
"""
import os, sys, json, re
from openpyxl import load_workbook
from collections import Counter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PLAN = os.path.join(ROOT, '.planning')
IMP  = os.path.join(ROOT, 'horoshop-import-2026-05-20.xlsx')
EXP  = os.path.join(ROOT, 'horoshop-export 20.05.26.xlsx')  # 6h pre-incident snapshot

LOG_PATH  = os.path.join(PLAN, '_recovery_audit.txt')
ROWS_PATH = os.path.join(PLAN, '_recovery_audit_rows.json')

LOG = open(LOG_PATH, 'w', encoding='utf-8')
def p(*args):
    s = ' '.join(str(a) for a in args)
    LOG.write(s + '\n')
LOG.flush()

# --- 1. Load broken import ---
p('=== STEP 1: read broken import xlsx ===')
wb = load_workbook(IMP, read_only=True)
ws = wb.active
hdr = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
p(f'IMP headers ({len(hdr)}): {hdr}')
imp_idx = {h:i for i,h in enumerate(hdr)}

# Required cols
art_i        = imp_idx['Артикул']
desc_ru_i    = imp_idx['Описание товара (RU)']
desc_ua_i    = imp_idx['Описание товара (UA)']
meta_kw_ru_i = imp_idx.get('META keywords (RU)')
meta_kw_ua_i = imp_idx.get('META keywords (UA)')
name_ru_i    = imp_idx.get('Название (RU)')
name_ua_i    = imp_idx.get('Название (UA)')

imp_rows = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    art = row[art_i]
    if art is None: continue
    art = str(art).strip()
    if not art: continue
    imp_rows[art] = {
        'desc_ru': (row[desc_ru_i] or '').strip(),
        'desc_ua': (row[desc_ua_i] or '').strip(),
        'meta_kw_ru': (row[meta_kw_ru_i] or '').strip() if meta_kw_ru_i is not None else '',
        'meta_kw_ua': (row[meta_kw_ua_i] or '').strip() if meta_kw_ua_i is not None else '',
        'name_ru': (row[name_ru_i] or '').strip() if name_ru_i is not None else '',
        'name_ua': (row[name_ua_i] or '').strip() if name_ua_i is not None else '',
    }
p(f'IMP rows with non-empty Артикул: {len(imp_rows)}')

# --- 2. Load pre-incident export (clean snapshot) ---
p()
p('=== STEP 2: read pre-incident export (13.05.26) ===')
wb_e = load_workbook(EXP, read_only=True)
ws_e = wb_e.active
hdr_e = list(next(ws_e.iter_rows(min_row=1, max_row=1, values_only=True)))
p(f'EXP headers ({len(hdr_e)}): {hdr_e[:30]}')
exp_idx = {h:i for i,h in enumerate(hdr_e)}
# Match canonical names
e_art_i        = exp_idx.get('Артикул')
e_desc_ua_i    = exp_idx.get('Описание товара (UA)')
e_desc_ru_i    = exp_idx.get('Описание товара (RU)')
e_meta_kw_ru_i = exp_idx.get('META keywords (RU)')
e_meta_kw_ua_i = exp_idx.get('META keywords (UA)')
e_name_ua_i    = exp_idx.get('Название (UA)')
e_name_ru_i    = exp_idx.get('Название (RU)')
p(f'EXP col indexes: art={e_art_i} desc_ua={e_desc_ua_i} desc_ru={e_desc_ru_i} meta_kw_ru={e_meta_kw_ru_i} meta_kw_ua={e_meta_kw_ua_i}')

exp_rows = {}
if e_art_i is not None:
    for row in ws_e.iter_rows(min_row=2, values_only=True):
        art = row[e_art_i]
        if art is None: continue
        art = str(art).strip()
        if not art: continue
        exp_rows[art] = {
            'desc_ua': (row[e_desc_ua_i] or '').strip() if e_desc_ua_i is not None else '',
            'desc_ru': (row[e_desc_ru_i] or '').strip() if e_desc_ru_i is not None else '',
            'meta_kw_ru': (row[e_meta_kw_ru_i] or '').strip() if e_meta_kw_ru_i is not None else '',
            'meta_kw_ua': (row[e_meta_kw_ua_i] or '').strip() if e_meta_kw_ua_i is not None else '',
            'name_ua': (row[e_name_ua_i] or '').strip() if e_name_ua_i is not None else '',
            'name_ru': (row[e_name_ru_i] or '').strip() if e_name_ru_i is not None else '',
        }
p(f'EXP rows with non-empty Артикул: {len(exp_rows)}')

# --- 3. Classify broken rows ---
p()
p('=== STEP 3: classify broken rows in IMP ===')

UA_ONLY_RE = re.compile(r'[іїєґІЇЄҐ]')
def has_ua_letters(s):
    return bool(UA_ONLY_RE.search(s))

def desc_is_broken(s):
    """Return (broken: bool, reason: str)."""
    if not s: return False, ''
    if s.startswith('<li>'): return True, 'starts <li>'
    # Leading backtick fragments
    if s.startswith('`'): return True, 'leading backtick fragment'
    if s.startswith(' `'): return True, 'leading backtick fragment'
    # short no block tags
    if len(s) < 80 and not re.search(r'<(p|h[1-6]|ul|div)\b', s, re.I):
        return True, 'short, no block tags'
    # not starting with block tag
    if not re.match(r'^\s*<(p|h[1-6]|ul|div|table|ol|hr|br)\b', s, re.I):
        return True, 'no opening block tag'
    return False, ''

def desc_has_ua_leak_in_ru(s):
    if not s: return False
    return has_ua_letters(s)

def kw_has_ua_leak_in_ru(s):
    if not s: return False
    return has_ua_letters(s)

decisions = {}
counters = Counter()
for art, r in imp_rows.items():
    d = {'fixes': []}

    # RU description
    bru, why_ru = desc_is_broken(r['desc_ru'])
    ualeak_ru   = desc_has_ua_leak_in_ru(r['desc_ru'])
    if bru or ualeak_ru:
        d['fixes'].append({
            'field': 'Описание товара (RU)',
            'reasons': [w for w in ([why_ru] if bru else []) + (['ua_leak_in_ru'] if ualeak_ru else []) if w],
        })
        counters['desc_ru_broken'] += 1

    # UA description
    bua, why_ua = desc_is_broken(r['desc_ua'])
    if bua:
        d['fixes'].append({
            'field': 'Описание товара (UA)',
            'reasons': [why_ua],
        })
        counters['desc_ua_broken'] += 1

    # RU meta kw
    if kw_has_ua_leak_in_ru(r['meta_kw_ru']):
        d['fixes'].append({'field': 'META keywords (RU)', 'reasons': ['ua_leak_in_ru']})
        counters['meta_ru_ua_leak'] += 1
    # UA meta kw (just check empty? skip — UA kw can have UA letters legitimately)

    # Coverage in pre-incident export
    if d['fixes']:
        e = exp_rows.get(art)
        d['exp_has'] = bool(e)
        if e:
            d['exp_desc_ru_len'] = len(e['desc_ru'])
            d['exp_desc_ua_len'] = len(e['desc_ua'])
            d['exp_meta_kw_ru_len'] = len(e['meta_kw_ru'])
        decisions[art] = d

p(f'Broken rows count by category:')
for k,v in counters.most_common():
    p(f'  {k}: {v}')

p()
covered = sum(1 for v in decisions.values() if v.get('exp_has'))
p(f'Total broken artikuls: {len(decisions)}')
p(f'Coverage in pre-incident export (13.05.26): {covered}/{len(decisions)} ({100*covered/max(1,len(decisions)):.1f}%)')

# Missing in export
missing = [a for a,v in decisions.items() if not v.get('exp_has')]
p(f'NOT in pre-incident export: {len(missing)}')
if missing:
    p('  first 30:', missing[:30])

# Where exp has empty desc_ru for our fix
empty_in_exp_for_desc_ru = []
for art, v in decisions.items():
    if not v.get('exp_has'): continue
    for fx in v['fixes']:
        if fx['field'] == 'Описание товара (RU)' and v.get('exp_desc_ru_len',0) == 0:
            empty_in_exp_for_desc_ru.append(art)
            break
p(f'Need RU desc fix but pre-incident export has EMPTY desc_ru: {len(empty_in_exp_for_desc_ru)}')
if empty_in_exp_for_desc_ru[:30]:
    p('  first 30:', empty_in_exp_for_desc_ru[:30])

empty_in_exp_for_desc_ua = []
for art, v in decisions.items():
    if not v.get('exp_has'): continue
    for fx in v['fixes']:
        if fx['field'] == 'Описание товара (UA)' and v.get('exp_desc_ua_len',0) == 0:
            empty_in_exp_for_desc_ua.append(art)
            break
p(f'Need UA desc fix but pre-incident export has EMPTY desc_ua: {len(empty_in_exp_for_desc_ua)}')

# Save
with open(ROWS_PATH, 'w', encoding='utf-8') as f:
    json.dump(decisions, f, ensure_ascii=False, indent=2)
p()
p(f'Wrote: {LOG_PATH}')
p(f'Wrote: {ROWS_PATH}')

LOG.close()
print(f'OK. Report: {LOG_PATH}')
print(f'JSON:    {ROWS_PATH}')
