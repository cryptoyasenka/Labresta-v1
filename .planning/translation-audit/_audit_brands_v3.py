"""W2 brand audit v3 — добавлено UA-cyrillic (хенді, сірман, бартшер з 'і') + c35 проверка симметрично c36.

v2 errors:
- 'хенді' не искался → UA description (c35) с украинской 'і' не покрыт
- сирман (RU) ≠ сірман (UA) — был только RU
"""
import openpyxl, os, re
from collections import defaultdict

CHUNKS_DIR = r'C:\Projects\labresta-sync-w2\.planning\translation-audit\chunks'
BUNDLE = r'C:\Projects\labresta-sync-w2\.planning\translation-audit\w2-horoshop-import-055-085.xlsx'
OUT_MD = r'C:\Projects\labresta-sync-w2\.planning\translation-audit\W2-POST-HANDOFF-AUDIT-BRANDS-v3.md'

# Brand patterns: BOTH russian (cyr-и) AND ukrainian (cyr-і) variants
BRAND_PATTERNS = {
    'Hendi':       [r'\bхенди\b', r'\bхенді\b'],
    'Tefcold':     [r'\bтефколд\w*', r'\bтэфколд\w*', r'\bтефкольд\w*'],
    'Bartscher':   [r'\bбарчер\w*', r'\bбартшер\w*', r'\bбартчер\w*', r'\bбарщер\w*'],
    'Sirman':      [r'\bсирман\w*', r'\bцирман\w*', r'\bсірман\w*'],
    'Asber':       [r'\bасбер\w*'],
    'Ceado':       [r'\bчеадо\b', r'\bцеадо\b', r'\bсеадо\b'],
    'Robot Coupe': [r'\bробот[\s-]куп\b', r'\bробокуп\b'],
    'Gi Metal':    [r'\bджи\s+метал\b', r'\bги\s+метал\b', r'\bджі\s+метал\b'],
    'REEDNEE':     [r'\bридни[еий]?\b', r'\bрідні\b', r'\bріднє\b'],
    'Tecnodom':    [r'\bтекнодом\w*'],
    'Forcar':      [r'\bфоркар\w*'],
    'GoodFood':    [r'\bгудфуд\w*', r'\bгудфут\w*', r'\bгуд\s+фуд\w*', r'\bгудфут\w*'],
    'Convito':     [r'\bконвито\b', r'\bконвіто\b'],
    'Krupps':      [r'\bкрупс\b', r'\bкрупп\b'],
    'Italgi':      [r'\bитальджи\b', r'\bіталджі\b', r'\bіталджи\b'],
    'Pavoni':      [r'\bпавони\b', r'\bпавоні\b'],
    'Cunill':      [r'\bкуніл\w*', r'\bкуниль\w*'],
    'Hurakan':     [r'\bхуракан\w*'],
    'Apach':       [r'\bапач\b'],
    'Fagor':       [r'\bфагор\w*'],
    'Astoria':     [r'\bастория\b', r'\bасторія\b'],
    'Arris':       [r'\bаррис\b', r'\bарріс\b'],
    'Maxima':      [r'\bмаксима\b'],
    'Tatra':       [r'\bтатра\b'],
    'Cold':        [r'\bколд\b'],
    'Project Systems': [r'\bпроект\s+системс\b', r'\bпроджект\s+системс\b', r'\bпроєкт\s+системс\b'],
    'Cookmax':     [r'\bкукмакс\w*', r'\bкукмаст\w*'],
    'Saro':        [r'\bсаро\b'],
    'Hupfer':      [r'\bхупфер\w*'],
    'Cancan':      [r'\bканкан\b'],
    'Brema':       [r'\bбрема\b'],
    'Apex':        [r'\bапекс\w*'],
    'Pizzagroup':  [r'\bпиццагруп\w*', r'\bпіцагруп\w*'],
    'Scan':        [r'\bскан\b(?!\w)'],
}

TRANS_COLS = [4, 5, 6, 7, 24, 25, 35, 36]
COL_NAMES = {4:'c04 name_mod_UA', 5:'c05 name_mod_RU', 6:'c06 name_UA', 7:'c07 name_RU',
             24:'c24 META_kw_UA', 25:'c25 META_kw_RU', 35:'c35 desc_UA', 36:'c36 desc_RU'}

compiled = {brand: [re.compile(p, re.IGNORECASE) for p in pats]
            for brand, pats in BRAND_PATTERNS.items()}

bundle_arts = set()
wb = openpyxl.load_workbook(BUNDLE, data_only=True)
ws = wb.active
for r in range(2, ws.max_row + 1):
    a = ws.cell(r, 1).value
    if a is not None:
        bundle_arts.add(str(a).strip())
wb.close()

results = defaultdict(lambda: defaultdict(list))

for cf in sorted(os.listdir(CHUNKS_DIR)):
    if not (cf.startswith('chunk-0') and cf.endswith('-fixed.xlsx')):
        continue
    n = int(cf.split('-')[1])
    if not (55 <= n <= 85):
        continue
    fpath = os.path.join(CHUNKS_DIR, cf)
    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb.active
    for r in range(2, ws.max_row + 1):
        art = ws.cell(r, 1).value
        if art is None:
            continue
        art_s = str(art).strip()
        for col in TRANS_COLS:
            v = ws.cell(r, col).value
            if v is None:
                continue
            text = str(v)
            for brand, pats in compiled.items():
                for pat in pats:
                    m = pat.search(text)
                    if m:
                        idx = m.start()
                        snip = text[max(0,idx-60):idx+80].replace('\n',' ').replace('\r',' ')
                        results[brand][col].append((cf, r, art_s, snip, m.group()))
                        break
    wb.close()

lines = []
lines.append('# W2 post-handoff audit v3 — RU + UA cyrillic brand audit (всё 8 trans cols, 31 chunks)')
lines.append('')
lines.append(f'**Дата:** 2026-05-21')
lines.append(f'**Триггер:** Yana показала что UA остался cyrillic «хенді» после patch v1 (RU только).')
lines.append(f'**v3 fix:** добавлены UA-variants (хенді с і, сірман, итд) + проверка c35 симметрично с c36.')
lines.append('')
lines.append('## Summary by brand × column (только non-zero)')
lines.append('')
lines.append('| Brand | c4 name_mod UA | c5 name_mod RU | c6 name UA | c7 name RU | c24 kw UA | c25 kw RU | c35 desc UA | c36 desc RU | LIVE SKU |')
lines.append('|---|---|---|---|---|---|---|---|---|---|')

unique_sku_c35 = set()
unique_sku_c36 = set()
unique_sku_names = set()
brands_total = 0

for brand in sorted(results.keys()):
    cells_by_col = results[brand]
    if not cells_by_col:
        continue
    counts = {c: len(cells_by_col.get(c, [])) for c in TRANS_COLS}
    arts_live = set()
    for col, hits in cells_by_col.items():
        for h in hits:
            if h[2] in bundle_arts:
                arts_live.add(h[2])
            if col == 35:
                unique_sku_c35.add(h[2])
            elif col == 36:
                unique_sku_c36.add(h[2])
            elif col in (4, 5, 6, 7):
                unique_sku_names.add(h[2])
    if sum(counts.values()) == 0:
        continue
    brands_total += 1
    row = f'| **{brand}** |'
    for c in TRANS_COLS:
        n = counts[c]
        row += f' {n if n>0 else "·"} |'
    row += f' {len(arts_live)} |'
    lines.append(row)

lines.append('')
lines.append(f'**Уникальных SKU c хитом в c35 (UA desc): {len(unique_sku_c35)}** — НУЖЕН UA fix (живут на сайте: {len(unique_sku_c35 & bundle_arts)})')
lines.append(f'**Уникальных SKU c хитом в c36 (RU desc): {len(unique_sku_c36)}** — RU fix частично применён (Hendi 4 patched в v1)')
lines.append(f'**Уникальных SKU c хитом в c4/c5/c6/c7 (names): {len(unique_sku_names)}**')
lines.append('')

# c35 UA desc detail
lines.append('## c35 (Описание товара UA) — главная зона UA для SEO')
lines.append('')
for brand in sorted(results.keys()):
    c35_hits = results[brand].get(35, [])
    if not c35_hits:
        continue
    arts = set(h[2] for h in c35_hits)
    arts_live = arts & bundle_arts
    by_chunk = defaultdict(int)
    for h in c35_hits:
        by_chunk[h[0]] += 1
    lines.append(f'### {brand} — {len(c35_hits)} ячеек / {len(arts)} unique SKU / {len(arts_live)} live')
    lines.append('')
    lines.append('**По чанкам:** ' + ', '.join(f'{c.replace("-fixed.xlsx","").replace("chunk-","")}={n}' for c, n in sorted(by_chunk.items())))
    lines.append('')
    lines.append('**Sample 5:**')
    lines.append('')
    for h in c35_hits[:5]:
        cf, r, art, snip, match = h
        in_b = '✓ LIVE' if art in bundle_arts else '✗ not-in-bundle'
        lines.append(f'- `{cf.replace("-fixed.xlsx","")}` r{r} ART={art} ({in_b}) match=`{match}` — `...{snip}...`')
    lines.append('')

# c36 RU desc (post-v1-patch state)
lines.append('## c36 (Описание товара RU) — post-v1-patch state')
lines.append('')
for brand in sorted(results.keys()):
    c36_hits = results[brand].get(36, [])
    if not c36_hits:
        continue
    arts = set(h[2] for h in c36_hits)
    arts_live = arts & bundle_arts
    lines.append(f'### {brand} — {len(c36_hits)} ячеек / {len(arts)} unique SKU / {len(arts_live)} live')
    lines.append('')
    for h in c36_hits[:5]:
        cf, r, art, snip, match = h
        in_b = '✓ LIVE' if art in bundle_arts else '✗ not-in-bundle'
        lines.append(f'- `{cf.replace("-fixed.xlsx","")}` r{r} ART={art} ({in_b}) match=`{match}` — `...{snip}...`')
    lines.append('')

# Names check (c4-c7)
lines.append('## c4-c7 (Названия UA + RU) — должно быть 0 для bundle live')
lines.append('')
for brand in sorted(results.keys()):
    name_cols_hits = []
    for col in [4, 5, 6, 7]:
        name_cols_hits.extend([(col, h) for h in results[brand].get(col, [])])
    if not name_cols_hits:
        continue
    lines.append(f'### {brand} ({len(name_cols_hits)} cells)')
    for col, h in name_cols_hits[:8]:
        cf, r, art, snip, match = h
        in_b = '✓ LIVE' if art in bundle_arts else '✗ not-in-bundle'
        lines.append(f'- {COL_NAMES[col]}: `{cf.replace("-fixed.xlsx","")}` r{r} ART={art} ({in_b}) `{match}` — `{snip}`')
    lines.append('')

with open(OUT_MD, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

print(f'[OK] v3 audit saved: {OUT_MD}')
print(f'Brands with hits: {brands_total}')
print(f'Unique SKU c35 (UA desc): {len(unique_sku_c35)} ({len(unique_sku_c35 & bundle_arts)} LIVE)')
print(f'Unique SKU c36 (RU desc): {len(unique_sku_c36)} ({len(unique_sku_c36 & bundle_arts)} LIVE)')
print(f'Unique SKU names (c4-c7): {len(unique_sku_names)}')
