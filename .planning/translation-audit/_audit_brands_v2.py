"""W2 brand audit v2 — fixes Maxima/Tatra/Astoria false positives via word-boundary regex.

v1 errors:
- 'максима' matched inside 'максимальный/максимальная' → 266 FP in c36
- 'татра' may match in placenames
- 'астория' name may match
"""
import openpyxl, os, re
from collections import defaultdict

CHUNKS_DIR = r'C:\Projects\labresta-sync-w2\.planning\translation-audit\chunks'
BUNDLE = r'C:\Projects\labresta-sync-w2\.planning\translation-audit\w2-horoshop-import-055-085.xlsx'
OUT_MD = r'C:\Projects\labresta-sync-w2\.planning\translation-audit\W2-POST-HANDOFF-AUDIT-BRANDS.md'

# Each brand: canonical latin → list of (regex, label) where regex uses word boundaries to avoid false positives.
# UNICODE flag is implicit in Python 3 re for str patterns. Use \b for proper word boundary.
BRAND_PATTERNS = {
    'Hendi':       [r'\bхенди\b', r'\bхенді\b'],
    'Tefcold':     [r'\bтефколд\w*', r'\bтэфколд\w*'],
    'Bartscher':   [r'\bбарчер\w*', r'\bбартшер\w*', r'\bбартчер\w*', r'\bбарщер\w*'],
    'Sirman':      [r'\bсирман\w*', r'\bцирман\w*', r'\bсірман\w*'],
    'Asber':       [r'\bасбер\w*'],
    'Ceado':       [r'\bчеадо\b', r'\bцеадо\b', r'\bсеадо\b'],
    'Robot Coupe': [r'\bробот[\s-]куп\b', r'\bробокуп\b'],
    'Gi Metal':    [r'\bджи\s+метал\b', r'\bги\s+метал\b'],
    'REEDNEE':     [r'\bридни[еий]?\b', r'\bрідні\b'],
    'Tecnodom':    [r'\bтекнодом\w*'],
    'Forcar':      [r'\bфоркар\w*'],
    'GoodFood':    [r'\bгудфуд\w*', r'\bгудфут\w*', r'\bгуд\s+фуд\w*'],
    'Convito':     [r'\bконвито\b', r'\bконвіто\b'],
    'Krupps':      [r'\bкрупс\b', r'\bкрупп\b'],
    'Italgi':      [r'\bитальджи\b', r'\bіталджі\b'],
    'Pavoni':      [r'\bпавони\b', r'\bпавоні\b'],
    'Cunill':      [r'\bкуніл\w*', r'\bкуниль\w*'],
    'Hurakan':     [r'\bхуракан\w*'],
    'Apach':       [r'\bапач\b'],
    'Fagor':       [r'\bфагор\w*'],
    'Astoria':     [r'\bастория\b(?!льн)', r'\bасторія\b'],  # астория ≠ no exclusion needed but careful
    'Arris':       [r'\bаррис\b', r'\bарріс\b'],
    'Maxima':      [r'\bмаксима\b'],  # word boundary excludes 'максимальный'
    'Tatra':       [r'\bтатра\b'],
    'Cold':        [r'\bколд\b'],
    'Project Systems': [r'\bпроект\s+системс\b', r'\bпроджект\s+системс\b'],
    'Cookmax':     [r'\bкукмакс\w*', r'\bкукмаст\w*'],
    'Saro':        [r'\bсаро\b'],
    'Hupfer':      [r'\bхупфер\w*'],
    'Cancan':      [r'\bканкан\b'],   # word boundary excludes 'канкановский'
    'Brema':       [r'\bбрема\b'],
    'Apex':        [r'\bапекс\w*'],
    'Pizzagroup':  [r'\bпиццагруп\w*', r'\bпіцагруп\w*'],
    'Tecnodom':    [r'\bтекнодом\w*'],
    'Scan':        [r'\bскан\b(?!\w)'],  # тоже бренд Scan? редкий, попробуем
}

TRANS_COLS = [4, 5, 6, 7, 24, 25, 35, 36]
COL_NAMES = {4:'c04 name_mod_UA', 5:'c05 name_mod_RU', 6:'c06 name_UA', 7:'c07 name_RU',
             24:'c24 META_kw_UA', 25:'c25 META_kw_RU', 35:'c35 desc_UA', 36:'c36 desc_RU'}

# Compile patterns
compiled = {brand: [re.compile(p, re.IGNORECASE) for p in pats]
            for brand, pats in BRAND_PATTERNS.items()}

# Bundle ARTs
bundle_arts = set()
wb = openpyxl.load_workbook(BUNDLE, data_only=True)
ws = wb.active
for r in range(2, ws.max_row + 1):
    a = ws.cell(r, 1).value
    if a is not None:
        bundle_arts.add(str(a).strip())
wb.close()

# Scan
results = defaultdict(lambda: defaultdict(list))

for cf in sorted(os.listdir(CHUNKS_DIR)):
    if not (cf.startswith('chunk-0') and cf.endswith('-fixed.xlsx')):
        continue
    n = int(cf.split('-')[1])
    if not (55 <= n <= 85):
        continue
    fpath = os.path.join(CHUNKS_DIR, cf)
    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb.active
    for r in range(2, ws.max_row + 1):
        art = ws.cell(r, 1).value
        if art is None:
            continue
        art_s = str(art).strip()
        for col in TRANS_COLS:
            v = ws.cell(r, col).value
            if v is None:
                continue
            text = str(v)
            for brand, pats in compiled.items():
                for pat in pats:
                    m = pat.search(text)
                    if m:
                        idx = m.start()
                        snip = text[max(0,idx-60):idx+80].replace('\n',' ').replace('\r',' ')
                        results[brand][col].append((cf, r, art_s, snip, m.group()))
                        break
    wb.close()

# Build report
lines = []
lines.append('# W2 post-handoff audit v2 — slavified brand names (cyrillic) in fixed.xlsx 055-085')
lines.append('')
lines.append(f'**Дата:** 2026-05-21')
lines.append(f'**Триггер:** Yana заметила «хенди» (кириллица) на live SKU `Тендерайзер Hendi 843468`.')
lines.append(f'**Скоп:** все 31 chunk W2 (055-085, 2168 SKU), 8 trans cols.')
lines.append(f'**v2 fix:** word-boundary regex (\\b) исключает FP типа `максима` в `максимальный`.')
lines.append('')
lines.append('## Summary')
lines.append('')
lines.append('| Brand (canon) | Cells found | c36 desc RU | c35 desc UA | c25 kw RU | c5/c7 names RU | SKU live (in bundle) |')
lines.append('|---|---|---|---|---|---|---|')

summary_total_cells = 0
summary_total_in_bundle = 0
brands_with_hits = []

for brand in sorted(results.keys()):
    cells_by_col = results[brand]
    total = sum(len(v) for v in cells_by_col.values())
    if total == 0:
        continue
    in_c36 = len(cells_by_col.get(36, []))
    in_c35 = len(cells_by_col.get(35, []))
    in_c25 = len(cells_by_col.get(25, []))
    in_names_ru = len(cells_by_col.get(5, [])) + len(cells_by_col.get(7, []))
    arts_in_bundle = set()
    for col, hits in cells_by_col.items():
        for h in hits:
            if h[2] in bundle_arts:
                arts_in_bundle.add(h[2])
    in_bundle_count = len(arts_in_bundle)
    summary_total_cells += total
    summary_total_in_bundle += in_bundle_count
    brands_with_hits.append(brand)
    lines.append(f'| **{brand}** | {total} | {in_c36} | {in_c35} | {in_c25} | {in_names_ru} | {in_bundle_count} |')

lines.append(f'| **TOTAL** | **{summary_total_cells}** |  |  |  |  | **{summary_total_in_bundle}** |')
lines.append('')

lines.append('## Per-brand detail — c36 (Описание товара RU) ONLY, главный SEO surface')
lines.append('')
lines.append('> c25 (META keywords) кириллица — допустима (Yana ack), не фиксим.')
lines.append('> c35 (UA descriptions) — forward к W1 (W2 UA не правит).')
lines.append('')

c36_total_unique_sku = set()
for brand in brands_with_hits:
    c36_hits = results[brand].get(36, [])
    if not c36_hits:
        continue
    arts = set(h[2] for h in c36_hits)
    arts_in_bundle = arts & bundle_arts
    c36_total_unique_sku |= arts
    lines.append(f'### {brand}  — {len(c36_hits)} ячеек в c36 / {len(arts)} unique SKU / {len(arts_in_bundle)} live в bundle')
    lines.append('')
    by_chunk = defaultdict(int)
    for h in c36_hits:
        by_chunk[h[0]] += 1
    lines.append('**По чанкам:** ' + ', '.join(f'{c.replace("-fixed.xlsx","").replace("chunk-","")}={n}' for c, n in sorted(by_chunk.items())))
    lines.append('')
    lines.append('**Sample 5:**')
    lines.append('')
    for h in c36_hits[:5]:
        cf, r, art, snip, match = h
        in_b = '✓ LIVE' if art in bundle_arts else '✗ not-in-bundle'
        lines.append(f'- `{cf.replace("-fixed.xlsx","")}` r{r} ART={art} ({in_b}) match=`{match}` — `...{snip}...`')
    lines.append('')

lines.append(f'**ИТОГО c36 (descriptions RU) — уникальных SKU затронуто: {len(c36_total_unique_sku)}**')
lines.append(f'**Из них LIVE в bundle сейчас: {len(c36_total_unique_sku & bundle_arts)}**')
lines.append('')

lines.append('## Per-brand — c5/c7 (Названия RU) — критическая зона (имя товара)')
lines.append('')
name_hits_total = 0
for brand in brands_with_hits:
    h5 = results[brand].get(5, [])
    h7 = results[brand].get(7, [])
    if not h5 and not h7:
        continue
    lines.append(f'### {brand} — c5: {len(h5)}, c7: {len(h7)}')
    for h in (h5 + h7)[:5]:
        cf, r, art, snip, match = h
        in_b = '✓ LIVE' if art in bundle_arts else '✗ not-in-bundle'
        lines.append(f'- `{cf.replace("-fixed.xlsx","")}` r{r} ART={art} ({in_b}) match=`{match}` — `{snip}`')
    name_hits_total += len(h5) + len(h7)
    lines.append('')

lines.append(f'**ИТОГО c5/c7: {name_hits_total} ячеек**')
lines.append('')

lines.append('## c25 META keywords (RU) — distribution (NOT to fix, Yana ack)')
lines.append('')
lines.append('| Brand | Cells in c25 |')
lines.append('|---|---|')
for brand in brands_with_hits:
    n = len(results[brand].get(25, []))
    if n > 0:
        lines.append(f'| {brand} | {n} |')
lines.append('')

# Tautology check (c36 with both cyr-form AND latin form)
lines.append('## Tautology check (c36) — одновременно cyr + lat одного бренда в той же ячейке')
lines.append('')
taut_hits = []
for cf in sorted(os.listdir(CHUNKS_DIR)):
    if not (cf.startswith('chunk-0') and cf.endswith('-fixed.xlsx')):
        continue
    n = int(cf.split('-')[1])
    if not (55 <= n <= 85):
        continue
    fpath = os.path.join(CHUNKS_DIR, cf)
    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb.active
    for r in range(2, ws.max_row + 1):
        art = ws.cell(r, 1).value
        if art is None:
            continue
        art_s = str(art).strip()
        v = ws.cell(r, 36).value
        if v is None:
            continue
        text = str(v)
        for brand, pats in compiled.items():
            cyr_match = None
            for pat in pats:
                m = pat.search(text)
                if m:
                    cyr_match = m
                    break
            if cyr_match is None:
                continue
            lat_pat = re.compile(r'\b' + re.escape(brand) + r'\b', re.IGNORECASE)
            if lat_pat.search(text):
                idx = cyr_match.start()
                snip = text[max(0,idx-100):idx+200].replace('\n',' ')
                taut_hits.append((cf, r, art_s, brand, cyr_match.group(), snip))
                break
    wb.close()

lines.append(f'**Найдено {len(taut_hits)} ячеек c36 c cyr+lat одного бренда:**')
lines.append('')
for h in taut_hits[:20]:
    cf, r, art, brand, var, snip = h
    in_b = '✓ LIVE' if art in bundle_arts else '✗ not-in-bundle'
    lines.append(f'- `{cf.replace("-fixed.xlsx","")}` r{r} ART={art} ({in_b}) — `{brand}` + cyr `{var}` — `...{snip}...`')
if len(taut_hits) > 20:
    lines.append(f'- … +{len(taut_hits)-20} ещё')
lines.append('')

lines.append('## Proposed fix plan')
lines.append('')
lines.append('1. **c36 (descriptions RU) — regex replace cyr→latin:**')
lines.append('   - Для каждого бренда: word-boundary case-insensitive replace cyr-variant → canonical latin')
lines.append('   - После replace проверить tautology (Hendi…Hendi одним абзацем) — рерайт лид-предложения')
lines.append('2. **c5/c7 (names RU) — точечный fix:**')
lines.append('   - Если имя модификации/название содержит cyr-бренд — fix вручную с учётом склонения')
lines.append('3. **c25 META keywords (RU) — НЕ ТРОГАТЬ** (Yana ack)')
lines.append('4. **c35 UA descriptions — forward к W1** (W2 UA не правит)')
lines.append('5. **Patch-bundle build:**')
lines.append('   - rows из 9-col fixed.xlsx, в которых c36 или c5/c7 изменились patch-passом')
lines.append('   - + 14 unchanged-but-live SKU (бренд cyr был в Horoshop до W2, мы пропустили) → вытащить из horoshop-export, fix, добавить')
lines.append('   - Schema 9 cols как основной bundle (важно: НЕ оставлять пустые → wipe)')
lines.append('6. **Спот-чек patch-bundle** (5-10 SKU live) → full re-import только этого patch')
lines.append('')
lines.append('## Files')
lines.append('')
lines.append('- Audit script v2: `.planning/translation-audit/_audit_brands_v2.py`')
lines.append('- This report (v2 overwrote v1): `W2-POST-HANDOFF-AUDIT-BRANDS.md`')
lines.append('- Pending: `_w2_apply_brand_fix.py` + `w2-horoshop-import-PATCH-brands.xlsx` (на след. сессии после approval)')

with open(OUT_MD, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

print(f'[OK] v2 audit saved: {OUT_MD}')
print(f'Total brand-hits: {summary_total_cells} cells')
print(f'Brands with hits: {len(brands_with_hits)}')
print(f'c36 unique SKUs affected: {len(c36_total_unique_sku)}')
print(f'c36 unique SKUs LIVE (in bundle): {len(c36_total_unique_sku & bundle_arts)}')
print(f'c5/c7 name hits: {name_hits_total}')
print(f'Tautology cells (cyr+lat same brand): {len(taut_hits)}')
