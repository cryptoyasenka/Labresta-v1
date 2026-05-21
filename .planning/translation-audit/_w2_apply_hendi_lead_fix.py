"""Apply Hendi-lead fix to chunk-055-fixed.xlsx r9-r12 col 36 (Description RU).

Yana approval: 'Соединить без дубля' — лид-абзац объединяется так что Hendi упоминается один раз.
Заодно убираем грамматическую опечатку 'рекомендованная' → 'рекомендована'.

Before:
  <p>Современный размягчитель мяса хенди, особенно полезен для приготовления мяса на гриле.
     Продукция Hendi отличается высоким качеством и универсальностью, представленная модель
     рекомендова{на|нная} для использования на профессиональной кухне, ее можно успешно использовать в домашних условиях.</p>

After:
  <p>Современный размягчитель мяса Hendi отличается высоким качеством и универсальностью,
     особенно полезен для приготовления мяса на гриле. Представленная модель рекомендована
     для использования на профессиональной кухне, ее можно успешно использовать в домашних условиях.</p>
"""
import openpyxl, re

FIXED = r'C:\Projects\labresta-sync-w2\.planning\translation-audit\chunks\chunk-055-fixed.xlsx'
ROWS = [9, 10, 11, 12]
COL = 36

NEW_LEAD = (
    '<p>Современный размягчитель мяса Hendi отличается высоким качеством и универсальностью, '
    'особенно полезен для приготовления мяса на гриле. Представленная модель рекомендована '
    'для использования на профессиональной кухне, ее можно успешно использовать в домашних условиях.</p>'
)

# Regex matching the OLD lead paragraph (any variant of «рекомендована/рекомендованная»)
OLD_LEAD_RE = re.compile(
    r'<p>Современный размягчитель мяса хенди, особенно полезен для приготовления мяса на гриле\. '
    r'Продукция Hendi отличается высоким качеством и универсальностью, представленная модель '
    r'рекомендован(?:на|нная|ная|нна|а|ая)? для использования на профессиональной кухне, '
    r'ее можно успешно использовать в домашних условиях\.</p>'
)

wb = openpyxl.load_workbook(FIXED)
ws = wb.active

changes = 0
for r in ROWS:
    art = ws.cell(r, 1).value
    v = ws.cell(r, COL).value
    if v is None:
        print(f'r{r} ART={art}: c{COL} empty — SKIP')
        continue
    s = str(v)
    new_s, n = OLD_LEAD_RE.subn(NEW_LEAD, s, count=1)
    if n == 0:
        print(f'r{r} ART={art}: regex NO MATCH — SKIP (unexpected)')
        print(f'  first 300 chars: {s[:300]}')
        continue
    ws.cell(r, COL).value = new_s
    changes += 1
    # Verify хенди gone
    assert 'хенди' not in new_s.lower(), f'still has хенди after replace in r{r}'
    # Verify Hendi present (latin) and present once (not дубль)
    h_count = new_s.lower().count('hendi')
    print(f'r{r} ART={art}: PATCHED OK (Hendi count in c36 = {h_count})')

if changes > 0:
    wb.save(FIXED)
    print(f'\n[OK] Saved {FIXED} ({changes} cells patched)')
else:
    print('\n[!!] No changes applied')
wb.close()
