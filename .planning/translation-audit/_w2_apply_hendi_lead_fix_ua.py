"""Apply UA lead-paragraph fix for the same 4 Hendi-tendariser SKUs (chunk-055 r9-r12 col35).
Mirrors the RU fix already applied in v1: 'хенді' → 'Hendi', no duplicate, grammar fix.

Before UA:
  <p>Сучасний пом'якшувач м'яса хенді, особливо корисний для приготування м'яса на грилі.
     Продукція Hendi вирізняється високою якістю та універсальністю, представлена модель
     рекомендована для використання на професійній кухні, її можна успішно використовувати
     в домашніх умовах.</p>

After UA:
  <p>Сучасний пом'якшувач м'яса Hendi вирізняється високою якістю та універсальністю,
     особливо корисний для приготування м'яса на грилі. Представлена модель рекомендована
     для використання на професійній кухні, її можна успішно використовувати в домашніх умовах.</p>
"""
import openpyxl, re

FIXED = r'C:\Projects\labresta-sync-w2\.planning\translation-audit\chunks\chunk-055-fixed.xlsx'
ROWS = [9, 10, 11, 12]
COL = 35  # UA description

# Apostrophe in UA xlsx is HTML-encoded as &#39; (verified in dump). Use the entity in regex.
APO = '&#39;'

NEW_LEAD_UA = (
    f"<p>Сучасний пом{APO}якшувач м{APO}яса Hendi вирізняється високою якістю та універсальністю, "
    f"особливо корисний для приготування м{APO}яса на грилі. Представлена модель рекомендована "
    f"для використання на професійній кухні, її можна успішно використовувати в домашніх умовах.</p>"
)

OLD_LEAD_RE = re.compile(
    r'<p>Сучасний пом' + re.escape(APO) + r'якшувач м' + re.escape(APO) + r'яса хенді, '
    r'особливо корисний для приготування м' + re.escape(APO) + r'яса на грилі\. '
    r'Продукція Hendi вирізняється високою якістю та універсальністю, представлена модель '
    r'рекомендован[аяні]+ для використання на професійній кухні, її можна успішно використовувати '
    r'в домашніх умовах\.</p>'
)

wb = openpyxl.load_workbook(FIXED)
ws = wb.active

changes = 0
for r in ROWS:
    art = ws.cell(r, 1).value
    v = ws.cell(r, COL).value
    if v is None:
        print(f'r{r} ART={art}: c{COL} empty — SKIP')
        continue
    s = str(v)
    new_s, n = OLD_LEAD_RE.subn(NEW_LEAD_UA, s, count=1)
    if n == 0:
        print(f'r{r} ART={art}: regex NO MATCH (unexpected)')
        print(f'  first 400 chars: {s[:400]}')
        continue
    ws.cell(r, COL).value = new_s
    changes += 1
    has_cyr = 'хенді' in new_s.lower()
    h_count = new_s.lower().count('hendi')
    print(f'r{r} ART={art}: PATCHED OK (хенді={has_cyr}, Hendi_count={h_count})')

if changes > 0:
    wb.save(FIXED)
    print(f'\n[OK] Saved {FIXED} ({changes} cells patched)')
else:
    print('\n[!!] No changes applied')
wb.close()
