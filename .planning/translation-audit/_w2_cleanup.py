# -*- coding: utf-8 -*-
"""
W2 cleanup wave 2 — universal processor.
Usage: python _w2_cleanup.py <chunk_NN> <batch_num> <start_row> <end_row>
Example: python _w2_cleanup.py 055 2 10 17

All 16 columns. Refined regex with Cyrillic word-boundaries.
"""
import re
import shutil
import sys
from openpyxl import load_workbook

COLS_RU = [5, 7, 23, 25, 27, 29, 36, 38]
COLS_UA = [4, 6, 22, 24, 26, 28, 35, 37]

CYR_RU = r'А-Яа-яЁё'
CYR_ALL = r'А-Яа-яЁёЇїІіЄєҐґ'

# --- RU rules ---

def drop_yo(s: str) -> str:
    return s.replace('Ё', 'Е').replace('ё', 'е')

def drop_apos_ru(s: str) -> str:
    s = re.sub(rf'(?<=[{CYR_RU}])&#39;(?=[{CYR_RU}])', '', s)
    s = re.sub(rf"(?<=[{CYR_RU}])'(?=[{CYR_RU}])", '', s)
    s = re.sub(rf"(?<=[{CYR_RU}])’(?=[{CYR_RU}])", '', s)
    return s

# Lex dict UA→RU (applied in RU cols only).
# Expanded based on findings per batch.
UA_RU_LEX = [
    # Hybrid typos (RU stem + UA letter)
    (r'\bпромішленній\b', 'промышленной'),
    (r'\bпромішленний\b', 'промышленный'),
    (r'\bпромышленній\b', 'промышленной'),
    (r'\bпрофессиональній\b', 'профессиональной'),
    (r'\bпрофессиональний\b', 'профессиональный'),
    # General UA stems
    (r'\bковбасний шприц\b', 'колбасный шприц'),
    (r'\bКовбасний шприц\b', 'Колбасный шприц'),
    (r'\bковбасний\b', 'колбасный'),
    (r'\bКовбасний\b', 'Колбасный'),
    (r'\bковбасні\b', 'колбасные'),
    (r'\bковбаси\b', 'колбасы'),
    (r'\bковбас\b', 'колбас'),
    (r'\bковбаса\b', 'колбаса'),
    (r'\bковбасу\b', 'колбасу'),
    (r'\bковбасою\b', 'колбасой'),
    (r'\bковбасі\b', 'колбасе'),
    (r'\bяловичина\b', 'говядина'),
    (r'\bЯловичина\b', 'Говядина'),
    (r'\bяловичини\b', 'говядины'),
    (r'\bяловичину\b', 'говядину'),
    (r'\bяловичиною\b', 'говядиной'),
    (r'\bхенді\b', 'хенди'),
    (r'\bХенді\b', 'Хенди'),
    (r'\bпривід\b', 'привод'),
    (r'\bПривід\b', 'Привод'),
    (r'\bелектричний\b', 'электрический'),
    (r'\bЕлектричний\b', 'Электрический'),
    (r'\bелектрична\b', 'электрическая'),
    (r'\bелектричні\b', 'электрические'),
    (r'\bелектро\b', 'электро'),
    (r'\bЕлектро\b', 'Электро'),
    (r'\bм&#39;яса\b', 'мяса'),
    (r"\bм'яса\b", 'мяса'),
    (r'\bм&#39;ясо\b', 'мясо'),
    (r"\bм'ясо\b", 'мясо'),
    (r'\bм&#39;ясу\b', 'мясу'),
    (r"\bм'ясу\b", 'мясу'),
    (r'\bм&#39;ясом\b', 'мясом'),
    (r"\bм'ясом\b", 'мясом'),
    (r'\bм&#39;ясний\b', 'мясной'),
    (r"\bм'ясний\b", 'мясной'),
    (r'\bм&#39;ясорубка\b', 'мясорубка'),
    (r"\bм'ясорубка\b", 'мясорубка'),
    (r'\bм&#39;ясорубки\b', 'мясорубки'),
    (r"\bм'ясорубки\b", 'мясорубки'),
    (r'\bм&#39;ясопухлин\b', 'мясорыхлитель'),
    (r"\bм'ясопухлин\b", 'мясорыхлитель'),
    (r'\bм&#39;ясорозпушувач\b', 'мясорыхлитель'),
    (r"\bм'ясорозпушувач\b", 'мясорыхлитель'),
    (r'\bпрофесійний\b', 'профессиональный'),
    (r'\bпрофесійна\b', 'профессиональная'),
    (r'\bпрофесійних\b', 'профессиональных'),
    (r'\bпрофесійні\b', 'профессиональные'),
    (r'\bПрофесійний\b', 'Профессиональный'),
    (r'\bпрофесіонал\b', 'профессионал'),
    (r'\bкорисний\b', 'полезный'),
    (r'\bкорисна\b', 'полезная'),
    (r'\bвидаляти\b', 'удалять'),
    (r'\bпідключення\b', 'подключение'),
    (r'\bПідключення\b', 'Подключение'),
    (r'\bвикористовується\b', 'используется'),
    (r'\bвикористання\b', 'использования'),
    (r'\bвикористовуються\b', 'используются'),
    (r'\bвикористовувати\b', 'использовать'),
    (r'\bоброблення\b', 'обработки'),
    (r'\bобробка\b', 'обработка'),
    (r'\bобробки\b', 'обработки'),
    (r'\bрозпушувач\b', 'разрыхлитель'),
    (r'\bрозм&#39;якшувач\b', 'размягчитель'),
    (r"\bрозм'якшувач\b", 'размягчитель'),
    (r'\bрозм&#39;якшення\b', 'размягчение'),
    (r"\bрозм'якшення\b", 'размягчение'),
    (r'\bпом&#39;якшувач\b', 'размягчитель'),
    (r"\bпом'якшувач\b", 'размягчитель'),
    (r'\bпом&#39;якшення\b', 'размягчение'),
    (r"\bпом'якшення\b", 'размягчение'),
    (r'\bвиконаний\b', 'выполнен'),
    (r'\bВиконаний\b', 'Выполнен'),
    (r'\bнержавіюча\b', 'нержавеющая'),
    (r'\bнержавіючої\b', 'нержавеющей'),
    (r'\bнержавіючий\b', 'нержавеющий'),
    (r'\bвищого\b', 'высшего'),
    (r'\bвища\b', 'высшая'),
    (r'\bвищий\b', 'высший'),
    (r'\bвищі\b', 'высшие'),
    (r'\bтехнічні\b', 'технические'),
    (r'\bТехнічні\b', 'Технические'),
    (r'\bкришка\b', 'крышка'),
    (r'\bсиліконові\b', 'силиконовые'),
    (r'\bсиліконовий\b', 'силиконовый'),
    (r'\bгерметичний\b', 'герметичный'),
    (r'\bгерметична\b', 'герметичная'),
    (r'\bциліндр\b', 'цилиндр'),
    (r'\bциліндра\b', 'цилиндра'),
    (r'\bциліндру\b', 'цилиндру'),
    (r'\bциліндром\b', 'цилиндром'),
    (r'\bстрижень\b', 'стержень'),
    (r'\bпорошня\b', 'поршня'),
    (r'\bбезшумний\b', 'бесшумный'),
    (r'\bвентильований\b', 'вентилируемый'),
    (r'\bдвигун\b', 'двигатель'),
    (r'\bобертів\b', 'оборотов'),
    (r'\bавтоматичне\b', 'автоматический'),
    (r'\bавтоматичний\b', 'автоматический'),
    (r'\bавтоматична\b', 'автоматическая'),
    (r'\bпульт керування\b', 'пульт управления'),
    (r'\bкерування\b', 'управления'),
    (r'\bвисокої точності\b', 'высокой точности'),
    (r'\bГабаритні розміри\b', 'Габаритные размеры'),
    (r'\bгабаритні\b', 'габаритные'),
    (r'\bупаковці\b', 'упаковке'),
    (r'\bПотужність\b', 'Мощность'),
    (r'\bпотужність\b', 'мощность'),
    (r'\bПродукція\b', 'Продукция'),
    (r'\bпродукція\b', 'продукция'),
    (r'\bвирізняється\b', 'отличается'),
    (r'\bякістю\b', 'качеством'),
    (r'\bуніверсальністю\b', 'универсальностью'),
    (r'\bпредставлена\b', 'представленная'),
    (r'\bрекомендована\b', 'рекомендованная'),
    (r'\bпрофесійній кухні\b', 'профессиональной кухне'),
    (r'\bумовах\b', 'условиях'),
    (r'\bдомашніх\b', 'домашних'),
    (r'\bтехнологія\b', 'технология'),
    (r'\bочищення\b', 'очистки'),
    (r'\bочищенні\b', 'очистке'),
    (r'\bпідрізають\b', 'подрезают'),
    (r'\bвалами\b', 'валами'),
    (r'\bріжучими\b', 'режущими'),
    (r'\bм&#39;язову\b', 'мышечную'),
    (r"\bм'язову\b", 'мышечную'),
    (r'\bтканину\b', 'ткань'),
    (r'\bвикидається\b', 'выбрасывается'),
    (r'\bмашини\b', 'машины'),
    (r'\bнаступної\b', 'последующей'),
    (r'\bтермообробки\b', 'термообработки'),
    (r'\bскорочується\b', 'сокращается'),
    (r'\bприблизно\b', 'примерно'),
    (r'\bрази\b', 'раза'),
    (r'\bкількість\b', 'количество'),
    (r'\bКількість\b', 'Количество'),
    (r'\bмаксимальна\b', 'максимальная'),
    (r'\bМаксимальна\b', 'Максимальная'),
    (r'\bширина\b', 'ширина'),
    (r'\bдовжина\b', 'длина'),
    (r'\bДовжина\b', 'Длина'),
    (r'\bвисота\b', 'высота'),
    (r'\bВисота\b', 'Высота'),
    (r'\bвага\b', 'вес'),
    (r'\bВага\b', 'Вес'),
    (r'\bодна швидкість\b', 'одна скорость'),
    (r'\bшвидкість\b', 'скорость'),
    (r'\bШвидкість\b', 'Скорость'),
    (r'\bобертання\b', 'вращения'),
    (r'\bохолоджується\b', 'охлаждается'),
    (r'\bвбудованим\b', 'встроенным'),
    (r'\bвентилятором\b', 'вентилятором'),
    (r'\bгумові\b', 'резиновые'),
    (r'\bніжки\b', 'ножки'),
    (r'\bніжки\b', 'ножки'),
    (r'\bматеріал\b', 'материал'),
    (r'\bматеріалу\b', 'материала'),
    (r'\bкорпус\b', 'корпус'),
    (r'\bкорпусу\b', 'корпуса'),
    (r'\bстрави\b', 'блюда'),
    (r'\bстрав\b', 'блюд'),
    # apostrophe variants (most should be caught by drop_apos_ru, but also lex)
]

def fix_ua_lex_in_ru(s: str) -> str:
    for pat, rep in UA_RU_LEX:
        s = re.sub(pat, rep, s)
    return s

def find_ua_stem_words_in_ru(s: str) -> list[str]:
    if not s:
        return []
    flagged = []
    for m in re.finditer(rf'[{CYR_ALL}]+', s):
        word = m.group()
        if re.search(r'[іїєґІЇЄҐ]', word):
            flagged.append(word)
    # uniq preserve order
    seen = set()
    out = []
    for w in flagged:
        if w.lower() not in seen:
            seen.add(w.lower())
            out.append(w)
    return out

# --- UA rules ---

UA_TYPO = [
    (r'\bпри зупинкі\b', 'при зупинці'),
]

def fix_ua_typo(s: str) -> str:
    for pat, rep in UA_TYPO:
        s = re.sub(pat, rep, s)
    return s

def has_yo(s: str) -> bool:
    return bool(re.search(r'[Ёё]', s or ''))

# --- main ---

def process_ru(value):
    if value is None:
        return value, []
    s = str(value)
    changes = []

    new = drop_yo(s)
    if new != s:
        changes.append('Ё→Е')
    s = new

    new = drop_apos_ru(s)
    if new != s:
        changes.append("apos→drop")
    s = new

    new = fix_ua_lex_in_ru(s)
    if new != s:
        changes.append("UA-lex→RU")
    s = new

    flagged = find_ua_stem_words_in_ru(s)
    if flagged:
        changes.append(f"FLAG-UA-stem-remaining: {flagged[:8]}")

    return s, changes

def process_ua(value):
    if value is None:
        return value, []
    s = str(value)
    changes = []

    new = fix_ua_typo(s)
    if new != s:
        changes.append("UA-typo-fix")
    s = new

    if has_yo(s):
        changes.append("FLAG-Ё-in-UA")

    return s, changes


def main():
    if len(sys.argv) != 5:
        print('Usage: python _w2_cleanup.py <chunk_NN> <batch_num> <start_row> <end_row>')
        sys.exit(1)

    chunk = sys.argv[1]
    batch = int(sys.argv[2])
    start_row = int(sys.argv[3])
    end_row = int(sys.argv[4])

    fixed_path = f'.planning/translation-audit/chunks/chunk-{chunk}-fixed.xlsx'
    backup_path = f'.planning/translation-audit/chunks/chunk-{chunk}-fixed.before_b{batch}.xlsx'
    diff_path = f'.planning/translation-audit/chunks/chunk-{chunk}-cleanup-wave2-b{batch}.md'

    shutil.copy(fixed_path, backup_path)
    wb = load_workbook(fixed_path)
    ws = wb.active

    lines = []
    lines.append(f'# chunk-{chunk} cleanup wave2 batch {batch} — SKU rows {start_row}-{end_row}, all 16 columns\n\n')
    lines.append('## Rules\n')
    lines.append('- RU (c5/c7/c23/c25/c27/c29/c36/c38): drop Ё→Е, drop apostrophe between RU letters, replace UA lex via expanded dict.\n')
    lines.append('- UA (c4/c6/c22/c24/c26/c28/c35/c37): fix зупинкі→зупинці; flag Ё.\n')
    lines.append('- FLAG (not auto-fixed): UA-stem words remaining in RU after lex pass; Ё in UA.\n\n')

    total = 0
    total_flags = 0
    for r in range(start_row, end_row + 1):
        sku = ws.cell(r, 1).value
        row_lines = []

        for c in COLS_RU:
            cell = ws.cell(r, c)
            new, ch = process_ru(cell.value)
            real = [x for x in ch if not x.startswith('FLAG')]
            flags = [x for x in ch if x.startswith('FLAG')]
            if real:
                cell.value = new
                row_lines.append(f'- c{c} RU: {"; ".join(real)}')
                total += 1
            for f in flags:
                row_lines.append(f'- c{c} RU [{f}]')
                total_flags += 1

        for c in COLS_UA:
            cell = ws.cell(r, c)
            new, ch = process_ua(cell.value)
            real = [x for x in ch if not x.startswith('FLAG')]
            flags = [x for x in ch if x.startswith('FLAG')]
            if real:
                cell.value = new
                row_lines.append(f'- c{c} UA: {"; ".join(real)}')
                total += 1
            for f in flags:
                row_lines.append(f'- c{c} UA [{f}]')
                total_flags += 1

        if row_lines:
            lines.append(f'## r{r} ART={sku}\n')
            lines.extend(rl + '\n' for rl in row_lines)
            lines.append('\n')

    wb.save(fixed_path)
    print(f'TOTAL real changes: {total}, FLAGs: {total_flags}')

    with open(diff_path, 'w', encoding='utf-8') as f:
        f.writelines(lines)
    print(f'Diff: {diff_path}')


if __name__ == '__main__':
    main()
