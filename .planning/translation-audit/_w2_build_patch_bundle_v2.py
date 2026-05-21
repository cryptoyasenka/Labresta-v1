"""Rebuild patch-bundle v2 — 5 SKUs (4 chunk-055 r9-r12 + 1 chunk-071 r66).

Schema 9 cols matches main bundle. Reads patched values (both RU + UA now).
"""
import openpyxl
from openpyxl import Workbook

HOROSHOP_EXPORT = r'C:\Projects\labresta-sync\horoshop-export 21.05.26.xlsx'
OUT_PATCH = r'C:\Projects\labresta-sync-w2\.planning\translation-audit\w2-horoshop-import-PATCH-hendi-lead.xlsx'

SOURCES = [
    (r'C:\Projects\labresta-sync-w2\.planning\translation-audit\chunks\chunk-055-fixed.xlsx', [9, 10, 11, 12]),
    (r'C:\Projects\labresta-sync-w2\.planning\translation-audit\chunks\chunk-071-fixed.xlsx', [66]),
]

SOURCE_COLS = [
    (1,  "Артикул"),
    (4,  "Название модификации (UA)"),
    (5,  "Название модификации (RU)"),
    (6,  "Название (UA)"),
    (7,  "Название (RU)"),
    (24, "META keywords (UA)"),
    (25, "META keywords (RU)"),
    (35, "Описание товара (UA)"),
    (36, "Описание товара (RU)"),
]
OUT_HEADER = [name for _, name in SOURCE_COLS]

def is_empty(v):
    return v is None or str(v).strip() == ''

hw = openpyxl.load_workbook(HOROSHOP_EXPORT, data_only=True)
hs = hw.active
horoshop = {}
for r in range(2, hs.max_row + 1):
    art = hs.cell(r, 1).value
    if art is None:
        continue
    horoshop[str(art).strip()] = {c: hs.cell(r, c).value for c, _ in SOURCE_COLS}
hw.close()

out = Workbook()
ws_out = out.active
ws_out.title = 'import'
ws_out.append(OUT_HEADER)

empty_count = 0
filled_from_horoshop = 0
written = 0
for fp, rows in SOURCES:
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb.active
    for r in rows:
        art_raw = ws.cell(r, 1).value
        if is_empty(art_raw):
            continue
        art = str(art_raw).strip()
        h_row = horoshop.get(art)
        row_out = []
        skip = False
        for src_col, col_name in SOURCE_COLS:
            if src_col == 1:
                row_out.append(art)
                continue
            v = ws.cell(r, src_col).value
            if is_empty(v):
                fallback = h_row.get(src_col) if h_row else None
                if is_empty(fallback):
                    empty_count += 1
                    print(f'!!! r{r} ART={art} c{src_col} EMPTY both — SKIP row')
                    skip = True
                    break
                v = fallback
                filled_from_horoshop += 1
                print(f'   r{r} ART={art} c{src_col}: filled from Horoshop')
            row_out.append(v)
        if not skip:
            ws_out.append(row_out)
            written += 1
            print(f'   r{r} ART={art}: appended')
    wb.close()

out.save(OUT_PATCH)
print(f'\n[OK] saved {OUT_PATCH}')
print(f'    rows written: {written}')
print(f'    fallback: {filled_from_horoshop}, excluded empty: {empty_count}')

# Verify
print('\n[VERIFY] re-reading patch bundle:')
wb2 = openpyxl.load_workbook(OUT_PATCH, data_only=True)
ws2 = wb2.active
for r in range(2, ws2.max_row + 1):
    art = ws2.cell(r, 1).value
    ua = ws2.cell(r, 8).value or ''  # col 8 in 9-col schema = c35 UA desc
    ru = ws2.cell(r, 9).value or ''  # col 9 = c36 RU desc
    ua_s, ru_s = str(ua), str(ru)
    ua_cyr = 'хенді' in ua_s.lower() or 'хенди' in ua_s.lower()
    ru_cyr = 'хенди' in ru_s.lower() or 'хенді' in ru_s.lower()
    print(f'    ART={art}: UA_cyr={ua_cyr}, UA_Hendi={ua_s.lower().count("hendi")}, RU_cyr={ru_cyr}, RU_Hendi={ru_s.lower().count("hendi")}')
wb2.close()
