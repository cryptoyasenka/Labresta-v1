# -*- coding: utf-8 -*-
"""
W2 cleanup wave 2 — chunk-055 batch 1 (SKU 1-8, rows 2-9)
All 16 columns. Refined regex with Cyrillic word-boundaries.
"""
import re
import shutil
from openpyxl import load_workbook

FIXED_PATH = '.planning/translation-audit/chunks/chunk-055-fixed.xlsx'
BACKUP_PATH = '.planning/translation-audit/chunks/chunk-055-fixed.before_b1.xlsx'

COLS_RU = [5, 7, 23, 25, 27, 29, 36, 38]
COLS_UA = [4, 6, 22, 24, 26, 28, 35, 37]

CYR_RU = r'А-Яа-яЁё'
CYR_ALL = r'А-Яа-яЁёЇїІіЄєҐґ'

# --- RU rules ---

def drop_yo(s: str) -> str:
    return s.replace('Ё', 'Е').replace('ё', 'е')

def drop_apos_ru(s: str) -> str:
    # apostrophe between RU letters → drop (handles HTML entity, plain, curly)
    s = re.sub(rf'(?<=[{CYR_RU}])&#39;(?=[{CYR_RU}])', '', s)
    s = re.sub(rf"(?<=[{CYR_RU}])'(?=[{CYR_RU}])", '', s)
    s = re.sub(rf"(?<=[{CYR_RU}])’(?=[{CYR_RU}])", '', s)
    return s

# UA lex → RU lex (applied in RU columns only)
UA_RU_LEX = [
    (r'\bковбасний шприц\b', 'колбасный шприц'),
    (r'\bКовбасний шприц\b', 'Колбасный шприц'),
    (r'\bковбасний\b', 'колбасный'),
    (r'\bКовбасний\b', 'Колбасный'),
    (r'\bяловичина\b', 'говядина'),
    (r'\bЯловичина\b', 'Говядина'),
    (r'\bхенді\b', 'хенди'),
    (r'\bХенді\b', 'Хенди'),
    (r'\bпривід\b', 'привод'),
    (r'\bПривід\b', 'Привод'),
    (r'\bелектричний\b', 'электрический'),
    (r'\bЕлектричний\b', 'Электрический'),
]

def fix_ua_lex_in_ru(s: str) -> str:
    for pat, rep in UA_RU_LEX:
        s = re.sub(pat, rep, s)
    return s

def find_ua_stem_words_in_ru(s: str) -> list[str]:
    """Cyrillic words containing UA-specific letters (і/ї/є/ґ); skip Latin-mixed tokens."""
    if not s:
        return []
    flagged = []
    for m in re.finditer(rf'[{CYR_ALL}]+', s):
        word = m.group()
        if re.search(r'[іїєґІЇЄҐ]', word):
            # ensure not Latin-glued (already excluded by class)
            flagged.append(word)
    return flagged

# --- UA rules ---

UA_TYPO = [
    (r'\bпри зупинкі\b', 'при зупинці'),
]

def fix_ua_typo(s: str) -> str:
    for pat, rep in UA_TYPO:
        s = re.sub(pat, rep, s)
    return s

def has_yo(s: str) -> bool:
    return bool(re.search(r'[Ёё]', s or ''))

# --- main ---

def process_ru(value):
    if value is None:
        return value, []
    s = str(value)
    changes = []

    new = drop_yo(s)
    if new != s:
        changes.append('Ё→Е')
    s = new

    new = drop_apos_ru(s)
    if new != s:
        changes.append("apos→drop")
    s = new

    new = fix_ua_lex_in_ru(s)
    if new != s:
        changes.append("UA-lex→RU")
    s = new

    flagged = find_ua_stem_words_in_ru(s)
    if flagged:
        changes.append(f"FLAG-UA-stem-remaining: {flagged[:5]}")

    return s, changes

def process_ua(value):
    if value is None:
        return value, []
    s = str(value)
    changes = []

    new = fix_ua_typo(s)
    if new != s:
        changes.append("UA-typo-fix(зупинкі→зупинці)")
    s = new

    if has_yo(s):
        changes.append("FLAG-Ё-in-UA")

    return s, changes


def main():
    shutil.copy(FIXED_PATH, BACKUP_PATH)
    wb = load_workbook(FIXED_PATH)
    ws = wb.active

    lines = []
    lines.append('# chunk-055 cleanup wave2 batch 1 — SKU 1-8 (rows 2-9), all 16 columns\n\n')
    lines.append('## Rules\n')
    lines.append('- RU (c5/c7/c23/c25/c27/c29/c36/c38): drop Ё→Е, drop apostrophe between RU letters, replace UA lex via dict (ковбасний→колбасный, яловичина→говядина, хенді→хенди, привід→привод, електричний→электрический).\n')
    lines.append('- UA (c4/c6/c22/c24/c26/c28/c35/c37): fix зупинкі→зупинці; flag Ё.\n')
    lines.append('- FLAGs (not auto-fixed): UA-stem words remaining in RU after lex pass; Ё in UA.\n\n')

    total = 0
    for r in range(2, 10):
        sku = ws.cell(r, 1).value
        row_lines = []

        for c in COLS_RU:
            cell = ws.cell(r, c)
            new, ch = process_ru(cell.value)
            real = [x for x in ch if not x.startswith('FLAG')]
            flags = [x for x in ch if x.startswith('FLAG')]
            if real:
                cell.value = new
                row_lines.append(f'- c{c} RU: {"; ".join(real)}')
                total += 1
            for f in flags:
                row_lines.append(f'- c{c} RU [{f}]')

        for c in COLS_UA:
            cell = ws.cell(r, c)
            new, ch = process_ua(cell.value)
            real = [x for x in ch if not x.startswith('FLAG')]
            flags = [x for x in ch if x.startswith('FLAG')]
            if real:
                cell.value = new
                row_lines.append(f'- c{c} UA: {"; ".join(real)}')
                total += 1
            for f in flags:
                row_lines.append(f'- c{c} UA [{f}]')

        if row_lines:
            lines.append(f'## r{r} ART={sku}\n')
            lines.extend(rl + '\n' for rl in row_lines)
            lines.append('\n')

    wb.save(FIXED_PATH)
    print(f'TOTAL real changes: {total}')

    with open('.planning/translation-audit/chunks/chunk-055-cleanup-wave2-b1.md', 'w', encoding='utf-8') as f:
        f.writelines(lines)


if __name__ == '__main__':
    main()
