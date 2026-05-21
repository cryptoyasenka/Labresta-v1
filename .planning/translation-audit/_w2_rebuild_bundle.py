"""W2 Horoshop import bundle rebuild — chunks 055-085, post-OQ-apply.

Schema: same as prior bundle (4 RU cols).
  c01 Артикул
  c02 Название модификации (RU)  ← source c05
  c03 Название (RU)              ← source c07
  c04 Описание товара (RU)       ← source c36

Filters:
  - SKIP-НП brands (HURAKAN/APACH/FAGOR/TATRA/COLD/PROJECT SYSTEMS/ASTORIA/ARRIS/MAXIMA) → excluded
  - Rows unchanged vs source chunk-NN.xlsx (W2 did no cleanup) → excluded

Safety rule (Yana): NO empty cells in bundle, иначе Horoshop сотрёт данные.
  - Empty in fixed.xlsx → fallback to Horoshop export (current live value)
  - Still empty after fallback → WARNING + row excluded
"""
import openpyxl
import os
from openpyxl import Workbook

CHUNKS_DIR = r"C:\Projects\labresta-sync-w2\.planning\translation-audit\chunks"
HOROSHOP_EXPORT = r"C:\Projects\labresta-sync\horoshop-export 21.05.26.xlsx"
OUT_PATH = r"C:\Projects\labresta-sync-w2\.planning\translation-audit\w2-horoshop-import-055-085.xlsx"
TEST_PATH = r"C:\Projects\labresta-sync-w2\.planning\translation-audit\w2-horoshop-import-TEST-1.xlsx"

# Schema: source col index in chunk-NN-fixed.xlsx → output col name
SOURCE_COLS = [
    (1,  "Артикул"),
    (5,  "Название модификации (RU)"),
    (7,  "Название (RU)"),
    (36, "Описание товара (RU)"),
]
OUT_HEADER = [name for _, name in SOURCE_COLS]
TRANS_COLS = [c for c, _ in SOURCE_COLS if c != 1]  # cols to validate non-empty: 5, 7, 36

SKIP_BRANDS = [
    "hurakan", "apach", "fagor", "tatra", "cold",
    "project systems", "astoria", "arris", "maxima",
    # cyrillic equivalents
    "хуракан", "апач", "фагор", "татра", "колд", "астория", "аррис", "максима",
]

_SKIP_BRANDS_SET = {sb.strip().lower() for sb in SKIP_BRANDS}

def is_skip_brand(b):
    """Exact match against SKIP-НП brand list (case-insensitive).
    NB: substring match is unsafe — 'cold' would falsely match 'Tefcold'.
    Brands are stored as canonical names in c08 (e.g. 'Tefcold', 'Hurakan'),
    so case-insensitive equality is sufficient."""
    if not b:
        return False
    return str(b).strip().lower() in _SKIP_BRANDS_SET

def is_empty(v):
    return v is None or str(v).strip() == ""

# ─── Load Horoshop export as fallback dict ───────────────────────
print("Loading Horoshop export...")
hw = openpyxl.load_workbook(HOROSHOP_EXPORT, data_only=True)
hs = hw.active
horoshop = {}
for r in range(2, hs.max_row + 1):
    art = hs.cell(r, 1).value
    if art is None:
        continue
    key = str(art).strip()
    horoshop[key] = {c: hs.cell(r, c).value for c, _ in SOURCE_COLS}
hw.close()
print(f"  {len(horoshop)} SKUs in Horoshop export")

# ─── Build bundle ────────────────────────────────────────────────
out = Workbook()
ws_out = out.active
ws_out.title = "import"
ws_out.append(OUT_HEADER)

stats = {
    "total_seen": 0,
    "skip_np": 0,
    "unchanged_vs_source": 0,
    "empty_filled_from_horoshop": 0,
    "row_excluded_empty_after_fallback": 0,
    "not_in_horoshop_warn": 0,
    "written": 0,
}
warnings = []
sku_first_per_chunk = {}  # chunk → first written row for TEST-1

CHUNK_FILES = sorted(
    f for f in os.listdir(CHUNKS_DIR)
    if f.startswith("chunk-0") and f.endswith("-fixed.xlsx")
    and 55 <= int(f.split("-")[1]) <= 85
)
print(f"\nProcessing {len(CHUNK_FILES)} chunks (055-085)...")

for cf in CHUNK_FILES:
    fpath = os.path.join(CHUNKS_DIR, cf)
    spath = os.path.join(CHUNKS_DIR, cf.replace("-fixed.xlsx", ".xlsx"))

    wf = openpyxl.load_workbook(fpath, data_only=True)
    wsf = wf.active

    ws_src = None
    ws_s_handle = None
    if os.path.exists(spath):
        ws_s_handle = openpyxl.load_workbook(spath, data_only=True)
        ws_src = ws_s_handle.active

    chunk_written = 0
    for r in range(2, wsf.max_row + 1):
        stats["total_seen"] += 1
        artikul_raw = wsf.cell(r, 1).value
        if is_empty(artikul_raw):
            continue
        artikul = str(artikul_raw).strip()

        brand = wsf.cell(r, 8).value
        if is_skip_brand(brand):
            stats["skip_np"] += 1
            continue

        # Detect unchanged vs source (W2 did no cleanup → skip)
        if ws_src is not None:
            same = all(
                wsf.cell(r, c).value == ws_src.cell(r, c).value
                for c in TRANS_COLS
            )
            if same:
                stats["unchanged_vs_source"] += 1
                continue

        # Collect cells with empty-cell guard
        row_out = []
        h_row = horoshop.get(artikul)
        if h_row is None:
            stats["not_in_horoshop_warn"] += 1
            warnings.append(f"NOT-IN-HOROSHOP {artikul} ({cf} r{r}) — fallback unavailable")

        row_has_unrecoverable_empty = False
        for src_col, col_name in SOURCE_COLS:
            if src_col == 1:
                row_out.append(artikul)
                continue
            v = wsf.cell(r, src_col).value
            if is_empty(v):
                # Fallback to Horoshop current value
                fallback = h_row.get(src_col) if h_row else None
                if not is_empty(fallback):
                    v = fallback
                    stats["empty_filled_from_horoshop"] += 1
                    warnings.append(f"FALLBACK {artikul} c{src_col} ({col_name}) [{cf} r{r}] — empty in fixed.xlsx, filled from Horoshop ({len(str(fallback))} chars)")
                else:
                    row_has_unrecoverable_empty = True
                    stats["row_excluded_empty_after_fallback"] += 1
                    warnings.append(f"EXCLUDED {artikul} c{src_col} ({col_name}) [{cf} r{r}] — BOTH empty (fixed AND Horoshop) — row excluded to prevent wipe")
                    break
            row_out.append(v)

        if row_has_unrecoverable_empty:
            continue

        ws_out.append(row_out)
        stats["written"] += 1
        chunk_written += 1
        if cf not in sku_first_per_chunk:
            sku_first_per_chunk[cf] = row_out

    print(f"  {cf}: written {chunk_written}")
    wf.close()
    if ws_s_handle is not None:
        ws_s_handle.close()

out.save(OUT_PATH)
print(f"\n[OK] bundle saved: {OUT_PATH}")

# ─── TEST-1: 1 SKU sample (chunk-055 first written row) ─────────
test_wb = Workbook()
test_ws = test_wb.active
test_ws.title = "import"
test_ws.append(OUT_HEADER)
first_chunk = sorted(sku_first_per_chunk.keys())[0]
test_ws.append(sku_first_per_chunk[first_chunk])
test_wb.save(TEST_PATH)
print(f"[OK] TEST-1 saved (1 SKU from {first_chunk}): {TEST_PATH}")

# ─── Final report ───────────────────────────────────────────────
print(f"\n{'='*60}\nSTATS:")
for k, v in stats.items():
    print(f"  {k}: {v}")

print(f"\nWARNINGS ({len(warnings)}):")
fallback_count = sum(1 for w in warnings if w.startswith("FALLBACK"))
excluded_count = sum(1 for w in warnings if w.startswith("EXCLUDED"))
notinh_count = sum(1 for w in warnings if w.startswith("NOT-IN-HOROSHOP"))
print(f"  FALLBACK (filled from Horoshop): {fallback_count}")
print(f"  EXCLUDED (both empty, row dropped): {excluded_count}")
print(f"  NOT-IN-HOROSHOP (SKU missing in export): {notinh_count}")

if excluded_count > 0:
    print(f"\n!!! EXCLUDED rows (both fixed and Horoshop empty):")
    for w in warnings:
        if w.startswith("EXCLUDED"):
            print(f"  {w}")

if notinh_count > 0 and notinh_count <= 30:
    print(f"\n  NOT-IN-HOROSHOP details:")
    for w in warnings:
        if w.startswith("NOT-IN-HOROSHOP"):
            print(f"  {w}")
