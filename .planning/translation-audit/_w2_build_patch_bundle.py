"""Build Horoshop patch-bundle ONLY for 4 Hendi-tendariser SKUs (chunk-055 r9-r12).

Schema identical to main bundle (9 cols RU+UA). Reads patched values from chunk-055-fixed.xlsx
and writes a tiny xlsx for Horoshop /import.

Safety: same empty-cell guard — fallback to horoshop-export if any cell empty (won't happen here).
"""
import openpyxl
from openpyxl import Workbook

FIXED = r'C:\Projects\labresta-sync-w2\.planning\translation-audit\chunks\chunk-055-fixed.xlsx'
HOROSHOP_EXPORT = r'C:\Projects\labresta-sync\horoshop-export 21.05.26.xlsx'
OUT_PATCH = r'C:\Projects\labresta-sync-w2\.planning\translation-audit\w2-horoshop-import-PATCH-hendi-lead.xlsx'

SOURCE_COLS = [
    (1,  "Артикул"),
    (4,  "Название модификации (UA)"),
    (5,  "Название модификации (RU)"),
    (6,  "Название (UA)"),
    (7,  "Название (RU)"),
    (24, "META keywords (UA)"),
    (25, "META keywords (RU)"),
    (35, "Описание товара (UA)"),
    (36, "Описание товара (RU)"),
]
OUT_HEADER = [name for _, name in SOURCE_COLS]
ROWS = [9, 10, 11, 12]  # 4 tendarisers

def is_empty(v):
    return v is None or str(v).strip() == ''

# Read horoshop export as fallback
hw = openpyxl.load_workbook(HOROSHOP_EXPORT, data_only=True)
hs = hw.active
horoshop = {}
for r in range(2, hs.max_row + 1):
    art = hs.cell(r, 1).value
    if art is None:
        continue
    horoshop[str(art).strip()] = {c: hs.cell(r, c).value for c, _ in SOURCE_COLS}
hw.close()

wb = openpyxl.load_workbook(FIXED, data_only=True)
ws = wb.active

out = Workbook()
ws_out = out.active
ws_out.title = 'import'
ws_out.append(OUT_HEADER)

empty_count = 0
filled_from_horoshop = 0
for r in ROWS:
    art_raw = ws.cell(r, 1).value
    if is_empty(art_raw):
        continue
    art = str(art_raw).strip()
    h_row = horoshop.get(art)
    row_out = []
    for src_col, col_name in SOURCE_COLS:
        if src_col == 1:
            row_out.append(art)
            continue
        v = ws.cell(r, src_col).value
        if is_empty(v):
            fallback = h_row.get(src_col) if h_row else None
            if is_empty(fallback):
                empty_count += 1
                print(f'!!! r{r} ART={art} c{src_col} ({col_name}): EMPTY in both fixed and Horoshop — SKIP')
                break
            v = fallback
            filled_from_horoshop += 1
            print(f'   r{r} ART={art} c{src_col}: filled from Horoshop')
        row_out.append(v)
    else:
        ws_out.append(row_out)
        print(f'   r{r} ART={art}: appended')
wb.close()

out.save(OUT_PATCH)
print(f'\n[OK] saved {OUT_PATCH}')
print(f'    rows: {ws_out.max_row - 1}')
print(f'    cells filled from Horoshop fallback: {filled_from_horoshop}')
print(f'    cells excluded (both empty): {empty_count}')

# Verify patch result
print('\n[VERIFY] re-reading patch bundle to confirm Hendi present and хенди absent:')
wb2 = openpyxl.load_workbook(OUT_PATCH, data_only=True)
ws2 = wb2.active
for r in range(2, ws2.max_row + 1):
    art = ws2.cell(r, 1).value
    desc = ws2.cell(r, 9).value  # col 9 = c36 in our schema (Описание товара RU)
    if desc is None:
        continue
    s = str(desc)
    has_xenди = 'хенди' in s.lower()
    h_count = s.lower().count('hendi')
    print(f'    ART={art}: хенди={has_xenди}, Hendi_count={h_count}, c36_len={len(s)}')
wb2.close()
