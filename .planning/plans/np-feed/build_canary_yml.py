"""Канареечный YML на 1 товар HKN-PICO12M. ТОЛЬКО план-каталог. Репо не трогаем.

Источники read-only: np-feed.xlsx (тело UA/RU + галерея) + labresta.db?mode=ro
(available). Состав offer: id+vendorCode+available+description+description_ru+
picture×N. БЕЗ name/price/oldprice/vendor (тест-1: изолируем тело+фото).
"""
import sqlite3
import openpyxl
from lxml import etree

XLSX = r"C:\Users\Yana\labresta-np-feed-plan\np-feed.xlsx"
DB = r"C:\Projects\labresta-sync\instance\labresta.db"
OUT = r"C:\Users\Yana\labresta-np-feed-plan\canary-HKN-PICO12M.yml"
EXTERNAL_ID = "1546341257"
ARTICLE = "HKN-PICO12M"

# col idx (FINDINGS §2): B=1 article, D=3 photos, H=7 desc_uk, Q=16 desc_ru
wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
row = None
for r in ws.iter_rows(min_row=2, values_only=True):
    if r and len(r) > 1 and str(r[1]).strip() == ARTICLE:
        row = r
        break
wb.close()
if row is None:
    raise SystemExit(f"{ARTICLE} not found in xlsx")

desc_uk = (row[7] or "").strip()
desc_ru = (row[16] or "").strip()
photos_raw = (row[3] or "").strip()
photos = [p.strip() for p in photos_raw.split(";") if p.strip()]
# dedup, preserve order
seen, pics = set(), []
for p in photos:
    if p not in seen:
        seen.add(p)
        pics.append(p)

con = sqlite3.connect(f"file:{DB}?mode=ro&immutable=1", uri=True)
av = con.execute(
    "SELECT available FROM supplier_products WHERE supplier_id=2 AND article=?",
    (ARTICLE,)).fetchone()
con.close()
available = "true" if (av and av[0]) else "false"

root = etree.Element("yml_catalog", date="canary-test")
shop = etree.SubElement(root, "shop")
etree.SubElement(shop, "name").text = "LabResta"
etree.SubElement(shop, "company").text = "LabResta"
etree.SubElement(shop, "url").text = "https://labresta.com"
cur = etree.SubElement(shop, "currencies")
etree.SubElement(cur, "currency", id="EUR", rate="1")
etree.SubElement(cur, "currency", id="UAH", rate="1")
etree.SubElement(shop, "categories")
offers = etree.SubElement(shop, "offers")

offer = etree.SubElement(offers, "offer", id=EXTERNAL_ID, available=available)
etree.SubElement(offer, "vendorCode").text = EXTERNAL_ID
if desc_uk:
    etree.SubElement(offer, "description").text = etree.CDATA(desc_uk)
if desc_ru:
    etree.SubElement(offer, "description_ru").text = etree.CDATA(desc_ru)
for p in pics:
    etree.SubElement(offer, "picture").text = p
# НЕТ: name / name_ru / price / oldprice / currencyId / vendor

tree = etree.ElementTree(root)
tree.write(OUT, xml_declaration=True, encoding="UTF-8", pretty_print=True)

with open(r"C:\Users\Yana\labresta-np-feed-plan\scratch_canary_meta.txt",
          "w", encoding="utf-8") as f:
    f.write(f"article={ARTICLE} external_id={EXTERNAL_ID} available={available}\n")
    f.write(f"desc_uk_len={len(desc_uk)} desc_ru_len={len(desc_ru)}\n")
    f.write(f"photos={len(pics)}\n")
    for i, p in enumerate(pics, 1):
        f.write(f"  pic{i}: {p}\n")
    f.write(f"uk_head: {desc_uk[:160]}\n")
    f.write(f"ru_head: {desc_ru[:160]}\n")
print("OK", OUT)
