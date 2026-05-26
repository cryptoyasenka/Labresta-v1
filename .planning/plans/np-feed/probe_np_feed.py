"""GATE probe for the NP ("Новый проект") dealer-export xlsx feed.

READ-ONLY. Downloads the feed once via GET, dumps structure + ru!=uk gate
to np-probe-dump.txt (UTF-8). No DB, no Horoshop, no writes outside this folder.
"""
import os
import random
import statistics

import requests
import openpyxl

OUT = os.path.dirname(os.path.abspath(__file__))
FEED_URL = "https://np.com.ua/dealer-export?dealer_id=69781&filetype=xlsx&platform=horoshop"
XLSX_PATH = os.path.join(OUT, "np-feed.xlsx")
DUMP_PATH = os.path.join(OUT, "np-probe-dump.txt")

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*",
}

_lines = []


def log(s=""):
    _lines.append(str(s))


def download():
    if os.path.exists(XLSX_PATH) and os.path.getsize(XLSX_PATH) > 10000:
        log(f"[cache] reuse {XLSX_PATH} ({os.path.getsize(XLSX_PATH)} bytes)")
        return
    r = requests.get(FEED_URL, headers=HEADERS, timeout=90)
    r.raise_for_status()
    with open(XLSX_PATH, "wb") as f:
        f.write(r.content)
    log(f"[get] {len(r.content)} bytes ctype={r.headers.get('Content-Type')}")


def col_letter(idx0):
    s, n = "", idx0 + 1
    while n:
        n, rem = divmod(n - 1, 26)
        s = chr(65 + rem) + s
    return s


def main():
    download()
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    log(f"sheets = {wb.sheetnames}")
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header = next(it)
    rows = [r for r in it]
    wb.close()

    log(f"\n=== HEADER ({len(header)} cols) ===")
    for i, h in enumerate(header):
        log(f"  [{i:>2}] {col_letter(i):>2}  {('' if h is None else str(h))!r}")

    log(f"\n=== ROW COUNT: {len(rows)} data rows ===")

    # Dump first 2 rows fully so mojibake/bracket columns are decoded properly.
    log("\n=== FIRST 2 DATA ROWS (full) ===")
    for ri, r in enumerate(rows[:2]):
        log(f"-- row {ri} --")
        for i, v in enumerate(r):
            vs = "" if v is None else str(v)
            log(f"  [{i:>2}] {col_letter(i):>2} {('' if header[i] is None else str(header[i]))!r}: "
                f"{vs[:300]!r}")

    # Column index map by exact header name
    by_name = {}
    for i, h in enumerate(header):
        by_name[("" if h is None else str(h)).strip()] = i

    ID = by_name.get("id")
    ART = 1  # col B per LabResta column_mapping {"1":"article"}
    UA = by_name.get("description_uk")
    RU = by_name.get("description_ru")
    TUA = by_name.get("title_uk")
    TRU = by_name.get("title_ru")
    BR = by_name.get("attr_brend_uk")

    log(f"\nresolved: id={ID} article(colB)={ART} title_uk={TUA} title_ru={TRU} "
        f"desc_uk={UA} desc_ru={RU} brand={BR}")

    # ---- Article column sanity (col B) ----
    log("\n=== ARTICLE (col B) sample 15 ===")
    for r in rows[:15]:
        log(f"  id={r[ID]!r:>14}  art={str(r[ART])[:40]!r:>42}  "
            f"title_uk={str(r[TUA])[:60]!r}")

    # ---- Photo/gallery columns: scan B..F bracket params for url-ish content ----
    log("\n=== COLS 0..6 first-nonempty sample (find photo/price/gallery) ===")
    for ci in range(0, 7):
        sample_vals = []
        for r in rows:
            v = r[ci] if ci < len(r) else None
            if v is not None and str(v).strip():
                sample_vals.append(str(v).strip())
            if len(sample_vals) >= 3:
                break
        log(f"  col {ci} {col_letter(ci)} {('' if header[ci] is None else str(header[ci]))!r}:")
        for sv in sample_vals:
            log(f"      {sv[:200]!r}")

    # ---- GATE: ru != uk across sample ----
    random.seed(42)
    sample = random.sample(rows, min(60, len(rows)))
    checked = both = diff = ident = 0
    ua_lens, ru_lens = [], []
    examples = []
    for r in sample:
        ua = str(r[UA]).strip() if UA is not None and r[UA] is not None else ""
        ru = str(r[RU]).strip() if RU is not None and r[RU] is not None else ""
        if not ua and not ru:
            continue
        checked += 1
        if ua and ru:
            both += 1
            ua_lens.append(len(ua))
            ru_lens.append(len(ru))
            if ua == ru:
                ident += 1
            else:
                diff += 1
                if len(examples) < 6:
                    examples.append((str(r[ID]), ua, ru))

    log("\n=== GATE RESULT (ru vs uk description) ===")
    log(f"sampled rows               : {len(sample)}")
    log(f"rows with any description  : {checked}")
    log(f"rows with BOTH ua & ru     : {both}")
    log(f"rows ru == ua (identical)  : {ident}")
    log(f"rows ru != ua (distinct)   : {diff}")
    if ua_lens:
        log(f"len uk: min={min(ua_lens)} med={int(statistics.median(ua_lens))} max={max(ua_lens)}")
        log(f"len ru: min={min(ru_lens)} med={int(statistics.median(ru_lens))} max={max(ru_lens)}")

    log("\n=== UP TO 6 EXAMPLES (id / UA head / RU head) ===")
    for k, (pid, ua, ru) in enumerate(examples, 1):
        log(f"\n[{k}] id={pid}")
        log(f"    UA: {ua[:280]!r}")
        log(f"    RU: {ru[:280]!r}")

    # ---- Brand distribution for the 9 NP-exclusive brands ----
    NP_BRANDS = ["hurakan", "хуракан", "apach", "апач", "fagor", "фагор",
                 "tatra", "татра", "cold", "колд", "project systems",
                 "astoria", "асторія", "астория", "arris", "arріс", "арріс",
                 "maxima", "максима"]
    from collections import Counter
    bc = Counter()
    np_hit = 0
    for r in rows:
        b = str(r[BR]).strip() if BR is not None and r[BR] is not None else ""
        bc[b] += 1
        bl = b.lower()
        if any(nb in bl for nb in NP_BRANDS):
            np_hit += 1
    log(f"\n=== BRAND DISTRIBUTION (attr_brend_uk) — total {len(rows)} ===")
    for b, c in bc.most_common(40):
        log(f"  {c:>4}  {b!r}")
    log(f"\nrows whose brand matches one of the 9 NP-exclusive names: {np_hit}")

    with open(DUMP_PATH, "w", encoding="utf-8") as f:
        f.write("\n".join(_lines))
    # Console: ascii-safe summary only
    print(f"dump written: {DUMP_PATH}")
    print(f"rows={len(rows)} both_ua_ru={both} ru!=ua={diff} identical={ident} "
          f"np_brand_rows={np_hit}")


if __name__ == "__main__":
    main()
