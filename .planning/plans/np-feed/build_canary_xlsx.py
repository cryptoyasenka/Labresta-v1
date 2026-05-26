"""Канарейка v6 — НАТИВНАЯ Horoshop-СХЕМА (XLSX dealer-export), а НЕ YML.

ПОЧЕМУ: Horoshop import авто-маппит колонку → поле каталога ТОЛЬКО когда имя
колонки фида == имя поля каталога (док help.horoshop + доказано живьём в
profile1: YML <name> Horoshop жёстко зовёт «Название модификации(RU)» →
украинский текст падал в русское поле; <name_ru>/<description_ru>/<picture> он
не узнаёт → «Не імпортувати»). np-feed.xlsx — это сам дилер-экспорт Horoshop,
его 24 родные колонки Horoshop узнаёт автоматически (round-trip). Поэтому
канарейку отдаём в ТОЙ ЖЕ нативной схеме: заголовки берём ДОСЛОВНО из
np-feed.xlsx (не набираем кириллицу руками), значения тела/фото/имени —
из фида НП, цену/наличие — из матчера (идемпотентно как live-фид).

Override vs сырая строка поставщика:
- Артикул (col1) := pp.external_id (=1546341257) — Horoshop матчит товар по
  Артикулу; ставим артикул НАШЕЙ карточки, не supplier 'HKN-PICO12M'.
- [КАТАЛОГ] Цена := наша скидочная цена из матчера (718.3), НЕ supplier 845.
- + 2 НАТИВНЫЕ колонки, которых нет в supplier-экспорте, но есть как поля
  каталога: «[КАТАЛОГ] Старая цена» (=retail 845, чтобы Horoshop сам считал %
  и не двоил скидку — как в основном фиде) и «[КАТАЛОГ] Валюта» (=EUR).
  Префикс «[КАТАЛОГ] » берём из реального заголовка col2, не из головы.
- categories_uk/ru и attr_* (dims/power/voltage/weight/contry) НЕ трогаем
  (пусто) — это характеристики/категория, отдельный import_type; канарейка =
  имя UA/RU + тело UA/RU + фото + цена + бренд. Страну/характеристики — позже.

Репо READ-ONLY (sqlite mode=ro). Пишет ТОЛЬКО в plan-каталог + _serve\.
"""
import math
import sqlite3
import openpyxl

DB = r"C:/Projects/labresta-sync/instance/labresta.db"   # mode=ro
NP_FEED = r"C:/Users/Yana/labresta-np-feed-plan/np-feed.xlsx"
EXTERNAL_ID = "1546341257"                                 # pp.external_id
_EUR_RATE_FALLBACK = 51.15


# ---- pricing.py mirror (verbatim, read-only) ----
def calculate_price_eur(retail_price_cents, discount_percent):
    discounted_cents = retail_price_cents * (100 - discount_percent) / 100
    tenths = math.floor(discounted_cents / 10 + 0.5)
    return tenths / 10.0


def resolve_discount_percent(match_discount, sup, brand, brand_rows):
    if match_discount is not None:
        return match_discount
    mode = (sup["pricing_mode"] or "flat")
    default = float(sup["discount_percent"] or 0.0)
    if mode == "per_brand" and brand:
        key = brand.strip().lower()
        if key:
            for row in brand_rows:
                if ((row["brand"] or "").strip().lower()) == key:
                    return float(row["discount_percent"])
        return default
    return default


def clamp_discount_for_min_margin(base_discount, retail_price_cents,
                                  eur_rate_uah, min_margin_uah, cost_rate=0.75):
    if retail_price_cents <= 0 or eur_rate_uah <= 0 or min_margin_uah <= 0:
        return int(base_discount)
    base_d = float(base_discount)
    retail_eur = retail_price_cents / 100.0
    max_margin_frac = 1 - cost_rate
    if retail_eur * max_margin_frac * eur_rate_uah < min_margin_uah:
        return 0
    base_frac = max_margin_frac - base_d / 100.0
    if base_frac * retail_eur * eur_rate_uah >= min_margin_uah:
        return int(math.floor(base_d))
    d_max = 100.0 * (max_margin_frac - min_margin_uah / (retail_eur * eur_rate_uah))
    d = int(math.floor(d_max))
    if d < 0:
        return 0
    if d > int(math.floor(base_d)):
        return int(math.floor(base_d))
    return d


def is_valid_price(price_cents):
    return price_cents is not None and price_cents > 0


def resolve_eur_rate(sup):
    raw = sup["eur_rate_uah"]
    rate = float(raw or 0)
    if rate > 0:
        return rate
    return _EUR_RATE_FALLBACK


# ---- matcher row (READ-ONLY) ----
con = sqlite3.connect(f"file:{DB}?mode=ro", uri=True)
con.row_factory = sqlite3.Row
pp = con.execute(
    "SELECT id, external_id, name, name_ru, brand FROM prom_products "
    "WHERE external_id=?", (EXTERNAL_ID,)).fetchone()
m = con.execute(
    "SELECT * FROM product_matches WHERE prom_product_id=? "
    "AND status IN ('confirmed','manual') AND published=1", (pp["id"],)).fetchone()
if m is None:
    raise SystemExit("No published confirmed/manual match")
sp = con.execute("SELECT * FROM supplier_products WHERE id=?",
                  (m["supplier_product_id"],)).fetchone()
sup = con.execute("SELECT * FROM suppliers WHERE id=?",
                  (sp["supplier_id"],)).fetchone()
brand_rows = con.execute(
    "SELECT brand, discount_percent FROM supplier_brand_discounts "
    "WHERE supplier_id=?", (sup["id"],)).fetchall()
con.close()

is_available = (is_valid_price(sp["price_cents"]) and bool(sp["available"])
                and not bool(sp["needs_review"]))
eff = resolve_discount_percent(m["discount_percent"], sup, sp["brand"], brand_rows)
if m["discount_percent"] is None and sup is not None:
    min_margin = float(sup["min_margin_uah"] or 0.0)
    if min_margin > 0:
        rate = resolve_eur_rate(sup)
        if (sp["currency"] or "EUR") == "UAH":
            rate = 1.0
        eff = clamp_discount_for_min_margin(
            eff, sp["price_cents"], rate, min_margin,
            float(sup["cost_rate"] or 0.75))
price_eur = calculate_price_eur(sp["price_cents"], eff) if is_valid_price(sp["price_cents"]) else 0.0
retail_eur = (sp["price_cents"] or 0) / 100.0
currency_id = (sp["currency"] or "EUR") if sp["currency"] in ("EUR", "UAH") else "EUR"

# ---- supplier native row (read-only) ----
_art = (sp["external_id"] or "").strip().lower()
wb = openpyxl.load_workbook(NP_FEED, read_only=True, data_only=True)
ws = wb["Worksheet"]
rows = ws.iter_rows(values_only=True)
HEADERS = list(next(rows))               # 24 НАТИВНЫХ заголовка Horoshop, дословно
src = None
for r in rows:
    if (str(r[1]) if r[1] is not None else "").strip().lower() == _art:
        src = list(r)
        break
wb.close()
if src is None:
    raise SystemExit(f"НП-фид: нет строки Артикул={_art!r}")

# col2 = «[КАТАЛОГ] Цена» -> вытащить точный префикс «[КАТАЛОГ] » без ручной кириллицы
_price_hdr = HEADERS[2]                                  # напр. "[КАТАЛОГ] Цена"
_prefix = _price_hdr[:_price_hdr.rfind(_price_hdr.split()[-1])]  # "[КАТАЛОГ] "
H_OLDPRICE = _prefix + "Старая цена"
H_CURRENCY = _prefix + "Валюта"

# ---- ЗАГОЛОВКИ В ФОРМЕ HOROSHOP АВТО-МАППИНГА (доказано live profile1 2026-05-19) ----
# Horoshop import авто-маппит колонку → поле каталога ТОЛЬКО когда заголовок ==
# полному квалифицированному имени поля "[КАТАЛОГ] <leaf>" (top-level поля — голым
# именем, напр. "Артикул"). На preview сели: Артикул, [КАТАЛОГ] Цена/Наличие/
# Отображать/Старая цена/Валюта. НЕ сели title_uk/description_uk/title_ru/
# description_ru/attr_brend — это ВНУТРЕННИЕ коды экспорта np.com.ua, а НЕ имена
# полей нашего каталога Horoshop → "Не імпортувати". Точные имена/группа сняты
# из vue-treeselect forest (id4/5/37/38/14/39/33, все под группой «КАТАЛОГ»):
#   id4  «Название модификации (UA)»  id5  «Название модификации (RU)»
#   id37 «Описание товара (UA)»       id38 «Описание товара (RU)»
#   id14 «Фото» (одиночное)           id39 «Галерея» (мультиизображения, наш url;url)
#   id33 «Бренд» (одно НЕлокализов. поле — таргетит ТОЛЬКО одна колонка, иначе
#        конфликт двух источников в одно поле; RU-дубль attr_brend оставляем
#        нераспознанным = "Не імпортувати", значение то же, безвредно).
# id (col0) сознательно оставляем нераспознанным — матч по Артикулу.
_HDR_OVERRIDE = {
    3:  _prefix + "Галерея",                       # был [КАТАЛОГ] Фото (url;url галерея)
    6:  _prefix + "Название модификации (UA)",      # был title_uk
    7:  _prefix + "Описание товара (UA)",           # был description_uk
    9:  _prefix + "Бренд",                          # был attr_brend_uk
    15: _prefix + "Название модификации (RU)",      # был title_ru
    16: _prefix + "Описание товара (RU)",           # был description_ru
}

# ---- собрать выходную строку в нативной схеме ----
out = list(src)                                          # копия supplier-значений
out[0] = ""                                              # id: пусто, матч по Артикулу
out[1] = str(pp["external_id"])                          # Артикул = НАША карточка
out[2] = f"{price_eur:.1f}"                               # [КАТАЛОГ] Цена = наша
out[4] = "В наличии" if is_available else "Нет в наличии"  # [КАТАЛОГ] Наличие
out[5] = "1"                                              # [КАТАЛОГ] Отображать
# тело/фото/имя/бренд (6,7,9,15,16,18 ...) — из supplier как есть.
# категории и характеристики НЕ трогаем (отдельный заход):
for i in (8, 10, 11, 12, 13, 14, 17, 19, 20, 21, 22, 23):
    out[i] = ""

_hdrs = list(HEADERS)
for _i, _h in _HDR_OVERRIDE.items():
    _hdrs[_i] = _h
out_headers = _hdrs + [H_OLDPRICE, H_CURRENCY]
out_row = out + [f"{retail_eur:.1f}", currency_id]

wbo = openpyxl.Workbook()
wso = wbo.active
wso.title = "Worksheet"
wso.append(out_headers)
wso.append(out_row)
for path in (r"C:/Users/Yana/labresta-np-feed-plan/canary-HKN-PICO12M.xlsx",
             r"C:/Users/Yana/labresta-np-feed-plan/_serve/canary-HKN-PICO12M.xlsx"):
    wbo.save(path)

print(f"OK match_id={m['id']} sp={sp['id']} pp={pp['id']} avail={is_available}")
print(f"  eff={eff}% retail={retail_eur:.2f} <Цена>={price_eur:.1f} "
      f"<Старая цена>={retail_eur:.1f} <Валюта>={currency_id}")
print(f"  Артикул(col1)={pp['external_id']}  prefix={_prefix!r}")
print(f"  headers({len(out_headers)}): {out_headers}")
print(f"  title_uk={src[6]!r}")
print(f"  title_ru={src[15]!r}")
