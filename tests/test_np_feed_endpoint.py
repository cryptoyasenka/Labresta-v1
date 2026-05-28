"""Endpoint tests for the NP Horoshop file UI (/feeds/np).

Verifies the wiring build_np_file's pure tests can't: form parsing, the
fetch→temp→build→download path (live fetch monkeypatched, no network), the
brand filter end-to-end, and that the downloaded file keys on the catalog
external_id. Seeds an NP supplier (slug=novyy-proekt) + one confirmed,
published match whose sp.article matches a row in the fixture feed.
"""

import io

import openpyxl

from app.models.catalog import PromProduct
from app.models.product_match import ProductMatch
from app.models.supplier import Supplier
from app.models.supplier_product import SupplierProduct


def _np_feed_bytes(article, *, brand="HURAKAN",
                   desc_ua="<p>опис</p>", desc_ru="<p>описание</p>",
                   photos="https://np/a.jpg;https://np/b.jpg"):
    """Build an NP-layout xlsx (header + one product row) as bytes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Worksheet"
    header = [None] * 24
    header[1] = "Артикул"
    header[2] = "[КАТАЛОГ] Цена"
    ws.append(header)
    row = [None] * 24
    row[1] = article
    row[3] = photos
    row[7] = desc_ua
    row[9] = brand
    row[16] = desc_ru
    ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


def _seed_np_match(session, *, article, brand="HURAKAN", external_id="999001",
                   price_cents=50000, feed_url="https://example.com/np.xlsx"):
    supplier = Supplier(
        name="Новый проект", slug="novyy-proekt", feed_url=feed_url,
        discount_percent=0, is_enabled=True,
    )
    session.add(supplier)
    session.flush()

    pp = PromProduct(
        external_id=external_id, name="Карточка", name_ru="Карточка",
        brand=brand, price=price_cents,
        page_url=f"https://labresta.com.ua/{external_id}/",
    )
    session.add(pp)
    session.flush()

    sp = SupplierProduct(
        supplier_id=supplier.id, external_id=f"ext-{external_id}",
        name="НП товар", brand=brand, article=article,
        price_cents=price_cents, available=True, needs_review=False,
    )
    session.add(sp)
    session.flush()

    match = ProductMatch(
        supplier_product_id=sp.id, prom_product_id=pp.id,
        score=100.0, status="confirmed", published=True, confirmed_by="test",
    )
    session.add(match)
    session.commit()
    return supplier, pp, sp, match


def test_np_page_renders_with_brands(client, session):
    _seed_np_match(session, article="ART-1", brand="HURAKAN")
    resp = client.get("/feeds/np")
    assert resp.status_code == 200
    body = resp.get_data(as_text=True)
    assert "Файл НП" in body
    assert "HURAKAN" in body
    assert "APACH" in body  # all 9 brand checkboxes present


def test_np_generate_downloads_native_xlsx(client, session, monkeypatch):
    _seed_np_match(session, article="ART-1", brand="HURAKAN", external_id="777")
    feed_bytes = _np_feed_bytes("ART-1", brand="HURAKAN")
    monkeypatch.setattr(
        "app.views.feed.fetch_feed_with_retry",
        lambda url, timeout=30: feed_bytes,
    )

    resp = client.post("/feeds/np/generate", data={"brands": ["HURAKAN"]})
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["Content-Type"]

    wb = openpyxl.load_workbook(io.BytesIO(resp.get_data()), read_only=True, data_only=True)
    ws = wb["Worksheet"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    assert rows[0][0] == "Артикул"
    assert len(rows) == 2  # header + one product
    data = rows[1]
    assert data[0] == "777"  # KEY = pp.external_id, NOT supplier article "ART-1"
    joined = " ".join(str(c) for c in data if c is not None)
    assert "np/a.jpg" in joined and "np/b.jpg" in joined  # gallery
    assert "описание" in joined                            # RU description


def test_np_generate_requires_brand_selection(client, session):
    _seed_np_match(session, article="ART-1")
    resp = client.post("/feeds/np/generate", data={})  # nothing ticked
    assert resp.status_code in (302, 303)  # redirect back with flash


def test_np_generate_excludes_unselected_brand(client, session, monkeypatch):
    # Product is APACH but operator ticks HURAKAN → header-only file, no rows.
    _seed_np_match(session, article="ART-2", brand="APACH", external_id="888")
    feed_bytes = _np_feed_bytes("ART-2", brand="APACH")
    monkeypatch.setattr(
        "app.views.feed.fetch_feed_with_retry",
        lambda url, timeout=30: feed_bytes,
    )

    resp = client.post("/feeds/np/generate", data={"brands": ["HURAKAN"]})
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(resp.get_data()), read_only=True, data_only=True)
    ws = wb["Worksheet"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    assert rows[0][0] == "Артикул"  # header present
    assert len(rows) == 1           # no data rows
