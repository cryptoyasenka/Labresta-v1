"""Tests for the НП Channel-2 catalog download endpoint (feed.download_np_catalog)."""

import io

import openpyxl

from app.models.catalog import PromProduct
from app.models.product_match import ProductMatch
from app.models.supplier import Supplier
from app.models.supplier_product import SupplierProduct

# Minimal native-feed header shape (same as the builder test).
_HEADERS = [
    "id", "Артикул", "[КАТАЛОГ] Цена", "[КАТАЛОГ] Фото", "[КАТАЛОГ] Наличие",
    "[КАТАЛОГ] Отображать", "title_uk", "description_uk", "categories_uk",
    "attr_brend_uk", "attr_contry_uk", "attr_dimensions_uk", "attr_power_uk",
    "attr_voltage_uk", "attr_weight_uk", "title_ru", "description_ru",
    "categories_ru", "attr_brend_ru", "attr_contry_ru", "attr_dimensions_ru",
    "attr_power_ru", "attr_voltage_ru", "attr_weight_ru",
]


def _make_feed(path, article, brand="HURAKAN"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Worksheet"
    ws.append(_HEADERS)
    r = [""] * 24
    r[1] = article
    r[3] = "https://np/a.jpg"
    r[6] = "Назва UA"
    r[7] = "Тіло UA"
    r[9] = brand
    r[15] = "Имя RU"
    r[16] = "Тело RU"
    ws.append(r)
    wb.save(path)
    wb.close()


def _np_match(session, *, article, external_id, brand="HURAKAN",
              status="confirmed", published=True):
    # The endpoint hardcodes supplier_id=2 (NP_SUPPLIER_ID); pin the PK to 2.
    if session.get(Supplier, 2) is None:
        session.add(Supplier(id=2, name="Новый проект", feed_url="http://np",
                             slug="novyy-proekt"))
        session.commit()
    sp = SupplierProduct(supplier_id=2, external_id=article, name="sp",
                         brand=brand, article=article, price_cents=100000,
                         currency="EUR", available=True)
    pp = PromProduct(external_id=external_id, name="pp", brand=brand)
    session.add_all([sp, pp])
    session.commit()
    session.add(ProductMatch(supplier_product_id=sp.id, prom_product_id=pp.id,
                             score=100.0, status=status, published=published))
    session.commit()


def test_download_serves_xlsx_attachment(client, session, app, tmp_path):
    feed = tmp_path / "np.xlsx"
    _make_feed(feed, "HKN-1")
    _np_match(session, article="HKN-1", external_id="EXT-9")
    app.config["NP_FEED_PATH"] = str(feed)

    resp = client.get("/feeds/np-catalog.xlsx")
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["Content-Type"]
    assert "attachment" in resp.headers["Content-Disposition"]
    assert "np-catalog-" in resp.headers["Content-Disposition"]

    wb = openpyxl.load_workbook(io.BytesIO(resp.data))
    ws = wb.active
    read = list(ws.iter_rows(values_only=True))
    wb.close()
    assert read[1][1] == "EXT-9"           # Артикул = our card external_id
    assert read[1][9] == "Имя RU"          # RU name landed in catalog column


def test_download_redirects_when_nothing_in_scope(client, session, app, tmp_path):
    feed = tmp_path / "np.xlsx"
    _make_feed(feed, "RC-1", brand="ROBOT COUPE")          # non-scope brand
    _np_match(session, article="RC-1", external_id="EXT-1", brand="ROBOT COUPE")
    app.config["NP_FEED_PATH"] = str(feed)

    resp = client.get("/feeds/np-catalog.xlsx")
    assert resp.status_code == 302
    assert "/feeds/custom" in resp.headers["Location"]


def test_download_redirects_when_feed_missing(client, app, tmp_path):
    app.config["NP_FEED_PATH"] = str(tmp_path / "does-not-exist.xlsx")
    resp = client.get("/feeds/np-catalog.xlsx")
    assert resp.status_code == 302
    assert "/feeds/custom" in resp.headers["Location"]


def test_download_requires_auth(app):
    # Unauthenticated client → login_required bounces (302/401), never 200.
    with app.test_client() as c:
        resp = c.get("/feeds/np-catalog.xlsx")
    assert resp.status_code in (302, 401)
