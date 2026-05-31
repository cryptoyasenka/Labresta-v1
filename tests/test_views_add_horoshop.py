"""Endpoint tests for the Horoshop create-file UI (/feeds/add).

Verifies the wiring the builder's pure tests can't: the supplier+brand picker,
the upload→temp→build→download path, exclusion of matched products, the
fallback category when no export is uploaded, and login_required on both routes.
Seeds one supplier with unmatched SupplierProducts (no ProductMatch) plus one
matched SP (confirmed) that must be excluded.
"""

import io

import openpyxl

from app.models.catalog import PromProduct
from app.models.product_match import ProductMatch
from app.models.supplier import Supplier
from app.models.supplier_product import SupplierProduct
from app.services.add_horoshop_file import (
    HEADERS, H_ARTICLE, H_CATEGORY, H_DESC_RU, H_NAME_RU, H_NAME_UA, H_VISIBLE,
)
from app.services.category_resolver import DEFAULT_FALLBACK_CATEGORY


def _export_bytes():
    """Tiny Horoshop-export xlsx (header + 2 data rows) for the upload field."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Worksheet"
    ws.append(["Артикул", "Название (UA)", "Бренд", "Раздел"])
    # The HURAKAN card shares the «Льодогенератор» token with the unmatched SP
    # ("Льодогенератор HKN") so the analogy tier can legitimately match it.
    ws.append(["EX-1", "Льодогенератор HKN-450", "HURAKAN", "Холодильне обладнання"])
    ws.append(["EX-2", "Піч конвекційна", "APACH", "Печі"])
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


def _np_feed_bytes(article, *, name_uk, name_ru, desc_uk, desc_ru, category):
    """Tiny NP-feed xlsx: header with title_uk/title_ru/categories_uk labels at
    their live indices (6/15/8) + one data row carrying RU text + category."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Worksheet"
    header = [None] * 24
    header[1] = "Артикул"
    header[2] = "[КАТАЛОГ] Цена"
    header[6] = "title_uk"
    header[7] = "description_uk"
    header[8] = "categories_uk"
    header[15] = "title_ru"
    header[16] = "description_ru"
    ws.append(header)
    row = [None] * 24
    row[1] = article
    row[6] = name_uk
    row[7] = desc_uk
    row[8] = category
    row[15] = name_ru
    row[16] = desc_ru
    ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


def _seed(session):
    """One supplier: 2 unmatched SPs (HURAKAN, APACH) + 1 matched SP (excluded)."""
    supplier = Supplier(
        name="Тест Постачальник", slug="test-supplier",
        discount_percent=15.0, pricing_mode="flat", is_enabled=True,
    )
    session.add(supplier)
    session.flush()

    sp_unmatched = SupplierProduct(
        supplier_id=supplier.id, external_id="ext-u1",
        name="Льодогенератор HKN", brand="HURAKAN", article="ART-UNMATCHED",
        price_cents=84500, currency="EUR", available=True, needs_review=False,
    )
    sp_other_brand = SupplierProduct(
        supplier_id=supplier.id, external_id="ext-u2",
        name="Піч APACH", brand="APACH", article="ART-APACH",
        price_cents=50000, currency="EUR", available=True, needs_review=False,
    )
    sp_matched = SupplierProduct(
        supplier_id=supplier.id, external_id="ext-m1",
        name="Матчений товар", brand="HURAKAN", article="ART-MATCHED",
        price_cents=60000, currency="EUR", available=True, needs_review=False,
    )
    session.add_all([sp_unmatched, sp_other_brand, sp_matched])
    session.flush()

    pp = PromProduct(
        external_id="999900", name="Картка", name_ru="Карточка",
        brand="HURAKAN", price=60000,
        page_url="https://labresta.com.ua/999900/",
    )
    session.add(pp)
    session.flush()

    match = ProductMatch(
        supplier_product_id=sp_matched.id, prom_product_id=pp.id,
        score=100.0, status="confirmed", published=True, confirmed_by="test",
    )
    session.add(match)
    session.commit()
    return supplier, sp_unmatched, sp_other_brand, sp_matched


def _load_rows(resp):
    wb = openpyxl.load_workbook(
        io.BytesIO(resp.get_data()), read_only=True, data_only=True
    )
    ws = wb["Worksheet"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return rows


def test_add_page_renders_supplier_picker(client, session):
    _seed(session)
    resp = client.get("/feeds/add")
    assert resp.status_code == 200
    body = resp.get_data(as_text=True)
    assert "Тест Постачальник" in body
    assert "Додати товари без матчу" in body


def test_add_page_lists_unmatched_brands(client, session):
    supplier, *_ = _seed(session)
    resp = client.get(f"/feeds/add?supplier_id={supplier.id}")
    assert resp.status_code == 200
    body = resp.get_data(as_text=True)
    # Brands of unmatched SPs appear.
    assert "HURAKAN" in body
    assert "APACH" in body
    # HURAKAN has exactly one UNMATCHED SP (the matched HURAKAN SP must not
    # inflate the badge to 2).
    assert ">2<" not in body.split("HURAKAN")[1].split("</li>")[0]


def test_generate_downloads_native_xlsx_with_category(client, session):
    supplier, *_ = _seed(session)
    resp = client.post(
        "/feeds/add/generate",
        data={
            "supplier_id": str(supplier.id),
            "brands": ["HURAKAN"],
            "export": (io.BytesIO(_export_bytes()), "export.xlsx"),
        },
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["Content-Type"]

    rows = _load_rows(resp)
    assert list(rows[0]) == HEADERS
    assert len(rows) == 2  # header + one HURAKAN unmatched product
    data = dict(zip(HEADERS, rows[1]))
    assert data[H_ARTICLE] == "ART-UNMATCHED"        # supplier article, not a pp id
    assert data[H_CATEGORY]                           # non-empty Раздел
    assert data[H_VISIBLE] == "1"


def test_generate_excludes_matched_products(client, session):
    supplier, *_ = _seed(session)
    resp = client.post(
        "/feeds/add/generate",
        data={"supplier_id": str(supplier.id)},  # no brands = all unmatched
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    rows = _load_rows(resp)
    articles = {dict(zip(HEADERS, r)).get(H_ARTICLE) for r in rows[1:]}
    assert "ART-MATCHED" not in articles
    assert "ART-UNMATCHED" in articles
    assert "ART-APACH" in articles


def test_generate_without_export_uses_fallback(client, session):
    supplier, *_ = _seed(session)
    resp = client.post(
        "/feeds/add/generate",
        data={"supplier_id": str(supplier.id), "brands": ["HURAKAN"]},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    rows = _load_rows(resp)
    assert len(rows) >= 2
    for r in rows[1:]:
        assert dict(zip(HEADERS, r))[H_CATEGORY] == DEFAULT_FALLBACK_CATEGORY


def test_generate_with_export_assigns_analogy_category(client, session):
    """With an export (HURAKAN card → «Холодильне обладнання») and NO feed, the
    HURAKAN unmatched SP gets that same-brand analog category — NOT the fallback.
    Proves the analogy tier overrides fallback when a corpus is present."""
    supplier, *_ = _seed(session)
    resp = client.post(
        "/feeds/add/generate",
        data={
            "supplier_id": str(supplier.id),
            "brands": ["HURAKAN"],
            "export": (io.BytesIO(_export_bytes()), "export.xlsx"),
        },
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    rows = _load_rows(resp)
    assert len(rows) == 2  # header + the single HURAKAN unmatched product
    data = dict(zip(HEADERS, rows[1]))
    # The only HURAKAN card in the export sits in «Холодильне обладнання»; the
    # analog must adopt it rather than the holding fallback category.
    assert data[H_CATEGORY] == "Холодильне обладнання"
    assert data[H_CATEGORY] != DEFAULT_FALLBACK_CATEGORY


def test_generate_np_feed_fills_ru_name_and_description(client, session):
    """FLAG-1/MINOR-B: an NP SupplierProduct carries NO RU (and no description) in
    the matcher DB. When an NP feed with title_ru/description_ru is uploaded, the
    created card's H_NAME_RU and H_DESC_RU must come out NON-EMPTY (decision D2)."""
    # Dedicated NP supplier: one unmatched SP with empty DB RU/description whose
    # article matches the feed row below.
    supplier = Supplier(
        name="Нова Пошта Фід", slug="np-feed-supplier",
        discount_percent=15.0, pricing_mode="flat", is_enabled=True,
    )
    session.add(supplier)
    session.flush()
    # NP SupplierProduct: the matcher DB has NO name_ru column at all and NP feeds
    # leave `description` empty — RU name + description live ONLY in the NP feed (D2).
    sp = SupplierProduct(
        supplier_id=supplier.id, external_id="np-ext-1",
        name="Шафа холодильна",          # UA name in DB; no RU, no description
        description=None,
        brand="POLAIR", article="NP-FRIDGE-1",
        price_cents=120000, currency="UAH", available=True, needs_review=False,
    )
    session.add(sp)
    session.commit()

    feed = _np_feed_bytes(
        "NP-FRIDGE-1",
        name_uk="Шафа холодильна POLAIR",
        name_ru="Шкаф холодильный POLAIR",
        desc_uk="Опис українською для шафи.",
        desc_ru="Описание по-русски для шкафа.",
        category="Холодильне обладнання/Шафи холодильні",
    )
    resp = client.post(
        "/feeds/add/generate",
        data={
            "supplier_id": str(supplier.id),
            "brands": ["POLAIR"],
            "np_feed": (io.BytesIO(feed), "np-feed.xlsx"),
        },
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    rows = _load_rows(resp)
    assert len(rows) == 2  # header + the single NP product
    data = dict(zip(HEADERS, rows[1]))
    assert data[H_NAME_RU] == "Шкаф холодильный POLAIR"     # RU name from the feed
    assert data[H_DESC_RU] == "Описание по-русски для шкафа."  # RU desc from the feed
    # UA name keeps the DB value (DB has it); the feed only backfills what's blank.
    assert data[H_NAME_UA]


def _seed_np_mapping_case(session):
    """NP supplier with one unmatched SP whose feed category does NOT reconcile
    to the store tree — the surface where Option B's mapping changes the «Раздел».
    Returns (supplier, article, feed_category, store_label)."""
    supplier = Supplier(
        name="НП Мапінг", slug="np-mapping-supplier",
        discount_percent=15.0, pricing_mode="flat", is_enabled=True,
    )
    session.add(supplier)
    session.flush()
    sp = SupplierProduct(
        supplier_id=supplier.id, external_id="np-map-ext-1",
        name="Шейкер для коктейлів", description=None,
        brand="HURAKAN", article="NP-SHAKER-1",
        price_cents=50000, currency="UAH", available=True, needs_review=False,
    )
    session.add(sp)
    session.commit()
    # Feed category uses a taxonomy the store export below does NOT carry, so it
    # cannot fuzzy-reconcile; only an explicit mapping can place it.
    return supplier, "NP-SHAKER-1", "Обладнання для барів/Шейкери ручні", "Барне обладнання/Шейкери"


def _mapping_export_bytes(store_label):
    """Tiny Horoshop export whose «Раздел» corpus contains `store_label` so a
    mapped feed value can exact-match it (HURAKAN brand for the analogy index)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Worksheet"
    ws.append(["Артикул", "Название (UA)", "Бренд", "Раздел"])
    ws.append(["EX-SH", "Шейкер бостонський", "HURAKAN", store_label])
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


def test_generate_no_category_mapping_is_baseline(client, session):
    """No category_mapping upload ≡ shipped default (Option A): the unreconcilable
    feed category does NOT become the store label — it falls back to the holding
    category (the analogy/fallback path), proving the default is unchanged."""
    supplier, article, feed_cat, store_label = _seed_np_mapping_case(session)
    export = _mapping_export_bytes(store_label)
    feed = _np_feed_bytes(
        article, name_uk="Шейкер UA", name_ru="Шейкер RU",
        desc_uk="опис", desc_ru="описание", category=feed_cat,
    )
    resp = client.post(
        "/feeds/add/generate",
        data={
            "supplier_id": str(supplier.id),
            "brands": ["HURAKAN"],
            "export": (io.BytesIO(export), "export.xlsx"),
            "np_feed": (io.BytesIO(feed), "np-feed.xlsx"),
            # NO category_mapping file → Option A.
        },
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    rows = _load_rows(resp)
    assert len(rows) == 2
    data = dict(zip(HEADERS, rows[1]))
    # Without the map, the feed label can't reconcile and there's no same-brand
    # analog token-match → the holding fallback category, NOT the store label.
    assert data[H_CATEGORY] != store_label
    assert data[H_CATEGORY] == DEFAULT_FALLBACK_CATEGORY


def test_generate_with_category_mapping_applies_store_label(client, session):
    """With a small category_mapping (.json) upload (Option B), the NP row whose
    feed category is mapped gets the mapped store «Раздел» — opt-in via the UI."""
    import json

    supplier, article, feed_cat, store_label = _seed_np_mapping_case(session)
    export = _mapping_export_bytes(store_label)
    feed = _np_feed_bytes(
        article, name_uk="Шейкер UA", name_ru="Шейкер RU",
        desc_uk="опис", desc_ru="описание", category=feed_cat,
    )
    mapping_json = json.dumps(
        {"_README": "test map", feed_cat: store_label}, ensure_ascii=False
    ).encode("utf-8")
    resp = client.post(
        "/feeds/add/generate",
        data={
            "supplier_id": str(supplier.id),
            "brands": ["HURAKAN"],
            "export": (io.BytesIO(export), "export.xlsx"),
            "np_feed": (io.BytesIO(feed), "np-feed.xlsx"),
            "category_mapping": (io.BytesIO(mapping_json), "map.json"),
        },
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    rows = _load_rows(resp)
    assert len(rows) == 2
    data = dict(zip(HEADERS, rows[1]))
    # The mapped feed label exact-matches the store set → the mapped «Раздел».
    assert data[H_CATEGORY] == store_label
    assert data[H_CATEGORY] != DEFAULT_FALLBACK_CATEGORY


def test_generate_unknown_supplier_redirects(client, session):
    _seed(session)
    resp = client.post(
        "/feeds/add/generate",
        data={"supplier_id": "999999"},
        content_type="multipart/form-data",
    )
    assert resp.status_code in (302, 303)


def test_routes_login_required(client, session):
    # The shared `client` fixture is authenticated; build a bare client.
    with client.application.test_client() as anon:
        get_resp = anon.get("/feeds/add", follow_redirects=False)
        assert get_resp.status_code in (301, 302)
        post_resp = anon.post(
            "/feeds/add/generate", data={"supplier_id": "1"},
            follow_redirects=False,
        )
        assert post_resp.status_code in (301, 302)
