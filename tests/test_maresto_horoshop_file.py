"""Tests for the MARESTO Horoshop matcher file builder (price + availability)."""

import io

import openpyxl

from app.services.maresto_horoshop_file import (
    H_ARTICLE,
    H_AVAIL,
    H_CURRENCY,
    H_OLDPRICE,
    H_PRICE,
    HEADERS,
    _shape_rows,
    _workbook_bytes,
)


def _m(**kw):
    base = {
        "external_id": "1",
        "stock_status": "In stock",
        "vendor": "Rational",
        "price_eur": 100.0,
        "retail_eur": 100.0,
        "currency": "EUR",
    }
    base.update(kw)
    return base


def test_shape_rows_maps_status_and_carries_price():
    matches = [
        _m(external_id="111", stock_status="In stock", vendor="Rational",
           price_eur=100.0, retail_eur=120.0),                       # discount
        _m(external_id="222", stock_status="Running low", vendor="Unox",
           price_eur=50.0, retail_eur=50.0),                         # no discount
        _m(external_id="333", stock_status="Reserved", vendor="Brema", price_eur=70.0),
        _m(external_id="444", stock_status="Out of stock", vendor="Forcar", price_eur=80.0),    # chinese
        _m(external_id="555", stock_status="Out of stock", vendor="Rational", price_eur=90.0),  # european
    ]
    rows, manifest = _shape_rows(matches)

    assert manifest["total"] == 5
    by_art = {r[H_ARTICLE]: r for r in rows}

    assert by_art["111"][H_AVAIL] == "В наявності"
    assert by_art["111"][H_PRICE] == "100.0"
    assert by_art["111"][H_OLDPRICE] == "120.0"        # real discount → oldprice present
    assert by_art["111"][H_CURRENCY] == "EUR"

    assert by_art["222"][H_AVAIL] == "В наявності"
    assert by_art["222"][H_OLDPRICE] == ""             # retail == price → no oldprice

    assert by_art["333"][H_AVAIL] == "Під замовлення"
    assert by_art["444"][H_AVAIL] == "Немає в наявності"   # chinese OOS
    assert by_art["555"][H_AVAIL] == "Під замовлення"      # european OOS

    assert manifest["per_status"] == {
        "В наявності": 2, "Під замовлення": 2, "Немає в наявності": 1,
    }


def test_shape_rows_skips_no_status_no_price_no_artikul():
    matches = [
        _m(external_id="1", stock_status=None),                 # no status
        _m(external_id="2", stock_status="In stock", price_eur=None),   # no price
        _m(external_id="3", stock_status="In stock", price_eur=0),      # zero price
        _m(external_id="", stock_status="In stock"),            # no artikul
    ]
    rows, manifest = _shape_rows(matches)
    assert rows == []
    assert manifest["skipped_no_status"] == 1
    assert manifest["skipped_no_price"] == 2
    assert manifest["skipped_no_artikul"] == 1
    assert manifest["total"] == 0


def test_workbook_roundtrips_headers_and_rows():
    rows, _ = _shape_rows([
        _m(external_id="111", stock_status="Out of stock", vendor="Rational",
           price_eur=90.0, retail_eur=110.0),
    ])
    data = _workbook_bytes(rows)

    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active
    assert [c.value for c in ws[1]] == HEADERS
    assert ws[2][0].value == "111"           # Артикул
    assert ws[2][1].value == "90.0"          # Цена
    assert ws[2][2].value == "110.0"         # Старая цена
    assert ws[2][3].value == "EUR"           # Валюта
    assert ws[2][4].value == "Під замовлення"  # Наличие
