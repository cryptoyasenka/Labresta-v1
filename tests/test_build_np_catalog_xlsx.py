"""Tests for scripts/build_np_catalog_xlsx — mass NP catalog import builder."""

import openpyxl
import pytest

from app.models.catalog import PromProduct
from app.models.product_match import ProductMatch
from app.models.supplier import Supplier
from app.models.supplier_product import SupplierProduct
from scripts.build_np_catalog_xlsx import (
    _gallery_str,
    build_catalog_rows,
    write_workbook,
)

# Reuse the feed-xlsx helper shape from the parser test.
HEADERS = [
    "id", "Артикул", "[КАТАЛОГ] Цена", "[КАТАЛОГ] Фото", "[КАТАЛОГ] Наличие",
    "[КАТАЛОГ] Отображать", "title_uk", "description_uk", "categories_uk",
    "attr_brend_uk", "attr_contry_uk", "attr_dimensions_uk", "attr_power_uk",
    "attr_voltage_uk", "attr_weight_uk", "title_ru", "description_ru",
    "categories_ru", "attr_brend_ru", "attr_contry_ru", "attr_dimensions_ru",
    "attr_power_ru", "attr_voltage_ru", "attr_weight_ru",
]


def _feed_row(article, title_uk, desc_uk, brand, title_ru, desc_ru, photo):
    r = [""] * 24
    r[1] = article
    r[3] = photo
    r[6] = title_uk
    r[7] = desc_uk
    r[9] = brand
    r[15] = title_ru
    r[16] = desc_ru
    return r


def _make_feed(path, data_rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Worksheet"
    ws.append(HEADERS)
    for r in data_rows:
        ws.append(r)
    wb.save(path)
    wb.close()


@pytest.fixture()
def np_supplier(session):
    s = Supplier(name="Новый проект", feed_url="http://np", slug="novyy-proekt",
                 discount_percent=0.0)
    session.add(s)
    session.commit()
    return s


def _matched(session, supplier_id, *, article, external_id, brand="HURAKAN",
             status="confirmed", published=True, price_cents=100000, available=True):
    sp = SupplierProduct(
        supplier_id=supplier_id, external_id=article, name="sp", brand=brand,
        article=article, price_cents=price_cents, currency="EUR",
        available=available,
    )
    pp = PromProduct(external_id=external_id, name="pp", brand=brand)
    session.add_all([sp, pp])
    session.commit()
    m = ProductMatch(supplier_product_id=sp.id, prom_product_id=pp.id,
                     score=100.0, status=status, published=published)
    session.add(m)
    session.commit()
    return sp, pp, m


class TestGalleryStr:
    def test_main_first_dedup(self):
        assert _gallery_str('["b.jpg","a.jpg"]', "a.jpg") == "a.jpg;b.jpg"

    def test_empty(self):
        assert _gallery_str(None, None) == ""


class TestBuildCatalogRows:
    def test_emits_row_with_content_from_feed(self, session, np_supplier, tmp_path):
        feed = tmp_path / "np.xlsx"
        _make_feed(feed, [
            _feed_row("HKN-1", "Назва UA", "Тіло UA", "HURAKAN",
                      "Имя RU", "Тело RU", "https://np/a.jpg;https://np/b.jpg"),
        ])
        _matched(session, np_supplier.id, article="HKN-1", external_id="EXT-9")

        headers, rows, errors = build_catalog_rows(str(feed), np_supplier.id)
        assert len(rows) == 1
        row = rows[0]
        # column order from _headers()
        assert row[0] == ""                       # id
        assert row[1] == "EXT-9"                  # Артикул = our card external_id
        assert row[4] == "В наличии"
        assert row[6] == "Назва UA"               # UA name
        assert row[8] == "HURAKAN"                # brand
        assert row[9] == "Имя RU"                 # RU name
        assert row[10] == "Тело RU"               # RU body
        assert row[3] == "https://np/a.jpg;https://np/b.jpg"  # gallery
        # native catalog headers present for auto-map
        assert any("Название модификации (RU)" in h for h in headers)
        assert any("Галерея" in h for h in headers)

    def test_skips_non_scope_brand(self, session, np_supplier, tmp_path):
        feed = tmp_path / "np.xlsx"
        _make_feed(feed, [
            _feed_row("RC-1", "u", "d", "ROBOT COUPE", "r", "dr", "x.jpg"),
        ])
        _matched(session, np_supplier.id, article="RC-1", external_id="EXT-1",
                 brand="ROBOT COUPE")
        _headers, rows, _errors = build_catalog_rows(str(feed), np_supplier.id)
        assert rows == []

    def test_skips_unpublished_and_candidate(self, session, np_supplier, tmp_path):
        feed = tmp_path / "np.xlsx"
        _make_feed(feed, [
            _feed_row("A-1", "u", "d", "APACH", "r", "dr", "x.jpg"),
            _feed_row("A-2", "u", "d", "APACH", "r", "dr", "x.jpg"),
        ])
        _matched(session, np_supplier.id, article="A-1", external_id="E1",
                 brand="APACH", published=False)
        _matched(session, np_supplier.id, article="A-2", external_id="E2",
                 brand="APACH", status="candidate")
        _headers, rows, _errors = build_catalog_rows(str(feed), np_supplier.id)
        assert rows == []

    def test_match_without_feed_row_warns(self, session, np_supplier, tmp_path):
        feed = tmp_path / "np.xlsx"
        _make_feed(feed, [
            _feed_row("OTHER", "u", "d", "FAGOR", "r", "dr", "x.jpg"),
        ])
        _matched(session, np_supplier.id, article="MISSING", external_id="E1",
                 brand="FAGOR")
        _headers, rows, errors = build_catalog_rows(str(feed), np_supplier.id)
        assert rows == []
        assert any("MISSING" in e for e in errors)

    def test_unavailable_when_no_stock(self, session, np_supplier, tmp_path):
        feed = tmp_path / "np.xlsx"
        _make_feed(feed, [
            _feed_row("T-1", "u", "d", "TATRA", "r", "dr", "x.jpg"),
        ])
        _matched(session, np_supplier.id, article="T-1", external_id="E1",
                 brand="TATRA", available=False)
        _headers, rows, _errors = build_catalog_rows(str(feed), np_supplier.id)
        assert rows[0][4] == "Нет в наличии"

    def test_write_workbook_roundtrip(self, session, np_supplier, tmp_path):
        feed = tmp_path / "np.xlsx"
        _make_feed(feed, [
            _feed_row("HKN-1", "Назва", "Тіло", "HURAKAN", "Имя", "Тело", "p.jpg"),
        ])
        _matched(session, np_supplier.id, article="HKN-1", external_id="EXT-9")
        headers, rows, _errors = build_catalog_rows(str(feed), np_supplier.id)
        out = tmp_path / "out.xlsx"
        write_workbook(headers, rows, str(out))

        wb = openpyxl.load_workbook(out)
        ws = wb.active
        read = list(ws.iter_rows(values_only=True))
        wb.close()
        assert list(read[0]) == headers
        assert read[1][1] == "EXT-9"
