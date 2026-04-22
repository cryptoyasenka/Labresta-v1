"""Tests for Excel parser service — URL detection, column mapping, product parsing."""

import logging

import openpyxl
import pytest

from app.services.excel_parser import (
    COLUMN_KEYWORDS,
    REQUIRED_FIELDS,
    convert_google_sheets_url,
    detect_columns,
    get_preview_data,
    is_google_sheets_url,
    parse_excel_products,
    validate_xlsx_response,
)


# ---------------------------------------------------------------------------
# Helpers: create temp .xlsx files for tests
# ---------------------------------------------------------------------------


def _create_xlsx(path, headers, rows):
    """Create a minimal .xlsx file with given headers and rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(str(path))
    wb.close()


def _create_xlsx_with_offset(path, blank_rows, headers, rows):
    """Create .xlsx where header is not in the first row (blank/title rows above)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    for _ in range(blank_rows):
        ws.append([])
    ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(str(path))
    wb.close()


# ===========================================================================
# URL Detection
# ===========================================================================


class TestIsGoogleSheetsUrl:
    def test_valid_google_sheets_url(self):
        url = "https://docs.google.com/spreadsheets/d/1-4UJcVAUefqV1NuSAGljghvzu-ulXf-A/edit?gid=1075932276#gid=1075932276"
        assert is_google_sheets_url(url) is True

    def test_invalid_non_google_url(self):
        assert is_google_sheets_url("https://example.com/feed.xml") is False

    def test_empty_string(self):
        assert is_google_sheets_url("") is False


# ===========================================================================
# URL Conversion
# ===========================================================================


class TestConvertGoogleSheetsUrl:
    def test_with_gid(self):
        url = "https://docs.google.com/spreadsheets/d/ABC123/edit?gid=456#gid=456"
        result = convert_google_sheets_url(url)
        assert result == "https://docs.google.com/spreadsheets/d/ABC123/export?format=xlsx&gid=456"

    def test_without_gid(self):
        url = "https://docs.google.com/spreadsheets/d/ABC123/edit"
        result = convert_google_sheets_url(url)
        assert result == "https://docs.google.com/spreadsheets/d/ABC123/export?format=xlsx"

    def test_invalid_url_raises_valueerror(self):
        with pytest.raises(ValueError):
            convert_google_sheets_url("https://example.com/feed.xml")


# ===========================================================================
# XLSX Validation
# ===========================================================================


class TestValidateXlsxResponse:
    def test_valid_pk_bytes(self):
        # PK magic bytes — valid .xlsx header
        valid_bytes = b"PK\x03\x04" + b"\x00" * 100
        # Should not raise
        validate_xlsx_response(valid_bytes)

    def test_html_response_raises(self):
        html = b"<!DOCTYPE html><html><body>Sign in</body></html>"
        with pytest.raises(ValueError, match="not shared publicly"):
            validate_xlsx_response(html)

    def test_invalid_bytes_raises(self):
        with pytest.raises(ValueError):
            validate_xlsx_response(b"\x00\x00\x00\x00")


# ===========================================================================
# Column Detection
# ===========================================================================


class TestDetectColumns:
    def test_russian_headers(self, tmp_path):
        xlsx = tmp_path / "ru.xlsx"
        _create_xlsx(xlsx, ["Название", "Бренд", "Модель", "Цена"], [])
        wb = openpyxl.load_workbook(str(xlsx), read_only=True)
        ws = wb[wb.sheetnames[0]]
        header_row, mapping = detect_columns(ws)
        wb.close()
        assert header_row == 0
        assert mapping[0] == "name"
        assert mapping[1] == "brand"
        assert mapping[2] == "model"
        assert mapping[3] == "price"

    def test_english_headers(self, tmp_path):
        xlsx = tmp_path / "en.xlsx"
        _create_xlsx(xlsx, ["#", "Name", "Brand", "Model", "Price", "Available"], [])
        wb = openpyxl.load_workbook(str(xlsx), read_only=True)
        ws = wb[wb.sheetnames[0]]
        header_row, mapping = detect_columns(ws)
        wb.close()
        assert header_row == 0
        assert mapping[1] == "name"
        assert mapping[2] == "brand"
        assert mapping[3] == "model"
        assert mapping[4] == "price"
        assert mapping[5] == "available"

    def test_ukrainian_headers(self, tmp_path):
        xlsx = tmp_path / "uk.xlsx"
        _create_xlsx(xlsx, ["Назва", "Бренд", "Модель", "Ціна", "Наявність"], [])
        wb = openpyxl.load_workbook(str(xlsx), read_only=True)
        ws = wb[wb.sheetnames[0]]
        header_row, mapping = detect_columns(ws)
        wb.close()
        assert header_row == 0
        assert mapping[0] == "name"
        assert mapping[1] == "brand"
        assert mapping[2] == "model"
        assert mapping[3] == "price"
        assert mapping[4] == "available"

    def test_no_match_returns_none(self, tmp_path):
        xlsx = tmp_path / "none.xlsx"
        _create_xlsx(xlsx, ["Foo", "Bar", "Baz"], [])
        wb = openpyxl.load_workbook(str(xlsx), read_only=True)
        ws = wb[wb.sheetnames[0]]
        header_row, mapping = detect_columns(ws)
        wb.close()
        assert header_row is None
        assert mapping == {}

    def test_header_not_first_row(self, tmp_path):
        """Header is in row 3 (index 2) — first rows are blank/title."""
        xlsx = tmp_path / "offset.xlsx"
        _create_xlsx_with_offset(
            xlsx,
            blank_rows=2,
            headers=["Название", "Бренд", "Модель", "Цена"],
            rows=[["Product1", "BrandA", "ModelX", 100]],
        )
        wb = openpyxl.load_workbook(str(xlsx), read_only=True)
        ws = wb[wb.sheetnames[0]]
        header_row, mapping = detect_columns(ws)
        wb.close()
        assert header_row == 2
        assert "name" in mapping.values()
        assert "brand" in mapping.values()


# ===========================================================================
# Product Parsing
# ===========================================================================


class TestParseExcelProducts:
    @pytest.fixture
    def basic_xlsx(self, tmp_path):
        """Create a basic .xlsx with header + 3 product rows."""
        xlsx = tmp_path / "products.xlsx"
        _create_xlsx(
            xlsx,
            ["Название", "Бренд", "Модель", "Цена"],
            [
                ["Pizza Oven Pro", "Apach", "APE-42", 15000],
                ["Mixer Industrial", "Sirman", "SIRPASTA", 8500],
                ["Grill Station", "Roller", "RG-520", 12000],
            ],
        )
        return str(xlsx)

    @pytest.fixture
    def basic_mapping(self):
        return {0: "name", 1: "brand", 2: "model", 3: "price"}

    def test_basic_parse(self, basic_xlsx, basic_mapping):
        products, errors = parse_excel_products(basic_xlsx, basic_mapping, header_row=0, supplier_id=1)
        assert len(products) == 3
        assert products[0]["name"] == "Pizza Oven Pro"
        assert products[0]["brand"] == "Apach"
        assert products[0]["model"] == "APE-42"
        assert products[0]["price_cents"] == 1500000
        assert products[0]["supplier_id"] == 1
        assert products[0]["currency"] == "EUR"

    def test_include_no_brand(self, tmp_path):
        """Rows without brand are included with name-based external_id."""
        xlsx = tmp_path / "no_brand.xlsx"
        _create_xlsx(
            xlsx,
            ["Название", "Бренд", "Модель", "Цена"],
            [
                ["Category Header", "", "ModelX", 100],
                ["Good Product", "Apach", "APE-42", 15000],
            ],
        )
        products, errors = parse_excel_products(
            str(xlsx), {0: "name", 1: "brand", 2: "model", 3: "price"}, header_row=0, supplier_id=1
        )
        assert len(products) == 2
        assert products[0]["brand"] == ""
        assert products[0]["external_id"] == "category header"
        assert products[1]["brand"] == "Apach"
        assert products[1]["external_id"] == "apach|ape-42"

    def test_include_no_model(self, tmp_path):
        """Rows without model are included with name-based external_id."""
        xlsx = tmp_path / "no_model.xlsx"
        _create_xlsx(
            xlsx,
            ["Название", "Бренд", "Модель", "Цена"],
            [
                ["Brand Divider", "Apach", "", 0],
                ["Good Product", "Apach", "APE-42", 15000],
            ],
        )
        products, errors = parse_excel_products(
            str(xlsx), {0: "name", 1: "brand", 2: "model", 3: "price"}, header_row=0, supplier_id=1
        )
        assert len(products) == 2
        assert products[0]["external_id"] == "brand divider"
        assert products[1]["external_id"] == "apach|ape-42"

    def test_skip_duplicate_brand_model(self, tmp_path, caplog):
        xlsx = tmp_path / "dup.xlsx"
        _create_xlsx(
            xlsx,
            ["Название", "Бренд", "Модель", "Цена"],
            [
                ["Oven Red", "Apach", "APE-42", 15000],
                ["Oven Blue", "Apach", "APE-42", 16000],
            ],
        )
        products, errors = parse_excel_products(
            str(xlsx), {0: "name", 1: "brand", 2: "model", 3: "price"}, header_row=0, supplier_id=1
        )
        assert len(products) == 1
        assert products[0]["name"] == "Oven Red"
        assert len(errors) >= 1
        assert "duplicate" in errors[0].lower()

    def test_unparseable_price(self, tmp_path):
        xlsx = tmp_path / "bad_price.xlsx"
        _create_xlsx(
            xlsx,
            ["Название", "Бренд", "Модель", "Цена"],
            [
                ["Oven", "Apach", "APE-42", "abc"],
            ],
        )
        products, errors = parse_excel_products(
            str(xlsx), {0: "name", 1: "brand", 2: "model", 3: "price"}, header_row=0, supplier_id=1
        )
        assert len(products) == 1
        assert products[0]["available"] is False
        assert products[0]["price_cents"] is None
        assert len(errors) >= 1

    def test_price_formats(self, tmp_path):
        """Comma decimal and space-thousands are parsed correctly."""
        xlsx = tmp_path / "prices.xlsx"
        _create_xlsx(
            xlsx,
            ["Название", "Бренд", "Модель", "Цена"],
            [
                ["Oven A", "Apach", "A1", "1 234.56"],
                ["Oven B", "Apach", "A2", "1234,56"],
            ],
        )
        products, errors = parse_excel_products(
            str(xlsx), {0: "name", 1: "brand", 2: "model", 3: "price"}, header_row=0, supplier_id=1
        )
        assert products[0]["price_cents"] == 123456
        assert products[1]["price_cents"] == 123456

    def test_default_available_true(self, tmp_path):
        """When no availability column mapped, all products are available=True."""
        xlsx = tmp_path / "no_avail.xlsx"
        _create_xlsx(
            xlsx,
            ["Название", "Бренд", "Модель", "Цена"],
            [
                ["Oven", "Apach", "APE-42", 15000],
            ],
        )
        products, errors = parse_excel_products(
            str(xlsx), {0: "name", 1: "brand", 2: "model", 3: "price"}, header_row=0, supplier_id=1
        )
        assert products[0]["available"] is True

    def test_external_id_format(self, tmp_path):
        xlsx = tmp_path / "ext_id.xlsx"
        _create_xlsx(
            xlsx,
            ["Название", "Бренд", "Модель", "Цена"],
            [
                ["Oven", "Apach", "APE-42", 15000],
            ],
        )
        products, errors = parse_excel_products(
            str(xlsx), {0: "name", 1: "brand", 2: "model", 3: "price"}, header_row=0, supplier_id=1
        )
        assert products[0]["external_id"] == "apach|ape-42"


# ===========================================================================
# Preview Data
# ===========================================================================


class TestGetPreviewData:
    def test_returns_expected_structure(self, tmp_path):
        xlsx = tmp_path / "preview.xlsx"
        _create_xlsx(
            xlsx,
            ["Название", "Бренд", "Модель", "Цена"],
            [
                ["Product1", "BrandA", "ModelX", 100],
                ["Product2", "BrandB", "ModelY", 200],
            ],
        )
        result = get_preview_data(str(xlsx))
        assert "header_row" in result
        assert "detected_mapping" in result
        assert "all_headers" in result
        assert "preview_rows" in result
        assert result["header_row"] == 0
        assert len(result["preview_rows"]) == 2
        assert len(result["all_headers"]) == 4


# ===========================================================================
# Constants
# ===========================================================================


class TestConstants:
    def test_column_keywords_has_required_fields(self):
        for field in REQUIRED_FIELDS:
            assert field in COLUMN_KEYWORDS, f"Missing keywords for required field: {field}"

    def test_required_fields(self):
        assert REQUIRED_FIELDS == {"name", "brand", "model", "price"}


# ===========================================================================
# Horoshop-style dealer exports (NP / НП: title_uk / attr_brend_uk / артикул)
# ===========================================================================


# Minimal subset of the 24-column NP dealer-export schema — enough to exercise
# detect_columns and parse_excel_products without dragging all 24 headers into
# every test. Order mirrors the real feed so column indices match.
NP_HEADERS = [
    "id",
    "артикул",
    "[оффера] цена",
    "[оффера] фото",
    "[оффера] наличие",
    "title_uk",
    "attr_brend_uk",
    "title_ru",
    "attr_brend_ru",
]


class TestHoroshopDealerExport:
    """Validate NP-style Horoshop dealer export ingestion end-to-end."""

    def _write_np_xlsx(self, path, rows):
        _create_xlsx(str(path), NP_HEADERS, rows)
        return str(path)

    def test_detect_columns_maps_title_brend_and_article(self, tmp_path):
        """title_uk → name, attr_brend_uk → brand, артикул → article, first-wins."""
        xlsx = self._write_np_xlsx(
            tmp_path / "np.xlsx",
            [[69773, "HKN-IMC25", 400.0, "img.jpg", "в наличии",
              "Льдогенератор Hurakan", "HURAKAN",
              "Льдогенератор Hurakan RU", "HURAKAN"]],
        )
        wb = openpyxl.load_workbook(xlsx, read_only=True)
        ws = wb[wb.sheetnames[0]]
        header_row, mapping = detect_columns(ws)
        wb.close()

        assert header_row == 0
        # ukr variants win (come first in the row) — ru columns stay unmapped.
        assert mapping == {
            1: "article",
            2: "price",
            4: "available",
            5: "name",
            6: "brand",
        }

    def test_article_becomes_external_id(self, tmp_path):
        """When article column is mapped, SKU — not brand+model — is external_id."""
        xlsx = self._write_np_xlsx(
            tmp_path / "np.xlsx",
            [
                [69773, "HKN-IMC25", 400.0, "img.jpg", "в наличии",
                 "Hurakan IMC25", "HURAKAN", "Hurakan IMC25 RU", "HURAKAN"],
                [69775, "HKN-IMC40", 505.0, "img.jpg", "в наличии",
                 "Hurakan IMC40", "HURAKAN", "Hurakan IMC40 RU", "HURAKAN"],
            ],
        )
        products, errors = parse_excel_products(
            xlsx,
            {1: "article", 2: "price", 4: "available", 5: "name", 6: "brand"},
            header_row=0,
            supplier_id=42,
        )
        assert errors == []
        assert len(products) == 2
        # external_id comes from the SKU column, not fabricated from brand+model
        assert products[0]["external_id"] == "hkn-imc25"
        assert products[0]["article"] == "HKN-IMC25"
        assert products[0]["brand"] == "HURAKAN"
        assert products[0]["name"] == "Hurakan IMC25"
        assert products[0]["price_cents"] == 40000  # 400.0 EUR → 40000 cents
        assert products[0]["available"] is True
        assert products[1]["external_id"] == "hkn-imc40"

    def test_duplicate_article_across_rows_is_skipped(self, tmp_path):
        """Two rows with the same SKU — second is dropped, error logged."""
        xlsx = self._write_np_xlsx(
            tmp_path / "np.xlsx",
            [
                [1, "HKN-IMC25", 400.0, "", "в наличии",
                 "Name one", "HURAKAN", "", "HURAKAN"],
                [2, "HKN-IMC25", 450.0, "", "в наличии",
                 "Name two", "HURAKAN", "", "HURAKAN"],
            ],
        )
        products, errors = parse_excel_products(
            xlsx,
            {1: "article", 2: "price", 4: "available", 5: "name", 6: "brand"},
            header_row=0,
            supplier_id=1,
        )
        assert len(products) == 1
        assert len(errors) == 1
        assert "HKN-IMC25" in errors[0]

    def test_fallback_external_id_when_no_article_column(self, tmp_path):
        """Without article mapping, external_id keeps old brand+model behavior."""
        xlsx = tmp_path / "no_article.xlsx"
        _create_xlsx(
            str(xlsx),
            ["Название", "Бренд", "Модель", "Цена"],
            [["Thing", "Apach", "APE-42", 15000]],
        )
        products, errors = parse_excel_products(
            str(xlsx),
            {0: "name", 1: "brand", 2: "model", 3: "price"},
            header_row=0,
            supplier_id=1,
        )
        assert products[0]["external_id"] == "apach|ape-42"
        assert products[0]["article"] is None
